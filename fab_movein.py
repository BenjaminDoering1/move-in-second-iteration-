#!/usr/bin/env python3
"""
Fab Move-In Simulator
=====================

Reads a fab layout (DXF) and a tool move-in schedule (Excel/CSV) that live on
YOUR machine, and writes a single self-contained HTML file that animates the
tools appearing on the layout, day by day.

Nothing is uploaded anywhere. The script runs locally; the HTML it produces has
the data baked in, so you just double-click it -- there is no file-upload step
and no internet connection required.

Quick start
-----------
    pip install ezdxf openpyxl
    python fab_movein.py --dxf layout.dxf --schedule movein.xlsx -o simulation.html

Not sure how your tools are identified in the drawing? Look first:

    python fab_movein.py --dxf layout.dxf --schedule movein.xlsx --inspect

That prints the block names, block attribute tags and text labels found in the
DXF alongside your spreadsheet columns, so you can see what to match on.

Move-in conflicts are detected and marked by default: two tools on the same
floor space, tools moving in within days of each other right next to each
other, a tool boxed in by neighbours that are already in place, and (with
--max-per-day N) days with too many move-ins. They are listed in the run
output, in the viewer's Conflicts tab, and with --conflicts-csv FILE.

Storage (laydown) space: give the zones with --storage NAME=X0,Y0,X1,Y1 or
--storage-layer NAME, and every tool whose delivery date (auto-detected
column, or --storage-lead DAYS before the move-in) precedes its move-in gets
a slot in the nearest zone that has room for the whole wait. The viewer
shows the crates in their slots until they move in; --storage-csv FILE
writes the plan.

Troubleshooting
---------------
* The page is slow / huge?  Real fab drawings carry a lot of geometry. This
  script already deduplicates repeated blocks, simplifies curves and caps the
  number of text labels; if the file is still big, run --inspect and drop the
  heavy layers you don't need with  --exclude-layer PIPING --exclude-layer "E-*"
* "[WinError 10053] An established connection was aborted..."  This script
  never opens a network connection -- that error comes from previewing the
  HTML through a local web server (VS Code Live Server, Jupyter,
  `python -m http.server`) which aborts on very large files. Open the HTML
  directly instead: double-click it in Explorer. It is fully self-contained.
* Tools all in one spot / not where they should be?  That happens when the
  tool blocks are proxy objects the DXF export did not include graphics for.
  The script now places such tools at their block insertion points and tells
  you how many were affected; for real footprints, re-export the DXF with
  proxy graphics (PROXYGRAPHICS=1) or explode the objects before SAVEAS.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import math
import os
import re
import sys
from collections import Counter, defaultdict

__version__ = "1.11.0"

_VIEWER_TEMPLATE = r'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Fab Move-In Simulator</title>
<style id="mainStyle">
/* ============ tokens ============ */
:root{
  color-scheme: light;
  --page:#f9f9f7; --surface:#fcfcfb; --surface-2:#f0efec;
  --ink:#0b0b0b; --ink-2:#52514e; --muted:#898781;
  --grid:#e1e0d9; --baseline:#c3c2b7; --hairline:rgba(11,11,11,.10);
  --s1:#2a78d6; --s2:#eb6834; --s3:#1baf7a; --s4:#eda100;
  --s5:#e87ba4; --s6:#008300; --s7:#4a3aa7; --s8:#e34948;
  --other:#898781; --good:#0ca30c;
  --map-ink:#a9a79e; --map-ink-strong:#6f6d66; --map-text:#898781;
  --accent:var(--s1);
  --shadow:0 1px 2px rgba(11,11,11,.06),0 4px 16px rgba(11,11,11,.06);
}
@media (prefers-color-scheme: dark){
  :root:where(:not([data-theme="light"])){
    color-scheme: dark;
    --page:#0d0d0d; --surface:#1a1a19; --surface-2:#242422;
    --ink:#ffffff; --ink-2:#c3c2b7; --muted:#898781;
    --grid:#2c2c2a; --baseline:#383835; --hairline:rgba(255,255,255,.10);
    --s1:#3987e5; --s2:#d95926; --s3:#199e70; --s4:#c98500;
    --s5:#d55181; --s6:#008300; --s7:#9085e9; --s8:#e66767;
    --map-ink:#4d4d49; --map-ink-strong:#75746e; --map-text:#6e6d67;
    --shadow:0 1px 2px rgba(0,0,0,.35),0 4px 16px rgba(0,0,0,.30);
  }
}
:root[data-theme="dark"]{
  color-scheme: dark;
  --page:#0d0d0d; --surface:#1a1a19; --surface-2:#242422;
  --ink:#ffffff; --ink-2:#c3c2b7; --muted:#898781;
  --grid:#2c2c2a; --baseline:#383835; --hairline:rgba(255,255,255,.10);
  --s1:#3987e5; --s2:#d95926; --s3:#199e70; --s4:#c98500;
  --s5:#d55181; --s6:#008300; --s7:#9085e9; --s8:#e66767;
  --map-ink:#4d4d49; --map-ink-strong:#75746e; --map-text:#6e6d67;
  --shadow:0 1px 2px rgba(0,0,0,.35),0 4px 16px rgba(0,0,0,.30);
}

*{box-sizing:border-box;margin:0;padding:0}
html,body{height:100%}
body{font:14px/1.45 system-ui,-apple-system,"Segoe UI",sans-serif;background:var(--page);color:var(--ink);
  display:flex;flex-direction:column;overflow:hidden}
button,select,input{font:inherit;color:inherit}
button{cursor:pointer;background:none;border:none}
h1{font-weight:600}
.num{font-variant-numeric:tabular-nums}

header{display:flex;align-items:center;gap:12px;padding:10px 16px;border-bottom:1px solid var(--hairline);
  background:var(--surface);flex:none}
header .logo{width:26px;height:26px;border-radius:7px;background:var(--accent);display:grid;place-items:center;
  color:#fff;font-weight:700;font-size:13px;flex:none}
header h1{font-size:15px}
header .sub{color:var(--muted);font-size:12px}
header .spacer{flex:1}
.warnpill{font-size:11.5px;color:var(--ink-2);background:var(--surface-2);border:1px solid var(--hairline);
  border-radius:999px;padding:4px 10px;cursor:help}

.btn{border:1px solid var(--hairline);background:var(--surface);border-radius:8px;padding:6px 12px;font-size:13px;
  color:var(--ink);display:inline-flex;align-items:center;gap:6px}
.btn:hover{background:var(--surface-2)}
.btn.icon{padding:6px 8px}

main{flex:1;min-height:0;display:flex;flex-direction:column;padding:12px;gap:12px}
.card{background:var(--surface);border:1px solid var(--hairline);border-radius:12px;box-shadow:var(--shadow)}

.controls{display:flex;align-items:center;gap:10px;flex-wrap:wrap;flex:none}
.controls .lbl{font-size:12px;color:var(--muted)}
select.sel,input.txt{background:var(--surface);border:1px solid var(--baseline);border-radius:8px;
  padding:7px 9px;font-size:13px}
select.sel{width:auto}
input.txt{width:170px}
#filters{display:flex;gap:6px;flex-wrap:wrap;align-items:center}
#filters select.sel{max-width:180px;padding:6px 8px;font-size:12.5px}
#filters select.sel.on{border-color:var(--accent);color:var(--accent)}
.clearf{font-size:12px;color:var(--muted);text-decoration:underline;cursor:pointer}
select.sel:focus,input.txt:focus{outline:2px solid var(--accent);outline-offset:-1px;border-color:transparent}
.legend{display:flex;gap:6px;flex-wrap:wrap;align-items:center}
.lg-chip{display:inline-flex;align-items:center;gap:6px;font-size:12px;color:var(--ink-2);
  border:1px solid var(--hairline);border-radius:999px;padding:3px 10px;background:var(--surface);cursor:pointer}
.lg-chip .sw{width:10px;height:10px;border-radius:3px;flex:none}
.lg-chip.off{opacity:.4}
.lg-chip .ct{color:var(--muted)}

.sim-body{flex:1;min-height:0;display:flex;gap:12px}
.map-card{flex:1;min-width:0;position:relative;overflow:hidden;display:flex}
#bgSvg,#mapSvg{position:absolute;inset:0;width:100%;height:100%;display:block}
/* own compositor layer: the heavy floor plan is rasterized once and cached,
   so tool animations never trigger a repaint of the background */
#bgSvg{pointer-events:none;will-change:transform;transform:translateZ(0);transform-origin:0 0}
#bgSvg.fast{shape-rendering:optimizeSpeed}
body.hide-fine .fine{display:none}
#mapSvg{touch-action:none;cursor:grab}
#mapSvg.panning{cursor:grabbing}
.map-tools-corner{position:absolute;top:10px;right:10px;display:flex;flex-direction:column;gap:6px}
.map-tools-corner .btn{box-shadow:var(--shadow)}
.map-title{position:absolute;top:10px;left:12px;font-size:12px;color:var(--muted);
  background:color-mix(in srgb,var(--surface) 82%,transparent);padding:3px 8px;border-radius:6px}

/* background strokes: width is set once per zoom on the root group (cheaper
   to rasterize than vector-effect:non-scaling-stroke on every element) */
.lyr{stroke:var(--map-ink);fill:none}
.lyr.strong{stroke:var(--map-ink-strong)}
.map-label{fill:var(--map-text);font-family:system-ui,sans-serif}
.tool-id{fill:var(--ink-2);font-family:system-ui,sans-serif;pointer-events:none}
g.tool{opacity:0;pointer-events:none}
g.tool.in{opacity:1;pointer-events:auto}
g.tool.pop .tool-shape{transform-box:fill-box;transform-origin:center;animation:pop .45s cubic-bezier(.2,1.6,.4,1) both}
g.tool.dim{opacity:.13}
g.tool.filt,#storeG g.stored.filt,#ghostG use.filt{display:none !important}
g.tool .halo{fill:none;stroke-width:2;vector-effect:non-scaling-stroke;opacity:0}
g.tool.today .halo,g.tool.flash .halo{opacity:1;animation:halo 1.6s ease-out infinite}
g.tool.in .hitbox{cursor:pointer}
g.tool .hitbox{fill:transparent;stroke:none}
#ghostG use{display:none;stroke:var(--map-ink);stroke-width:1;stroke-dasharray:3 3;fill:none;opacity:.55}
body.show-ghost #ghostG use.pending{display:block}
#ghostG use.smatch{display:block;stroke:var(--accent);opacity:.9}
body.hide-ids .tool-id{display:none}

/* conflicts: involved tools get a dashed warning outline, each conflict a marker */
g.tool.cf .tool-shape{stroke:var(--s8);stroke-width:2;stroke-dasharray:4 2}
g.tool.cf.sev-medium .tool-shape{stroke:var(--s2)}
g.tool.cf.sev-low .tool-shape{stroke:var(--s4)}
#cfG g.cfm{display:none;cursor:pointer}
#cfG g.cfm.in{display:block}
#cfG g.cfm path{fill:var(--s8);stroke:var(--surface);stroke-width:1.5;vector-effect:non-scaling-stroke}
#cfG g.cfm.sev-medium path{fill:var(--s2)}
#cfG g.cfm.sev-low path{fill:var(--s4)}
#cfG g.cfm text{fill:#fff;font-family:system-ui,sans-serif;font-weight:700;pointer-events:none}
#cfG g.cfm.flash path{animation:cfpulse .5s ease-in-out 4}
@keyframes cfpulse{50%{fill-opacity:.25}}
body.hide-cf #cfG{display:none}
/* storage zones and tools waiting in them */
#zoneG rect.zone{fill:var(--s3);fill-opacity:.07;stroke:var(--s3);stroke-width:1.2;stroke-dasharray:6 3;
  vector-effect:non-scaling-stroke;cursor:help}
#zoneG text{fill:var(--s3);font-family:system-ui,sans-serif;font-weight:600;pointer-events:none;opacity:.85}
#storeG g.stored{display:none;cursor:pointer}
#storeG g.stored.in{display:block}
#storeG g.stored use{fill:var(--s3);fill-opacity:.15;stroke:var(--s3);stroke-dasharray:2 2}
#storeG g.stored text{fill:var(--ink-2);font-family:system-ui,sans-serif;pointer-events:none}
#storeG g.stored.flash use{animation:cfpulse .5s ease-in-out 4}
body.hide-zones #zoneG,body.hide-zones #storeG{display:none}
.warnpill.cf-pill{color:var(--s8);border-color:color-mix(in srgb,var(--s8) 40%,transparent);cursor:pointer}
.tl-row .kind{font-size:10px;text-transform:uppercase;letter-spacing:.04em;padding:1px 6px;border-radius:4px;
  color:#fff;background:var(--s8);flex:none}
.tl-row .kind.medium{background:var(--s2)}
.tl-row .kind.low{background:var(--s4)}
.tl-row.cf-row{flex-wrap:wrap;row-gap:2px}
.tl-row.cf-row .nm{flex-basis:100%;white-space:normal;line-height:1.3}
.tl-row .st{color:var(--s3);font-size:11.5px;white-space:nowrap}
.tip .t-warn{color:var(--s8);font-size:12px;margin-top:3px}
.tip .t-store{color:var(--s3);font-size:12px;margin-top:2px}
body.wide .side{width:310px}
body.wide .tabs button{padding:9px 2px;font-size:12px}
@keyframes pop{0%{transform:scale(.4);opacity:0}100%{transform:scale(1);opacity:1}}
@keyframes halo{0%{stroke-opacity:.9;stroke-width:1.5}100%{stroke-opacity:0;stroke-width:7}}
@media (prefers-reduced-motion: reduce){
  g.tool.pop .tool-shape{animation:none}
  g.tool.today .halo,g.tool.flash .halo{animation:none;stroke-opacity:.7}
}

.side{width:280px;flex:none;display:flex;flex-direction:column;gap:12px;min-height:0}
.stat-tile{padding:14px 16px}
.stat-tile .lbl{font-size:12px;color:var(--ink-2)}
.stat-tile .val{font-size:34px;font-weight:650;line-height:1.15;margin-top:2px}
.stat-tile .val .of{font-size:16px;color:var(--muted);font-weight:500}
.stat-tile .delta{font-size:12.5px;margin-top:2px;color:var(--muted)}
.stat-tile .delta.pos{color:var(--good)}
.side .list-card{flex:1;min-height:0;display:flex;flex-direction:column}
.tabs{display:flex;border-bottom:1px solid var(--hairline);flex:none}
.tabs button{flex:1;padding:9px 4px;font-size:12.5px;color:var(--muted);border-bottom:2px solid transparent}
.tabs button.on{color:var(--ink);font-weight:600;border-bottom-color:var(--accent)}
.tool-list{flex:1;overflow:auto;padding:6px 0}
.tl-day{padding:7px 14px 3px;font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.04em;
  position:sticky;top:0;background:var(--surface)}
.tl-row{display:flex;align-items:center;gap:9px;padding:6px 14px;font-size:12.5px;cursor:pointer}
.tl-row:hover{background:var(--surface-2)}
.tl-row .sw{width:9px;height:9px;border-radius:3px;flex:none}
.tl-row .id{font-weight:600;white-space:nowrap}
.tl-row .nm{color:var(--muted);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex:1}
.tl-row .dt{color:var(--muted);font-size:11.5px;white-space:nowrap}
.tl-row.future{opacity:.6}
.tl-empty{padding:16px 14px;color:var(--muted);font-size:12.5px}

.playbar{flex:none;padding:10px 16px 12px;display:flex;flex-direction:column;gap:6px}
.pb-top{display:flex;align-items:center;gap:12px}
.pb-date{min-width:210px}
.pb-date .d1{font-size:19px;font-weight:650}
.pb-date .d2{font-size:12px;color:var(--muted)}
.pb-btns{display:flex;align-items:center;gap:6px}
.pb-btns .play{width:38px;height:38px;border-radius:50%;background:var(--accent);color:#fff;
  display:grid;place-items:center;border:none}
.pb-btns .play:hover{filter:brightness(1.07)}
.pb-btns .btn.icon{padding:7px 9px}
.pb-spacer{flex:1}
#chartWrap{position:relative;height:76px}
#chartWrap svg{position:absolute;inset:0;width:100%;height:100%;display:block}
#chartWrap .crosshair{position:absolute;top:0;bottom:14px;width:1px;background:var(--baseline);
  display:none;pointer-events:none}
input[type=range].scrub{width:100%;appearance:none;-webkit-appearance:none;height:18px;background:none;margin:0}
input[type=range].scrub::-webkit-slider-runnable-track{height:4px;border-radius:2px;
  background:linear-gradient(to right,var(--accent) 0 var(--pv,0%),var(--surface-2) var(--pv,0%) 100%)}
input[type=range].scrub::-webkit-slider-thumb{-webkit-appearance:none;width:16px;height:16px;border-radius:50%;
  background:var(--accent);border:2.5px solid var(--surface);box-shadow:var(--shadow);margin-top:-6px}
input[type=range].scrub::-moz-range-track{height:4px;border-radius:2px;background:var(--surface-2)}
input[type=range].scrub::-moz-range-progress{height:4px;border-radius:2px;background:var(--accent)}
input[type=range].scrub::-moz-range-thumb{width:12px;height:12px;border-radius:50%;background:var(--accent);
  border:2.5px solid var(--surface)}

.tip{position:fixed;z-index:50;pointer-events:none;background:var(--surface);border:1px solid var(--hairline);
  border-radius:9px;box-shadow:var(--shadow);padding:8px 11px;font-size:12.5px;display:none;max-width:260px}
.tip .t-val{font-size:14px;font-weight:650}
.tip .t-sub{color:var(--ink-2)}
.tip .t-mut{color:var(--muted);font-size:11.5px;margin-top:2px}
.tip .krow{display:flex;align-items:center;gap:7px;margin-top:3px}
.tip .krow .key{width:12px;height:3px;border-radius:2px;flex:none}

details.menu{position:relative}
details.menu>summary{list-style:none;cursor:pointer}
details.menu>summary::-webkit-details-marker{display:none}
details.menu .pop{position:absolute;right:0;top:calc(100% + 6px);z-index:40;background:var(--surface);
  border:1px solid var(--hairline);border-radius:10px;box-shadow:var(--shadow);padding:10px 12px;
  min-width:230px;max-height:60vh;overflow:auto}
.pop h4{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;margin:6px 0 4px}
.pop label.ck{display:flex;align-items:center;gap:8px;font-size:13px;padding:4px 0;cursor:pointer}
.pop label.ck input{accent-color:var(--accent)}
.pop .cnt{color:var(--muted);font-size:11.5px;margin-left:auto}
.pop .unmatched{font-size:12px;color:var(--ink-2);max-height:150px;overflow:auto}
.pop .unmatched .chip{display:inline-block;font-size:11px;background:var(--surface-2);border-radius:5px;
  padding:1px 6px;margin:2px 2px 0 0}

::-webkit-scrollbar{width:10px;height:10px}
::-webkit-scrollbar-thumb{background:var(--baseline);border-radius:5px;border:2px solid var(--surface)}
::-webkit-scrollbar-track{background:none}

</style>
</head>
<body>
<header>
  <div class="logo">▦</div>
  <div>
    <h1 id="hTitle">Fab Move-In Simulator</h1>
    <div class="sub" id="hSub"></div>
  </div>
  <div class="spacer"></div>
  <span class="warnpill cf-pill" id="cfPill" style="display:none" title="Potential move-in conflicts — click to list them"></span>
  <span class="warnpill" id="warnPill" style="display:none"></span>
  <button class="btn icon" id="pngBtn" title="Save current view as PNG">⤓ PNG</button>
  <button class="btn icon" id="themeBtn" title="Toggle light/dark">◐</button>
</header>

<main>
  <div class="controls">
    <span class="lbl">Color by</span>
    <select class="sel" id="colorBy"></select>
    <input class="txt" id="search" type="search" placeholder="Find tool… (Enter zooms)">
    <div id="filters"></div>
    <div class="legend" id="legend"></div>
    <div style="flex:1"></div>
    <details class="menu" id="viewMenu">
      <summary class="btn">View ▾</summary>
      <div class="pop">
        <h4>Options</h4>
        <label class="ck"><input type="checkbox" id="ckGhost"> Pending footprints (dashed)</label>
        <label class="ck"><input type="checkbox" id="ckText" checked> Drawing text</label>
        <label class="ck">Tool ID labels
          <select class="sel" id="labelMode" style="margin-left:auto;padding:3px 6px;font-size:12px">
            <option value="auto" selected>when zoomed in</option>
            <option value="on">always</option>
            <option value="off">never</option>
          </select>
        </label>
        <label class="ck">Fine drawing detail
          <select class="sel" id="detailMode" style="margin-left:auto;padding:3px 6px;font-size:12px">
            <option value="auto" selected>when zoomed in</option>
            <option value="on">always</option>
            <option value="off">never</option>
          </select>
        </label>
        <label class="ck"><input type="checkbox" id="ckFast"> Fast rendering (no anti-aliasing)</label>
        <label class="ck" id="lblCf" style="display:none"><input type="checkbox" id="ckCf" checked> Conflict markers</label>
        <label class="ck" id="lblZones" style="display:none"><input type="checkbox" id="ckZones" checked> Storage zones and waiting tools</label>
        <h4>Drawing layers</h4>
        <div id="layerList"></div>
        <div id="unmatchedBox" style="display:none">
          <h4 id="unmatchedH"></h4>
          <div class="unmatched" id="unmatchedList"></div>
        </div>
      </div>
    </details>
  </div>

  <div class="sim-body">
    <div class="card map-card">
      <svg id="bgSvg" xmlns="http://www.w3.org/2000/svg"></svg>
      <svg id="mapSvg" xmlns="http://www.w3.org/2000/svg"></svg>
      <div class="map-title" id="mapTitle"></div>
      <div class="map-tools-corner">
        <button class="btn icon" id="zoomIn" title="Zoom in">＋</button>
        <button class="btn icon" id="zoomOut" title="Zoom out">－</button>
        <button class="btn icon" id="zoomFit" title="Fit to view">⤢</button>
      </div>
    </div>

    <div class="side">
      <div class="card stat-tile">
        <div class="lbl">Tools moved in</div>
        <div class="val num"><span id="statN">0</span> <span class="of">/ <span id="statTotal">0</span></span></div>
        <div class="delta" id="statDelta"></div>
        <div class="delta" id="statExtra" style="display:none"></div>
      </div>
      <div class="card list-card">
        <div class="tabs">
          <button class="on" data-tab="day">This day</button>
          <button data-tab="next">Up next</button>
          <button data-tab="all">All tools</button>
          <button data-tab="cf" id="tabCf" style="display:none">Conflicts</button>
          <button data-tab="st" id="tabSt" style="display:none">Storage</button>
        </div>
        <div class="tool-list" id="toolList"></div>
      </div>
    </div>
  </div>

  <div class="card playbar">
    <div class="pb-top">
      <div class="pb-date">
        <div class="d1" id="dateBig">—</div>
        <div class="d2" id="dateSmall">—</div>
      </div>
      <div class="pb-btns">
        <button class="btn icon" id="stepBack" title="Previous day (←)">⏮</button>
        <button class="play" id="playBtn" title="Play / pause (space)"><span id="playIco">▶</span></button>
        <button class="btn icon" id="stepFwd" title="Next day (→)">⏭</button>
      </div>
      <span class="lbl">Speed</span>
      <select class="sel" id="speed">
        <option value="1">1 day/s</option>
        <option value="2" selected>2 days/s</option>
        <option value="4">4 days/s</option>
        <option value="8">8 days/s</option>
        <option value="16">16 days/s</option>
      </select>
      <div class="pb-spacer"></div>
      <span class="lbl num" id="dayCounter"></span>
    </div>
    <div id="chartWrap"><svg id="chartSvg"></svg><div class="crosshair" id="chartX"></div></div>
    <input type="range" class="scrub" id="scrub" min="0" max="0" value="0" step="1" aria-label="Simulation day">
  </div>
</main>

<div class="tip" id="tip"></div>

<script id="simData" type="application/json">/*__DATA__*/</script>
<script>(() => {
  const txt = document.getElementById('simData').textContent;
  const t0 = performance.now();
  window.DATA = JSON.parse(txt);
  window.DATA.__parseMs = performance.now() - t0;
  window.DATA.__jsonKB = txt.length / 1024;
})();</script>
<script>
'use strict';
/* Fab Move-In Simulator — viewer. Data is baked in by fab_movein.py as window.DATA. */

const $ = id => document.getElementById(id);
const el = (t, c, x) => { const n = document.createElement(t); if (c) n.className = c; if (x != null) n.textContent = x; return n; };
const svgEl = (t, a) => { const n = document.createElementNS('http://www.w3.org/2000/svg', t); if (a) for (const k in a) n.setAttribute(k, a[k]); return n; };
const norm = s => String(s == null ? '' : s).trim().toUpperCase();
const fmt = n => n.toLocaleString('en-US');
const MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
const WDAYS = ['Sun','Mon','Tue','Wed','Thu','Fri','Sat'];
const SLOTS = ['--s1','--s2','--s3','--s4','--s5','--s6','--s7','--s8'];
const REDUCED = matchMedia('(prefers-reduced-motion: reduce)').matches;

function kp(k){ const [y,m,d] = k.split('-').map(Number); return { y, m, d, wd: new Date(Date.UTC(y, m-1, d)).getUTCDay() }; }
const fmtLong  = k => { const p = kp(k); return `${MONTHS[p.m-1]} ${p.d}, ${p.y}`; };
const fmtShort = k => { const p = kp(k); return `${MONTHS[p.m-1]} ${p.d}`; };
const fmtWd    = k => WDAYS[kp(k).wd];

const D = window.DATA;
const PERF = { start: performance.now() };
const S = { idx: 0, prevK: -2, playing: false, speed: 2, colorField: null, colorMap: null,
            hidden: new Set(), search: '', layerVis: {}, tab: 'day', labelMode: 'auto',
            detailMode: 'auto', ready: false };

/* ---------- theme ---------- */
try { const t = localStorage.getItem('fabsim-theme'); if (t) document.documentElement.dataset.theme = t; } catch(e){}
$('themeBtn').addEventListener('click', () => {
  const r = document.documentElement;
  const dark = r.dataset.theme === 'dark' || (!r.dataset.theme && matchMedia('(prefers-color-scheme: dark)').matches);
  r.dataset.theme = dark ? 'light' : 'dark';
  try { localStorage.setItem('fabsim-theme', r.dataset.theme); } catch(e){}
  drawChart();
});

/* ---------- day index ----------
   Everything below is precomputed once so that stepping a day is O(changed
   tools), not O(all tools). Each tool id occurs on exactly one day (the
   generator dedupes), but one id may map to several drawn footprints.       */
const DAYS = D.days;
const DI = new Map(DAYS.map((k, i) => [k, i]));
D.tools.forEach(t => { t.di = DI.has(t.day) ? DI.get(t.day) : (t.day < DAYS[0] ? -1 : DAYS.length); });
D.tools.sort((a, b) => a.di - b.di || (a.id < b.id ? -1 : a.id > b.id ? 1 : 0));

const byDay = new Map(DAYS.map(k => [k, []]));
D.tools.forEach(t => { if (t.di >= 0 && t.di < DAYS.length) byDay.get(DAYS[t.di]).push(t); });

/* CUME[b] = number of tool entries with di+1 < b (entries sorted by di).    */
const CUME = [0];
{
  const bucket = new Array(DAYS.length + 2).fill(0);
  D.tools.forEach(t => bucket[Math.min(Math.max(t.di, -1), DAYS.length) + 1]++);
  bucket.forEach(n => CUME.push(CUME[CUME.length - 1] + n));
}
const idxUpTo = k => CUME[Math.min(Math.max(k + 2, 0), CUME.length - 1)];   // entries with di <= k

/* unique-id counts, over the tools passing the facet filters (recounted on
   every filter change -- O(tools) once, not per tick) */
const FIELDS = D.fields || [];
const newIds = new Array(DAYS.length).fill(0);
let NEWP = [0], TOTAL = 0, TOTAL_ALL = 0;
const passes = t => { for (const k in S.filter) if ((t[k] || '(blank)') !== S.filter[k]) return false; return true; };
function recount(){
  newIds.fill(0);
  let preIds = 0;
  const seen = new Set();
  D.tools.forEach(t => { if (seen.has(t.id) || !passes(t)) return; seen.add(t.id);
    if (t.di < 0) preIds++; else if (t.di < DAYS.length) newIds[t.di]++; });
  TOTAL = seen.size;
  NEWP = [preIds];
  newIds.forEach(n => NEWP.push(NEWP[NEWP.length - 1] + n));
}
S.filter = {};
recount(); TOTAL_ALL = TOTAL;
const cumAt = k => NEWP[Math.max(0, Math.min(k + 1, NEWP.length - 1))];

/* ---------- conflicts, storage, day events ----------
   Storage arrivals / departures and conflict markers are day EVENTS: sorted
   by day index with a prefix count, so stepping between two days applies
   only the events in between (forward) or undoes them (backward).          */
const dayIdx = k => DI.has(k) ? DI.get(k) : (k < DAYS[0] ? -1 : DAYS.length);
const CF = D.conflicts || [];
const ST = D.storage && D.storage.zones && D.storage.zones.length ? D.storage : null;
const CF_TITLE = { overlap: 'Footprints overlap', crowding: 'Crowded move-ins', access: 'Boxed in',
                   capacity: 'Too many move-ins in one day', storage: 'No storage space' };
CF.forEach((c, i) => { c.i = i; c.di = dayIdx(c.day); c.tools = []; });
const TOOLBYID = new Map();
D.tools.forEach(t => { if (!TOOLBYID.has(t.id)) TOOLBYID.set(t.id, t); });
CF.forEach(c => c.ids.forEach(id => { const t = TOOLBYID.get(id); if (t) c.tools.push(t); }));
/* one crate per tool id: the first placement carrying storage data */
const STORED = [];
D.tools.forEach(t => { if (t.ad && t.sz != null && TOOLBYID.get(t.id) === t) { t.adi = dayIdx(t.ad); STORED.push(t); } });
const CFCUM = new Array(DAYS.length + 1).fill(0);      // conflicts with di <= i
const INSTORE = new Array(DAYS.length + 1).fill(0);    // tools waiting in storage on day i
{
  const cc = new Array(DAYS.length + 2).fill(0), ss = new Array(DAYS.length + 2).fill(0);
  CF.forEach(c => cc[Math.min(Math.max(c.di, -1), DAYS.length) + 1]++);
  STORED.forEach(t => { ss[Math.min(Math.max(t.adi, -1), DAYS.length) + 1]++; ss[Math.min(Math.max(t.di, -1), DAYS.length) + 1]--; });
  let a = 0, b = 0;
  for (let i = 0; i <= DAYS.length; i++) { a += cc[i]; b += ss[i]; CFCUM[i] = a; INSTORE[i] = b; }
}
const cfAt = i => CFCUM[Math.max(0, Math.min(i + 1, DAYS.length))];
const storedAt = i => INSTORE[Math.max(0, Math.min(i + 1, DAYS.length))];
const inStorage = (t, i) => t.adi != null && t.adi <= i && i < t.di;
const EV = [];                                          // {di, fn(forward)}
let EVC = [0];
const evUpTo = k => EVC[Math.min(Math.max(k + 2, 0), EVC.length - 1)];
function indexEvents(){
  EV.sort((a, b) => a.di - b.di);
  const bucket = new Array(DAYS.length + 2).fill(0);
  EV.forEach(e => bucket[Math.min(Math.max(e.di, -1), DAYS.length) + 1]++);
  EVC = [0]; bucket.forEach(n => EVC.push(EVC[EVC.length - 1] + n));
}
const MED = (() => { const a = D.tools.map(t => Math.max(t.w, t.h)).sort((x, y) => x - y); return a.length ? a[a.length >> 1] : 5; })();

/* ---------- map ---------- */
const MAP = { vb: null, full: null, rendered: null, bgRoot: null, layerGroups: {},
              tiles: [], textBuckets: [], hasIds: false, medFs: 0 };

function addTextBuckets(tg, texts){
  if (!texts || !texts.length) return;
  const arr = texts.slice().sort((a, b) => b[2] - a[2]);
  const nb = Math.max(1, Math.min(6, Math.ceil(arr.length / 300)));
  const per = Math.ceil(arr.length / nb);
  for (let i = 0; i < arr.length; i += per) {
    const chunk = arr.slice(i, i + per);
    const g = svgEl('g');
    g.dataset.mh = chunk[Math.floor(chunk.length / 2)][2];
    chunk.forEach(t => {
      const [x, y, h, rot, str] = t;
      if (!str) return;
      const e = svgEl('text', { x, y, 'font-size': Math.max(h, 0.001), 'class': 'map-label' });
      if (rot) e.setAttribute('transform', `rotate(${-rot} ${x} ${y})`);
      e.textContent = str;
      g.appendChild(e);
    });
    MAP.textBuckets.push(g);
    tg.appendChild(g);
  }
}

function makeMapSkeleton(){
  /* The static background lives in its OWN <svg> so that playback (class
     toggles, halo pulses) only ever repaints the small foreground svg --
     never re-rasterizes the heavy floor-plan geometry. */
  const bgSvg = $('bgSvg'), svg = $('mapSvg');
  bgSvg.textContent = ''; svg.textContent = '';

  const bdefs = svgEl('defs');
  (D.bsyms || []).forEach((d, i) => bdefs.appendChild(svgEl('path', { id: 'b' + i, d })));
  bgSvg.appendChild(bdefs);
  const bg = svgEl('g');
  MAP.bgRoot = bg;
  D.layers.forEach(L => {
    const g = svgEl('g'); g.dataset.layer = L.n;
    bg.appendChild(g);
    MAP.layerGroups[L.n] = g;
    S.layerVis[L.n] = true;
  });
  bgSvg.appendChild(bg);

  const defs = svgEl('defs');
  D.syms.forEach((d, i) => defs.appendChild(svgEl('path', { id: 'sym' + i, d, 'vector-effect': 'non-scaling-stroke' })));
  svg.appendChild(defs);
  const root = svgEl('g');
  MAP.zoneG = svgEl('g', { id: 'zoneG' });
  root.appendChild(MAP.zoneG);
  MAP.ghostG = svgEl('g', { id: 'ghostG' });
  MAP.ghostG.setAttribute('pointer-events', 'none');
  root.appendChild(MAP.ghostG);
  MAP.storeG = svgEl('g', { id: 'storeG' });
  root.appendChild(MAP.storeG);
  MAP.toolsG = svgEl('g');
  root.appendChild(MAP.toolsG);
  MAP.cfG = svgEl('g', { id: 'cfG' });
  root.appendChild(MAP.cfG);
  svg.appendChild(root);

  MAP.storeG.addEventListener('pointerover', ev => {
    const g = ev.target.closest('g.stored'); if (!g) return;
    const t = TOOLBYID.get(g.dataset.id); if (t) showToolTip(t, ev, true);
  });
  MAP.storeG.addEventListener('pointermove', ev => { if ($('tip').style.display === 'block') positionTip(ev); });
  MAP.storeG.addEventListener('pointerout', ev => { if (ev.target.closest('g.stored')) hideTip(); });
  MAP.storeG.addEventListener('click', ev => { const g = ev.target.closest('g.stored'); if (g) zoomToTool(TOOLBYID.get(g.dataset.id)); });
  MAP.cfG.addEventListener('pointerover', ev => {
    const g = ev.target.closest('g.cfm'); if (!g) return;
    showConflictTip(CF[+g.dataset.i], ev);
  });
  MAP.cfG.addEventListener('pointermove', ev => { if ($('tip').style.display === 'block') positionTip(ev); });
  MAP.cfG.addEventListener('pointerout', ev => { if (ev.target.closest('g.cfm')) hideTip(); });
  MAP.cfG.addEventListener('click', ev => { const g = ev.target.closest('g.cfm'); if (g) focusConflict(CF[+g.dataset.i], false); });
  MAP.zoneG.addEventListener('pointerover', ev => {
    const r = ev.target.closest('rect.zone'); if (!r) return;
    showZoneTip(ST.zones[+r.dataset.z], ev);
  });
  MAP.zoneG.addEventListener('pointermove', ev => { if ($('tip').style.display === 'block') positionTip(ev); });
  MAP.zoneG.addEventListener('pointerout', ev => { if (ev.target.closest('rect.zone')) hideTip(); });

  /* delegated tool events: 2 listeners total instead of 3 per tool */
  MAP.toolsG.addEventListener('pointerover', ev => {
    const g = ev.target.closest('g.tool'); if (!g) return;
    const t = D.tools[+g.dataset.i];
    if (t && t.el && t.el.classList.contains('in')) showToolTip(t, ev);
  });
  MAP.toolsG.addEventListener('pointermove', ev => { if ($('tip').style.display === 'block') positionTip(ev); });
  MAP.toolsG.addEventListener('pointerout', ev => { if (ev.target.closest('g.tool')) hideTip(); });

  const b = D.bounds;
  const px = (b[2] - b[0]) * .04 || 1, py = (b[3] - b[1]) * .04 || 1;
  MAP.full = { x: b[0] - px, y: b[1] - py, w: (b[2] - b[0]) + px*2, h: (b[3] - b[1]) + py*2 };
  fitView();
  commitBg();                         // sets the initial background view + detail level
}

/* The background fills in over animation frames: giant path strings are split
   into ~1 MB pieces and <use> placements added in batches, so a heavy drawing
   paints progressively instead of freezing the tab. Every layer is stored as
   spatial tiles (hidden when off-screen) with a 'fine' sub-group of sub-pixel
   detail that is hidden until the zoom makes it visible. */
function backgroundTasks(){
  const tasks = [], SPLIT = 1 << 20;
  const pathTasks = (parent, d, cls) => {
    let off = 0;
    while (off < d.length) {
      let end = off + SPLIT >= d.length ? d.length : d.indexOf('M', off + SPLIT);
      if (end < 0) end = d.length;
      const piece = d.slice(off, end);
      tasks.push(() => parent.appendChild(svgEl('path', { d: piece, 'class': cls })));
      off = end;
    }
  };
  const useTasks = (parent, uses, cls) => {
    for (let i = 0; i < uses.length; i += 2500) {
      const part = uses.slice(i, i + 2500);
      tasks.push(() => part.forEach(u => parent.appendChild(svgEl('use', { href: '#b' + u[0], 'class': cls,
        transform: `matrix(${u[1]} ${u[2]} ${u[3]} ${u[4]} ${u[5]} ${u[6]})` }))));
    }
  };
  D.layers.forEach(L => {
    const g = MAP.layerGroups[L.n];
    const cls = 'lyr' + (L.strong ? ' strong' : '');
    (L.tiles || []).forEach(T => {
      const [ti, bb, dc, df, uc, uf] = T;
      const tg = svgEl('g'); tg.dataset.tile = ti;
      const fine = svgEl('g', { 'class': 'fine' });
      tg.appendChild(fine); g.appendChild(tg);
      MAP.tiles.push({ el: tg, bb });
      pathTasks(tg, dc, cls); useTasks(tg, uc, cls);
      pathTasks(fine, df, cls); useTasks(fine, uf, cls);
    });
    tasks.push(() => { const tg = svgEl('g', { 'class': 'bg-texts' }); addTextBuckets(tg, L.t); g.appendChild(tg); });
  });
  return tasks;
}

function buildMap(done){
  makeMapSkeleton();
  const tasks = backgroundTasks();
  let i = 0;
  (function step(){
    if (i >= tasks.length) { updateZoomVis(); done(); return; }
    const t0 = performance.now();
    while (i < tasks.length && performance.now() - t0 < 24) tasks[i++]();
    $('mapTitle').textContent = `drawing background… ${Math.round(i / tasks.length * 100)}%`;
    requestAnimationFrame(step);
  })();
}

const SHOW_IDS = D.tools.length <= 4000;
const fsList = [];
const worstSev = t => (t.cf || []).reduce((w, i) => { const s = CF[i].s; return (s === 'high' || w === 'high') ? 'high' : (s === 'medium' || w === 'medium') ? 'medium' : 'low'; }, '');
function addTool(t, i){
  const g = svgEl('g', { 'class': 'tool' + (t.cf ? ' cf sev-' + worstSev(t) : ''), 'data-i': i });
  const u = svgEl('use', { href: '#sym' + t.s, x: t.x, y: t.y, 'class': 'tool-shape' });
  g.appendChild(u);
  if (SHOW_IDS) {
    const fs = Math.max(Math.min(Math.min(t.w, t.h) * .32, 1.3 * t.w / Math.max(1, t.id.length)), .001);
    fsList.push(fs);
    const te = svgEl('text', { x: t.x + t.w/2, y: t.y + t.h/2, 'font-size': fs,
                               'text-anchor': 'middle', 'dominant-baseline': 'central', 'class': 'tool-id' });
    te.textContent = t.id;
    g.appendChild(te);
  }
  const pad = Math.max(t.w, t.h) * .15;
  g.appendChild(svgEl('rect', { 'class': 'hitbox', x: t.x - pad, y: t.y - pad,
                                width: t.w + pad*2, height: t.h + pad*2 }));
  t.el = g; t.use = u;
  return g;
}

/* tools are appended in animation-frame chunks so the floor plan paints first */
function buildToolsChunked(done){
  const per = 600; let i = 0;
  (function step(){
    const frag = document.createDocumentFragment();
    const end = Math.min(D.tools.length, i + per);
    for (; i < end; i++) frag.appendChild(addTool(D.tools[i], i));
    MAP.toolsG.appendChild(frag);
    if (i < D.tools.length) {
      $('mapTitle').textContent = `placing tools… ${Math.round(i / D.tools.length * 100)}%`;
      requestAnimationFrame(step);
    } else {
      fsList.sort((a, b) => a - b);
      MAP.medFs = fsList.length ? fsList[fsList.length >> 1] : 0;
      MAP.hasIds = SHOW_IDS;
      if (!SHOW_IDS) document.body.classList.add('hide-ids');
      done();
    }
  })();
}

/* storage zones, tools waiting in their slots, conflict markers */
function buildExtras(){
  if (ST) {
    ST.zones.forEach((z, zi) => {
      MAP.zoneG.appendChild(svgEl('rect', { 'class': 'zone', x: z.x, y: z.y, width: z.w, height: z.h, 'data-z': zi }));
      const fs = Math.max(Math.min(z.h * .16, z.w * 1.6 / Math.max(4, z.n.length)), .001);
      const te = svgEl('text', { x: z.x + z.w * .03, y: z.y + fs * 1.1, 'font-size': fs });
      te.textContent = z.n; MAP.zoneG.appendChild(te);
    });
    const frag = document.createDocumentFragment();
    STORED.forEach(t => {
      const g = svgEl('g', { 'class': 'stored', 'data-id': t.id });
      g.appendChild(svgEl('use', { href: '#sym' + t.s, x: t.sx, y: t.sy, 'class': 'tool-shape' }));
      if (SHOW_IDS) {
        const fs = Math.max(Math.min(Math.min(t.w, t.h) * .32, 1.3 * t.w / Math.max(1, t.id.length)), .001);
        const te = svgEl('text', { x: t.sx + t.w/2, y: t.sy + t.h/2, 'font-size': fs, 'text-anchor': 'middle',
                                   'dominant-baseline': 'central', 'class': 'tool-id' });
        te.textContent = t.id; g.appendChild(te);
      }
      t.stEl = g; frag.appendChild(g);
      EV.push({ di: t.adi, fn: f => g.classList.toggle('in', f) });
      EV.push({ di: t.di,  fn: f => g.classList.toggle('in', !f) });
    });
    MAP.storeG.appendChild(frag);
  }
  if (CF.length) {
    const r = MED * .38, frag = document.createDocumentFragment();
    CF.forEach(c => {
      if (c.x == null) return;
      const g = svgEl('g', { 'class': 'cfm sev-' + c.s, 'data-i': c.i });
      g.appendChild(svgEl('path', { d: `M${c.x} ${c.y - r}L${c.x + r} ${c.y}L${c.x} ${c.y + r}L${c.x - r} ${c.y}Z` }));
      const te = svgEl('text', { x: c.x, y: c.y, 'font-size': r * 1.2, 'text-anchor': 'middle', 'dominant-baseline': 'central' });
      te.textContent = '!'; g.appendChild(te);
      c.el = g; frag.appendChild(g);
      EV.push({ di: c.di, fn: f => g.classList.toggle('in', f) });
    });
    MAP.cfG.appendChild(frag);
  }
  indexEvents();
}

function ensureHalo(t){
  if (t.haloEl || !t.el) return;
  const r = svgEl('rect', { 'class': 'halo', x: t.x - t.w*.08, y: t.y - t.h*.08,
                            width: t.w*1.16, height: t.h*1.16, rx: Math.min(t.w, t.h)*.1 });
  r.setAttribute('stroke', toolColor(t));
  t.el.insertBefore(r, t.el.firstChild);
  t.haloEl = r;
}

let ghostsBuilt = false;
function buildGhosts(){
  if (ghostsBuilt) return; ghostsBuilt = true;
  const frag = document.createDocumentFragment();
  D.tools.forEach(t => {
    const u = svgEl('use', { href: '#sym' + t.s, x: t.x, y: t.y });
    if (t.di > S.idx) u.classList.add('pending');
    t.ghost = u; frag.appendChild(u);
  });
  MAP.ghostG.appendChild(frag);
}

/* viewBox writes are collapsed to one per animation frame. The foreground
   (tools) re-renders live. The heavy background is a cached bitmap: during a
   pan/zoom gesture it is only scaled and shifted (compositor-only, instant);
   once the gesture settles it is re-rendered sharp at the new view. */
let vbRaf = 0, commitT = 0;
function svgMetrics(v){
  const r = $('mapSvg').getBoundingClientRect();
  const s = Math.min(r.width / v.w, r.height / v.h);
  return { s, ox: (r.width - v.w * s) / 2, oy: (r.height - v.h * s) / 2 };
}
function setVB(vb){
  MAP.vb = vb;
  if (!vbRaf) vbRaf = requestAnimationFrame(() => {
    vbRaf = 0;
    const v = MAP.vb;
    $('mapSvg').setAttribute('viewBox', `${v.x} ${v.y} ${v.w} ${v.h}`);
    updateFgZoom();
    const m0 = MAP.rendered && svgMetrics(MAP.rendered), m1 = svgMetrics(v);
    const a = m0 ? m1.s / m0.s : NaN;
    if (isFinite(a) && a > 0) {
      const bx = m1.ox + (MAP.rendered.x - v.x) * m1.s - a * m0.ox;
      const by = m1.oy + (MAP.rendered.y - v.y) * m1.s - a * m0.oy;
      $('bgSvg').style.transform = `translate(${bx}px, ${by}px) scale(${a})`;
      clearTimeout(commitT); commitT = setTimeout(commitBg, 160);
    } else commitBg();
  });
}
function commitBg(){
  clearTimeout(commitT); commitT = 0;
  const v = MAP.vb, bg = $('bgSvg');
  bg.style.transform = '';
  bg.setAttribute('viewBox', `${v.x} ${v.y} ${v.w} ${v.h}`);
  MAP.rendered = { ...v };
  updateBgZoom();
}
function pxPerUnit(){
  const r = $('mapSvg').getBoundingClientRect();
  return (!r.width || !MAP.vb) ? 0 : Math.min(r.width / MAP.vb.w, r.height / MAP.vb.h);
}
/* tool ID labels appear once they are actually readable at this zoom */
function updateFgZoom(){
  if (!MAP.hasIds) return;
  const ppu = pxPerUnit(); if (!ppu) return;
  const readable = S.labelMode !== 'off' && (S.labelMode === 'on' || MAP.medFs * ppu >= 5.5);
  document.body.classList.toggle('hide-ids', !readable);
}
/* background detail: off-screen tiles, sub-pixel detail, small texts are hidden */
function updateBgZoom(){
  const ppu = pxPerUnit(); if (!ppu) return;
  const v = MAP.vb;
  const fineOn = S.detailMode === 'on' || (S.detailMode === 'auto' && ppu * (D.fineUnit || 0) >= 2.5);
  document.body.classList.toggle('hide-fine', !fineOn);
  if (MAP.bgRoot) MAP.bgRoot.setAttribute('stroke-width', 1 / ppu);
  const mx = v.w * .35, my = v.h * .35;
  MAP.tiles.forEach(t => {
    const b = t.bb;
    const vis = !(b[2] < v.x - mx || b[0] > v.x + v.w + mx || b[3] < v.y - my || b[1] > v.y + v.h + my);
    if (vis !== (t.el.style.display !== 'none')) t.el.style.display = vis ? '' : 'none';
  });
  const showT = $('ckText').checked;
  MAP.textBuckets.forEach(g => {
    const vis = showT && (+g.dataset.mh) * ppu >= 3.2;
    const cur = g.style.display !== 'none';
    if (vis !== cur) g.style.display = vis ? '' : 'none';
  });
}
function updateZoomVis(){ updateFgZoom(); updateBgZoom(); }
function fitView(){ setVB({ ...MAP.full }); }
function zoomAt(f, cx, cy){ const v = MAP.vb; setVB({ x: cx - (cx - v.x)*f, y: cy - (cy - v.y)*f, w: v.w*f, h: v.h*f }); }
function clientToSvg(ev){
  const r = $('mapSvg').getBoundingClientRect(), v = MAP.vb;
  const s = Math.min(r.width / v.w, r.height / v.h);
  return [ v.x + (ev.clientX - r.left - (r.width - v.w*s)/2) / s, v.y + (ev.clientY - r.top - (r.height - v.h*s)/2) / s ];
}
function setupMap(){
  const svg = $('mapSvg');
  svg.addEventListener('wheel', ev => { ev.preventDefault(); const [x, y] = clientToSvg(ev); zoomAt(Math.pow(1.0018, ev.deltaY), x, y); }, { passive: false });
  let pan = null;
  svg.addEventListener('pointerdown', ev => {
    if (ev.button !== 0) return;
    pan = { x: ev.clientX, y: ev.clientY, vb: { ...MAP.vb } };
    svg.classList.add('panning'); svg.setPointerCapture(ev.pointerId);
  });
  svg.addEventListener('pointermove', ev => {
    if (!pan) return;
    const r = svg.getBoundingClientRect(), s = Math.min(r.width / pan.vb.w, r.height / pan.vb.h);
    setVB({ x: pan.vb.x - (ev.clientX - pan.x)/s, y: pan.vb.y - (ev.clientY - pan.y)/s, w: pan.vb.w, h: pan.vb.h });
  });
  const end = () => { pan = null; svg.classList.remove('panning'); };
  svg.addEventListener('pointerup', end); svg.addEventListener('pointercancel', end);
  $('zoomIn').onclick  = () => zoomAt(.7,  MAP.vb.x + MAP.vb.w/2, MAP.vb.y + MAP.vb.h/2);
  $('zoomOut').onclick = () => zoomAt(1.45, MAP.vb.x + MAP.vb.w/2, MAP.vb.y + MAP.vb.h/2);
  $('zoomFit').onclick = fitView;
  new ResizeObserver(() => { if (MAP.vb) { updateFgZoom(); commitBg(); } }).observe(svg.parentElement);
}
function zoomToBox(x, y, w, h, f){
  const size = Math.max(w, h) * (f || 14);
  setVB({ x: x + w/2 - size/2, y: y + h/2 - size/2, w: size, h: size });
}
function flashTool(t){
  if (!t.el) return;
  ensureHalo(t); t.el.classList.add('flash');
  setTimeout(() => t.el.classList.remove('flash'), 2400);
}
function focusConflict(c, jump){
  if (jump && c.di > S.idx && c.di < DAYS.length) { setPlaying(false); setDay(c.di); }
  const ts = c.tools.filter(t => t.el);
  if (ts.length) {
    const x0 = Math.min(...ts.map(t => t.x)), y0 = Math.min(...ts.map(t => t.y));
    const x1 = Math.max(...ts.map(t => t.x + t.w)), y1 = Math.max(...ts.map(t => t.y + t.h));
    zoomToBox(x0, y0, x1 - x0, y1 - y0, c.k === 'capacity' ? 1.3 : 6);
    ts.forEach(flashTool);
    if (c.k === 'storage' && ts[0].stEl) ts[0].stEl.classList.add('flash');
    setTimeout(() => ts.forEach(t => t.stEl && t.stEl.classList.remove('flash')), 2400);
  }
  if (c.el) { c.el.classList.add('flash'); setTimeout(() => c.el.classList.remove('flash'), 2400); }
}
function zoomToTool(t){
  if (!t.el) return;
  if (inStorage(t, S.idx) && t.stEl) { zoomToBox(t.sx, t.sy, t.w, t.h); t.stEl.classList.add('flash');
    setTimeout(() => t.stEl.classList.remove('flash'), 2400); return; }
  const size = Math.max(t.w, t.h) * 14;
  setVB({ x: t.x + t.w/2 - size/2, y: t.y + t.h/2 - size/2, w: size, h: size });
  ensureHalo(t);
  t.el.classList.add('flash');
  if (!t.el.classList.contains('in')) {          // not moved in yet: reveal its outline briefly
    buildGhosts();
    t.ghost.classList.add('smatch');
    setTimeout(() => { if (t.ghost && !isSearchMatch(t)) t.ghost.classList.remove('smatch'); }, 2600);
  }
  setTimeout(() => t.el.classList.remove('flash'), 2400);
}

/* ---------- colors ---------- */
function buildColorControls(){
  const sel = $('colorBy'); sel.textContent = '';
  const opts = [{ k: '', l: 'Single color' }, ...FIELDS];
  opts.forEach(o => { const e = el('option', null, o.l); e.value = o.k; sel.appendChild(e); });
  sel.value = FIELDS.some(f => f.k === D.colorBy) ? D.colorBy : '';
  S.colorField = sel.value || null;
  sel.onchange = () => { S.colorField = sel.value || null; S.hidden.clear(); applyColors(); renderList(); };
}
/* one dropdown per schedule field with a manageable number of values; the
   filters combine (AND) and hide the tools that do not pass, so the counts,
   the chart and the lists describe the selection */
function buildFilters(){
  const box = $('filters'); box.textContent = '';
  const uniq = dedupeById(D.tools);
  FIELDS.forEach(f => {
    const counts = new Map();
    uniq.forEach(t => { const v = t[f.k] || '(blank)'; counts.set(v, (counts.get(v) || 0) + 1); });
    if (counts.size < 2 || counts.size > 400) return;
    const sel = el('select', 'sel'); sel.title = `Filter by ${f.l}`;
    const all = el('option', null, `All ${f.l}`); all.value = ''; sel.appendChild(all);
    [...counts.keys()].sort((a, b) => String(a).localeCompare(String(b), undefined, { numeric: true })).forEach(v => {
      const o = el('option', null, `${v} (${counts.get(v)})`); o.value = v; sel.appendChild(o);
    });
    sel.onchange = () => {
      if (sel.value) S.filter[f.k] = sel.value; else delete S.filter[f.k];
      sel.classList.toggle('on', !!sel.value);
      applyFilter();
    };
    box.appendChild(sel);
  });
  if (box.children.length) {
    const clr = el('span', 'clearf', 'clear');
    clr.onclick = () => { S.filter = {}; box.querySelectorAll('select').forEach(x => { x.value = ''; x.classList.remove('on'); }); applyFilter(); };
    box.appendChild(clr);
  }
}
function applyFilter(){
  D.tools.forEach(t => {
    const off = !passes(t);
    if (t.el) t.el.classList.toggle('filt', off);
    if (t.stEl) t.stEl.classList.toggle('filt', off);
    if (t.ghost) t.ghost.classList.toggle('filt', off);
  });
  recount();
  allBuilt = false; allBuilding = false; ROWBYID.clear();   // 'All tools' is rebuilt lazily
  renderLegend(); drawChart(); setDay(S.idx);
  $('mapTitle').textContent = Object.keys(S.filter).length
    ? `${fmt(TOTAL)} of ${fmt(TOTAL_ALL)} tools match the filter · ${fmt(DAYS.length)} days`
    : `${fmt(TOTAL_ALL)} tools · ${fmt(DAYS.length)} days`;
}
function computeColorMap(){
  if (!S.colorField) { S.colorMap = null; return; }
  const first = new Map();
  D.tools.forEach(t => {
    const v = t[S.colorField] || '(blank)';
    if (!first.has(v) || t.day < first.get(v)) first.set(v, t.day);
  });
  const cats = [...first.keys()].sort((a, b) => first.get(a) < first.get(b) ? -1 : first.get(a) > first.get(b) ? 1 : String(a).localeCompare(String(b)));
  S.colorMap = new Map(cats.map((c, i) => [c, i < 8 ? `var(${SLOTS[i]})` : 'var(--other)']));
}
const toolColor = t => (!S.colorField || !S.colorMap) ? 'var(--s1)' : (S.colorMap.get(t[S.colorField] || '(blank)') || 'var(--other)');
function applyColors(){
  computeColorMap();
  D.tools.forEach(t => {
    if (!t.use) return;
    const c = toolColor(t);
    t.use.setAttribute('stroke', c); t.use.setAttribute('fill', c); t.use.setAttribute('fill-opacity', '.20');
    if (t.haloEl) t.haloEl.setAttribute('stroke', c);
  });
  renderLegend(); applyDim();
}
function renderLegend(){
  const box = $('legend'); box.textContent = '';
  if (!S.colorField || !S.colorMap || S.colorMap.size < 2) return;
  const counts = new Map(), seen = new Set();
  D.tools.forEach(t => { if (seen.has(t.id) || !passes(t)) return; seen.add(t.id);
    const v = t[S.colorField] || '(blank)'; counts.set(v, (counts.get(v) || 0) + 1); });
  for (const [cat, color] of S.colorMap) {
    if (!counts.has(cat)) continue;
    const c = el('button', 'lg-chip' + (S.hidden.has(cat) ? ' off' : ''));
    const sw = el('span', 'sw'); sw.style.background = color; c.appendChild(sw);
    c.appendChild(el('span', null, cat));
    c.appendChild(el('span', 'ct num', String(counts.get(cat) || 0)));
    c.title = 'Click to show/hide';
    c.onclick = () => { S.hidden.has(cat) ? S.hidden.delete(cat) : S.hidden.add(cat); c.classList.toggle('off'); applyDim(); };
    box.appendChild(c);
  }
}
const isSearchMatch = t => !!S.search && (norm(t.id).includes(S.search) || norm(t.nm).includes(S.search)
  || FIELDS.some(f => norm(t[f.k]).includes(S.search)));
function applyDim(){
  D.tools.forEach(t => {
    if (!t.el) return;
    const hid = S.colorField && S.hidden.has(t[S.colorField] || '(blank)');
    const miss = S.search && !isSearchMatch(t);
    t.el.classList.toggle('dim', !!(hid || miss));
  });
}

/* ---------- search ---------- */
let searchPos = 0;
function applySearch(){
  S.search = norm($('search').value);
  searchPos = 0;
  applyDim();
  if (!S.search) {
    if (ghostsBuilt) D.tools.forEach(t => t.ghost && t.ghost.classList.remove('smatch'));
    return;
  }
  buildGhosts();
  let shown = 0;
  D.tools.forEach(t => {
    const m = isSearchMatch(t) && t.di > S.idx && shown < 2000;
    if (m) shown++;
    t.ghost.classList.toggle('smatch', m);
  });
}

/* ---------- layers menu ---------- */
function buildMenu(){
  const box = $('layerList'); box.textContent = '';
  D.layers.forEach(L => {
    const lab = el('label', 'ck');
    const ck = el('input'); ck.type = 'checkbox'; ck.checked = true;
    ck.onchange = () => { S.layerVis[L.n] = ck.checked; applyLayerVis(); };
    lab.appendChild(ck); lab.appendChild(el('span', null, L.n));
    const kb = Math.round(((L.tiles || []).reduce((s, T) => s + T[2].length + T[3].length + 45 * (T[4].length + T[5].length), 0)
                          + 40 * (L.t || []).length + (L.sw || 0)) / 1024);
    lab.appendChild(el('span', 'cnt num', `${fmt(L.c || 0)} · ${fmt(kb)} KB`));
    box.appendChild(lab);
  });
  $('ckText').onchange = updateBgZoom;
  const dm = $('detailMode');
  dm.onchange = () => { S.detailMode = dm.value; updateBgZoom(); };
  $('ckFast').onchange = () => $('bgSvg').classList.toggle('fast', $('ckFast').checked);
  $('ckGhost').onchange = () => {
    if ($('ckGhost').checked) buildGhosts();
    document.body.classList.toggle('show-ghost', $('ckGhost').checked);
  };
  if (CF.length) {
    $('lblCf').style.display = ''; $('tabCf').style.display = '';
    $('ckCf').onchange = () => document.body.classList.toggle('hide-cf', !$('ckCf').checked);
    const p = $('cfPill'); p.style.display = '';
    const hi = CF.filter(c => c.s === 'high').length;
    p.textContent = `⚠ ${fmt(CF.length)} conflict${CF.length === 1 ? '' : 's'}` + (hi ? ` · ${hi} high` : '');
    p.onclick = () => selectTab('cf');
  }
  if (ST) {
    $('lblZones').style.display = ''; $('tabSt').style.display = '';
    $('ckZones').onchange = () => document.body.classList.toggle('hide-zones', !$('ckZones').checked);
    ST.zones.forEach((z, i) => z.i = i);
  }
  if (CF.length || ST) document.body.classList.add('wide');
  const lm = $('labelMode');
  if (!SHOW_IDS) {
    lm.disabled = true;
    lm.title = `${fmt(D.tools.length)} tools — ID labels are off for performance`;
  }
  lm.onchange = () => { S.labelMode = lm.value; updateZoomVis(); };
  const unTotal = D.unmatchedTotal || (D.unmatched || []).length;
  if (unTotal) {
    $('unmatchedBox').style.display = '';
    const box2 = $('unmatchedList');
    $('unmatchedH').textContent = `Not found in drawing (${unTotal})`;
    D.unmatched.slice(0, 300).forEach(id => box2.appendChild(el('span', 'chip', id)));
    if (unTotal > 300) box2.appendChild(el('span', 'chip', `+${unTotal - 300} more`));
  }
}
function applyLayerVis(){
  for (const n in MAP.layerGroups)
    MAP.layerGroups[n].style.display = S.layerVis[n] === false ? 'none' : '';
}

/* ---------- playback ---------- */
const curKey = () => DAYS[S.idx];
const POPPED = []; let popT = 0;
const ROWBYID = new Map();       // 'All tools' rows, for incremental future-class updates
let allListBox = null, allBuilt = false, allBuilding = false;

/* toggle 'in' only for the tools whose state changes between two day indexes */
function applyRange(fromK, toK){
  if (fromK === toK) return;
  const on = toK > fromK, lo = Math.min(fromK, toK), hi = Math.max(fromK, toK);
  const a = idxUpTo(lo), b = idxUpTo(hi);
  const doPop = on && !REDUCED && (b - a) <= 400;
  for (let i = a; i < b; i++) {
    const t = D.tools[i];
    if (!t.el) continue;
    t.el.classList.toggle('in', on);
    if (doPop) { t.el.classList.add('pop'); POPPED.push(t.el); }
    if (t.ghost) t.ghost.classList.toggle('pending', !on);
    if (allBuilt) { const r = ROWBYID.get(t.id); if (r) r.classList.toggle('future', !on); }
  }
  if (POPPED.length) {
    clearTimeout(popT);
    popT = setTimeout(() => { POPPED.forEach(e => e.classList.remove('pop')); POPPED.length = 0; }, 600);
  }
  const ea = evUpTo(lo), eb = evUpTo(hi);
  if (on) for (let i = ea; i < eb; i++) EV[i].fn(true);
  else for (let i = eb - 1; i >= ea; i--) EV[i].fn(false);
}
let todayEls = [];
function setToday(k){
  todayEls.forEach(e => e.classList.remove('today')); todayEls = [];
  if (k < 0 || k >= DAYS.length) return;
  const a = idxUpTo(k - 1), b = idxUpTo(k);
  for (let i = a; i < b; i++) {
    const t = D.tools[i];
    if (!t.el) continue;
    ensureHalo(t);
    t.el.classList.add('today');
    todayEls.push(t.el);
  }
}
function setDay(i){
  i = Math.max(0, Math.min(DAYS.length - 1, i));
  S.idx = i;
  if (S.ready) { applyRange(S.prevK, i); setToday(i); S.prevK = i; }
  const k = curKey();
  $('statN').textContent = fmt(cumAt(i));
  $('statTotal').textContent = fmt(TOTAL);
  const today = newIds[i] || 0;
  const dl = $('statDelta');
  dl.textContent = today > 0 ? `+${today} today` : 'no move-ins today';
  dl.className = 'delta' + (today > 0 ? ' pos' : '');
  if (ST || CF.length) {
    const ex = [];
    if (ST) ex.push(`${fmt(storedAt(i))} waiting in storage`);
    if (CF.length) ex.push(`${fmt(cfAt(i))} of ${fmt(CF.length)} conflicts so far`);
    const x = $('statExtra'); x.style.display = ''; x.textContent = ex.join(' · ');
  }
  $('dateBig').textContent = fmtLong(k);
  $('dateSmall').textContent = `${fmtWd(k)} · day ${i + 1} of ${DAYS.length}`;
  $('dayCounter').textContent = `${fmt(cumAt(i))} of ${fmt(TOTAL)} tools in place`;
  const sc = $('scrub'); sc.max = DAYS.length - 1; sc.value = i;
  sc.style.setProperty('--pv', (i / Math.max(1, DAYS.length - 1) * 100) + '%');
  drawCursor(); renderList();
}
let timer = null;
function setPlaying(p){
  S.playing = p; $('playIco').textContent = p ? '⏸' : '▶';
  if (timer) { clearInterval(timer); timer = null; }
  if (p) {
    if (S.idx >= DAYS.length - 1) setDay(0);
    timer = setInterval(() => { if (S.idx >= DAYS.length - 1) setPlaying(false); else setDay(S.idx + 1); }, 1000 / S.speed);
  }
}

/* ---------- chart ---------- */
const CH = {};
function drawChart(){
  const wrap = $('chartWrap'), svg = $('chartSvg');
  svg.textContent = '';
  const W = wrap.clientWidth || 800, H = wrap.clientHeight || 76;
  svg.setAttribute('viewBox', `0 0 ${W} ${H}`);
  const binDays = Math.max(1, Math.ceil(DAYS.length / Math.floor(W / 4)));
  const bins = [];
  for (let i = 0; i < DAYS.length; i += binDays)
    bins.push({ i0: i, count: NEWP[Math.min(i + binDays, DAYS.length)] - NEWP[i] });
  CH.bins = bins; CH.binDays = binDays; CH.W = W;
  const axisH = 14, plotH = H - axisH - 4, top = 2;
  const maxC = Math.max(1, ...bins.map(b => b.count));
  const bw = W / bins.length, barW = Math.max(1.5, Math.min(24, bw - 2));
  const g = svgEl('g');
  g.appendChild(svgEl('line', { x1: 0, x2: W, y1: top + 2, y2: top + 2, stroke: 'var(--grid)', 'stroke-width': 1 }));
  const ml = svgEl('text', { x: 4, y: top + 11, 'font-size': 9.5, fill: 'var(--muted)' });
  ml.textContent = maxC + (binDays > 1 ? ` / ${binDays} days` : ' / day');
  g.appendChild(ml);
  g.appendChild(svgEl('line', { x1: 0, x2: W, y1: top + plotH, y2: top + plotH, stroke: 'var(--baseline)', 'stroke-width': 1 }));
  bins.forEach((b, i) => {
    if (!b.count) return;
    const h = Math.max(2, (plotH - 4) * b.count / maxC);
    const x = i * bw + (bw - barW) / 2, y = top + plotH - h, r = Math.min(2, barW / 2);
    const p = svgEl('path', { d: `M${x} ${top+plotH}V${y+r}Q${x} ${y} ${x+r} ${y}H${x+barW-r}Q${x+barW} ${y} ${x+barW} ${y+r}V${top+plotH}Z`,
                              fill: 'var(--s1)', 'fill-opacity': b.i0 <= S.idx ? '1' : '.32' });
    p.dataset.bin = i; g.appendChild(p);
  });
  let lastM = -1, lastX = -1e9;
  bins.forEach((b, i) => {
    const p = kp(DAYS[b.i0]), x = i * bw + 2, isM = p.m !== lastM;
    if (!isM && i % Math.max(1, Math.ceil(70 / bw)) !== 0) return;
    const txt = isM ? `${MONTHS[p.m-1]} ${p.d}` : String(p.d);
    if (x - lastX < 34) return;
    const t = svgEl('text', { x, y: H - 3, 'font-size': 9.5, fill: 'var(--muted)' });
    t.textContent = txt; g.appendChild(t);
    lastM = p.m; lastX = x + txt.length * 5;
  });
  const cx = (S.idx / Math.max(1, DAYS.length - 1)) * W;
  g.appendChild(svgEl('line', { id: 'cursor', x1: cx, x2: cx, y1: top, y2: top + plotH,
                                stroke: 'var(--ink)', 'stroke-width': 1.5, 'stroke-opacity': .75 }));
  svg.appendChild(g);
}
function drawCursor(){
  const c = $('cursor'); if (!c) { drawChart(); return; }
  const cx = (S.idx / Math.max(1, DAYS.length - 1)) * CH.W;
  c.setAttribute('x1', cx); c.setAttribute('x2', cx);
  $('chartSvg').querySelectorAll('path[data-bin]').forEach(p =>
    p.setAttribute('fill-opacity', CH.bins[+p.dataset.bin].i0 <= S.idx ? '1' : '.32'));
}
function setupChart(){
  const wrap = $('chartWrap');
  const toIdx = ev => { const r = wrap.getBoundingClientRect();
    return Math.round(Math.max(0, Math.min(1, (ev.clientX - r.left) / r.width)) * (DAYS.length - 1)); };
  let drag = false;
  wrap.addEventListener('pointerdown', ev => { drag = true; wrap.setPointerCapture(ev.pointerId); setPlaying(false); setDay(toIdx(ev)); });
  wrap.addEventListener('pointermove', ev => {
    if (drag) setDay(toIdx(ev));
    const r = wrap.getBoundingClientRect();
    const ch = $('chartX'); ch.style.display = 'block'; ch.style.left = (ev.clientX - r.left) + 'px';
    const i = toIdx(ev), k = DAYS[i];
    const n = newIds[i] || 0;
    const tip = $('tip'); tip.textContent = '';
    tip.appendChild(el('div', 't-val', `${fmtWd(k)} ${fmtLong(k)}`));
    tip.appendChild(el('div', 't-sub', `${n} move-in${n === 1 ? '' : 's'} this day`));
    tip.appendChild(el('div', 't-mut', `${fmt(cumAt(i))} of ${fmt(TOTAL)} tools in place`));
    tip.style.display = 'block'; positionTip(ev);
  });
  wrap.addEventListener('pointerup', () => drag = false);
  wrap.addEventListener('pointerleave', () => { $('chartX').style.display = 'none'; hideTip(); drag = false; });
  new ResizeObserver(() => drawChart()).observe(wrap);
}

/* ---------- tooltip ---------- */
function showToolTip(t, ev){
  const tip = $('tip'); tip.textContent = '';
  const kr = el('div', 'krow');
  const key = el('span', 'key'); key.style.background = toolColor(t);
  kr.appendChild(key); kr.appendChild(el('span', 't-val', t.id));
  tip.appendChild(kr);
  if (t.nm) tip.appendChild(el('div', 't-sub', t.nm));
  const bits = []; if (t.a) bits.push(t.a); if (t.g) bits.push(t.g);
  if (bits.length) tip.appendChild(el('div', 't-sub', bits.join(' · ')));
  if (D.byType && t.ty) tip.appendChild(el('div', 't-mut', `Type ${t.ty} — provisional slot`));
  const k = curKey();
  const state = t.day > k ? '' : (t.day === k ? ' — moving in today' : ' — in place');
  tip.appendChild(el('div', 't-mut', `Move-in: ${fmtWd(t.day)} ${fmtLong(t.day)}${state}`));
  if (t.ad) {
    const z = (ST && t.sz != null) ? ST.zones[t.sz] : null;
    const wait = Math.round((Date.parse(t.day) - Date.parse(t.ad)) / 864e5);
    tip.appendChild(el('div', 't-store', `Arrives ${fmtShort(t.ad)} · ${wait} day${wait === 1 ? '' : 's'} in `
      + (z ? `${z.n}, slot ${t.sl}` : 'storage — NO SLOT FOUND') + (inStorage(t, S.idx) ? ' (there now)' : '')));
  }
  (t.cf || []).slice(0, 3).forEach(i => tip.appendChild(el('div', 't-warn', `⚠ ${CF_TITLE[CF[i].k]}: ${CF[i].d}`)));
  if ((t.cf || []).length > 3) tip.appendChild(el('div', 't-warn', `… ${t.cf.length - 3} more (Conflicts tab)`));
  tip.style.display = 'block'; positionTip(ev);
}
function showConflictTip(c, ev){
  const tip = $('tip'); tip.textContent = '';
  tip.appendChild(el('div', 't-val', `⚠ ${CF_TITLE[c.k]}`));
  tip.appendChild(el('div', 't-sub', c.ids.slice(0, 6).join(', ') + (c.ids.length > 6 ? ` … +${c.ids.length - 6}` : '')));
  tip.appendChild(el('div', 't-mut', c.d));
  tip.appendChild(el('div', 't-mut', `${c.s} · from ${fmtLong(c.day)} · click to list`));
  tip.style.display = 'block'; positionTip(ev);
}
function zoneOccupants(z, i){ return STORED.filter(t => t.sz === z.i && inStorage(t, i)); }
function showZoneTip(z, ev){
  const tip = $('tip'); tip.textContent = '';
  const occ = zoneOccupants(z, S.idx);
  tip.appendChild(el('div', 't-val', z.n));
  tip.appendChild(el('div', 't-sub', `Storage zone · ${z.cols}×${z.rows} slots` + (z.cap < z.cols * z.rows ? ` · max ${z.cap} tools` : '')));
  tip.appendChild(el('div', 't-mut', `${occ.length} of ${z.cap} in use today` + (occ.length ? ': ' + occ.slice(0, 8).map(t => t.id).join(', ') + (occ.length > 8 ? ' …' : '') : '')));
  tip.style.display = 'block'; positionTip(ev);
}
function positionTip(ev){
  const tip = $('tip'), r = tip.getBoundingClientRect(), pad = 14;
  let x = ev.clientX + pad, y = ev.clientY + pad;
  if (x + r.width > innerWidth - 8) x = ev.clientX - r.width - pad;
  if (y + r.height > innerHeight - 8) y = ev.clientY - r.height - pad;
  tip.style.left = x + 'px'; tip.style.top = y + 'px';
}
const hideTip = () => $('tip').style.display = 'none';

/* ---------- side list ---------- */
const dedupeById = a => { const s = new Set(); return a.filter(t => s.has(t.id) ? false : (s.add(t.id), true)); };
function makeRow(t, future){
  const r = el('div', 'tl-row' + (future ? ' future' : ''));
  const sw = el('span', 'sw'); sw.style.background = toolColor(t); r.appendChild(sw);
  r.appendChild(el('span', 'id', t.id));
  r.appendChild(el('span', 'nm', t.nm || t.f0 || t.a || ''));
  r.appendChild(el('span', 'dt num', fmtShort(t.day)));
  r.onclick = () => {
    if (t.di > S.idx) { setPlaying(false); setDay(Math.min(DAYS.length - 1, Math.max(0, t.di))); }
    zoomToTool(t);
  };
  return r;
}
/* the 'All tools' list is built once (in chunks) and then only patched */
function buildAllList(){
  if (allBuilt || allBuilding) return; allBuilding = true;
  allListBox = el('div');
  const uniq = dedupeById(D.tools).filter(t => t.di < DAYS.length && passes(t));
  let i = 0, last = null;
  (function step(){
    const end = Math.min(uniq.length, i + 500);
    const frag = document.createDocumentFragment();
    for (; i < end; i++) {
      const t = uniq[i];
      if (t.day !== last) { frag.appendChild(el('div', 'tl-day', `${fmtWd(t.day)} · ${fmtLong(t.day)}`)); last = t.day; }
      const r = makeRow(t, t.di > S.idx);
      ROWBYID.set(t.id, r);
      frag.appendChild(r);
    }
    allListBox.appendChild(frag);
    if (i < uniq.length) requestAnimationFrame(step);
    else { allBuilt = true; allBuilding = false; if (S.tab === 'all') renderList(); }
  })();
}
function makeConflictRow(c){
  const r = el('div', 'tl-row cf-row' + (c.di > S.idx ? ' future' : ''));
  r.appendChild(el('span', 'kind ' + c.s, c.k));
  r.appendChild(el('span', 'id', c.ids.slice(0, 3).join(', ') + (c.ids.length > 3 ? ` +${c.ids.length - 3}` : '')));
  r.appendChild(el('span', 'dt num', fmtShort(c.day)));
  r.appendChild(el('span', 'nm', `${CF_TITLE[c.k]} — ${c.d}`));
  r.onclick = () => focusConflict(c, true);
  return r;
}
function makeStoredRow(t, i){
  const r = el('div', 'tl-row' + (t.adi > i ? ' future' : ''));
  const sw = el('span', 'sw'); sw.style.background = toolColor(t); r.appendChild(sw);
  r.appendChild(el('span', 'id', t.id));
  const z = (ST && t.sz != null) ? ST.zones[t.sz] : null;
  r.appendChild(el('span', 'st', z ? `${z.n} · slot ${t.sl}` : 'no slot'));
  r.appendChild(el('span', 'dt num', `${fmtShort(t.ad)} → ${fmtShort(t.day)}`));
  r.onclick = () => { if (t.adi > S.idx) { setPlaying(false); setDay(Math.max(0, t.adi)); } zoomToTool(t); };
  return r;
}
function renderList(){
  const box = $('toolList');
  const k = curKey();
  if (S.tab === 'day') {
    box.textContent = '';
    const list = dedupeById(byDay.get(k) || []).filter(passes);
    if (!list.length) box.appendChild(el('div', 'tl-empty', 'No tools move in on this day.'));
    list.forEach(t => box.appendChild(makeRow(t, false)));
    const arriving = STORED.filter(t => t.adi === S.idx);
    if (arriving.length) {
      box.appendChild(el('div', 'tl-day', 'Arriving into storage'));
      arriving.forEach(t => box.appendChild(makeStoredRow(t, S.idx)));
    }
    const cfs = CF.filter(c => c.di === S.idx);
    if (cfs.length) {
      box.appendChild(el('div', 'tl-day', 'Conflicts from today'));
      cfs.forEach(c => box.appendChild(makeConflictRow(c)));
    }
  } else if (S.tab === 'cf') {
    box.textContent = '';
    if (!CF.length) box.appendChild(el('div', 'tl-empty', 'No potential conflicts found.'));
    let last = null;
    CF.forEach(c => {
      if (c.s !== last) { box.appendChild(el('div', 'tl-day', `${c.s} (${CF.filter(x => x.s === c.s).length})`)); last = c.s; }
      box.appendChild(makeConflictRow(c));
    });
  } else if (S.tab === 'st') {
    box.textContent = '';
    ST.zones.forEach(z => {
      const occ = zoneOccupants(z, S.idx);
      box.appendChild(el('div', 'tl-day', `${z.n} · ${occ.length} / ${z.cap} in use`));
      occ.forEach(t => box.appendChild(makeStoredRow(t, S.idx)));
    });
    const soon = STORED.filter(t => t.adi > S.idx).sort((a, b) => a.adi - b.adi).slice(0, 40);
    if (soon.length) {
      box.appendChild(el('div', 'tl-day', 'Arriving next'));
      soon.forEach(t => box.appendChild(makeStoredRow(t, S.idx)));
    }
    const none = CF.filter(c => c.k === 'storage');
    if (none.length) {
      box.appendChild(el('div', 'tl-day', `No slot found (${none.length})`));
      none.forEach(c => box.appendChild(makeConflictRow(c)));
    }
    if (!STORED.length && !none.length) box.appendChild(el('div', 'tl-empty', 'No tool arrives before its move-in day, so nothing needs storage.'));
  } else if (S.tab === 'next') {
    box.textContent = '';
    const src = [], seen = new Set();
    for (let j = idxUpTo(S.idx); j < D.tools.length && src.length < 60; j++) {
      const t = D.tools[j];
      if (t.di >= DAYS.length) break;
      if (!seen.has(t.id) && passes(t)) { seen.add(t.id); src.push(t); }
    }
    if (!src.length) box.appendChild(el('div', 'tl-empty', 'All tools are in place.'));
    let last = null;
    src.forEach(t => {
      if (t.day !== last) { box.appendChild(el('div', 'tl-day', `${fmtWd(t.day)} · ${fmtLong(t.day)}`)); last = t.day; }
      box.appendChild(makeRow(t, true));
    });
  } else {   // 'all'
    if (!allBuilt) {
      box.textContent = '';
      box.appendChild(el('div', 'tl-empty', 'Building list…'));
      buildAllList();
      return;
    }
    if (box.firstChild !== allListBox) { box.textContent = ''; box.appendChild(allListBox); }
  }
}
function selectTab(name){
  document.querySelectorAll('.tabs button').forEach(x => x.classList.toggle('on', x.dataset.tab === name));
  S.tab = name; renderList();
}
document.querySelectorAll('.tabs button').forEach(b => b.onclick = () => selectTab(b.dataset.tab));

/* ---------- PNG export (composes background + foreground svg) ---------- */
$('pngBtn').onclick = () => {
  commitBg();                                   // background must be sharp, not bitmap-scaled
  const r = $('mapSvg').getBoundingClientRect(), scale = 2;
  const cs = getComputedStyle(document.body);
  const vars = ['--map-ink','--map-ink-strong','--map-text','--ink-2','--surface','--s1','--s2','--s3','--s4','--s5','--s6','--s7','--s8','--other'];
  let css = ':root{' + vars.map(v => `${v}:${cs.getPropertyValue(v)}`).join(';') + '}';
  css += document.getElementById('mainStyle').textContent;
  const cv = document.createElement('canvas');
  cv.width = r.width * scale; cv.height = r.height * scale;
  const ctx = cv.getContext('2d');
  ctx.fillStyle = cs.getPropertyValue('--surface') || '#fff';
  ctx.fillRect(0, 0, cv.width, cv.height);
  const draw = svg => new Promise((res, rej) => {
    const clone = svg.cloneNode(true);
    const st = document.createElementNS('http://www.w3.org/2000/svg', 'style');
    st.textContent = css;
    clone.insertBefore(st, clone.firstChild);
    clone.setAttribute('width', r.width); clone.setAttribute('height', r.height);
    clone.setAttribute('xmlns', 'http://www.w3.org/2000/svg');
    const blob = new Blob([new XMLSerializer().serializeToString(clone)], { type: 'image/svg+xml;charset=utf-8' });
    const url = URL.createObjectURL(blob), img = new Image();
    img.onload = () => { ctx.drawImage(img, 0, 0, cv.width, cv.height); URL.revokeObjectURL(url); res(); };
    img.onerror = () => { URL.revokeObjectURL(url); rej(new Error('render')); };
    img.src = url;
  });
  draw($('bgSvg')).then(() => draw($('mapSvg'))).then(() => {
    const a = document.createElement('a');
    a.download = `fab-movein-${curKey()}.png`;
    a.href = cv.toDataURL('image/png');
    a.click();
  }).catch(() => alert('PNG export failed in this browser — use a screenshot instead.'));
};

/* ---------- wiring ---------- */
$('playBtn').onclick  = () => setPlaying(!S.playing);
$('stepBack').onclick = () => { setPlaying(false); setDay(S.idx - 1); };
$('stepFwd').onclick  = () => { setPlaying(false); setDay(S.idx + 1); };
$('speed').onchange   = () => { S.speed = +$('speed').value; if (S.playing) setPlaying(true); };
$('scrub').oninput    = () => { setPlaying(false); setDay(+$('scrub').value); };
$('search').oninput   = applySearch;
$('search').addEventListener('keydown', ev => {
  if (ev.key !== 'Enter' || !S.search) return;
  const ms = D.tools.filter(isSearchMatch);
  if (!ms.length) return;
  zoomToTool(ms[searchPos++ % ms.length]);
});
document.addEventListener('keydown', ev => {
  if (/^(INPUT|SELECT|TEXTAREA)$/.test(document.activeElement.tagName)) return;
  if (ev.key === ' ') { ev.preventDefault(); setPlaying(!S.playing); }
  else if (ev.key === 'ArrowRight') { setPlaying(false); setDay(S.idx + 1); }
  else if (ev.key === 'ArrowLeft')  { setPlaying(false); setDay(S.idx - 1); }
  else if (ev.key === 'Home') setDay(0);
  else if (ev.key === 'End')  setDay(DAYS.length - 1);
});
document.addEventListener('pointerdown', ev => {
  const m = $('viewMenu'); if (m.open && !m.contains(ev.target)) m.open = false;
});

/* ---------- go ---------- */
document.title = D.title;
$('hTitle').textContent = D.title;
$('hSub').textContent = D.subtitle;
$('mapTitle').textContent = 'building map…';
{
  const unTotal = D.unmatchedTotal || (D.unmatched || []).length;
  if (unTotal) {
    const w = $('warnPill'); w.style.display = '';
    w.textContent = `${unTotal} not placed`;
    w.title = `${unTotal} schedule row(s) had no matching block in the drawing. See View ▾ for the list.`;
  }
}
buildMap(() => {
  PERF.bg = performance.now();
  buildToolsChunked(() => {
    buildExtras();
    applyColors();
    S.ready = true; S.prevK = -2;
    setDay(S.idx);
    PERF.done = performance.now();
    const secs = ms => (ms / 1000).toFixed(1) + 's';
    /* hover this label for a breakdown of where load time went */
    const perf = `ready in ${secs(PERF.done)} · payload ${(D.__jsonKB / 1024).toFixed(1)} MB `
               + `(parse ${secs(D.__parseMs || 0)}) · map build ${secs(PERF.bg - PERF.start)} `
               + `· tools ${secs(PERF.done - PERF.bg)} · rest = HTML load + first paint`;
    $('mapTitle').textContent = `${fmt(TOTAL)} tools · ${fmt(DAYS.length)} days`;
    $('mapTitle').title = perf;
    console.log('[fab-movein] ' + perf);
    updateZoomVis();
  });
});
setupMap(); buildColorControls(); buildFilters(); buildMenu(); setupChart();
setDay(S.idx);                       // UI numbers while the map streams in

</script>
</body>
</html>
'''


# --------------------------------------------------------------------------
# small helpers
# --------------------------------------------------------------------------

def norm(s) -> str:
    """Normalise an identifier for matching: trim, collapse spaces, uppercase."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip()).upper()


class Match:
    """Global switch set by --loose: ignore punctuation/spacing in identifiers."""
    loose = False


def compact(s) -> str:
    """Identifier reduced to letters+digits only: DT-112 == DT_112 == DT 112."""
    return re.sub(r"[^A-Z0-9]", "", norm(s))


def mkey(s) -> str:
    """Key used wherever schedule values are matched against the drawing."""
    return compact(s) if Match.loose else norm(s)


class BlockDetail:
    """--block-detail: how much of each block's internal geometry to keep."""
    mode = "auto"           # auto | full | outline | box


def convex_hull(pts):
    """Andrew's monotone chain; returns the hull counter-clockwise (open)."""
    pts = sorted(set(pts))
    if len(pts) < 3:
        return pts

    def cross(o, a, b):
        return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])

    lower = []
    for q in pts:
        while len(lower) >= 2 and cross(lower[-2], lower[-1], q) <= 0:
            lower.pop()
        lower.append(q)
    upper = []
    for q in reversed(pts):
        while len(upper) >= 2 and cross(upper[-2], upper[-1], q) <= 0:
            upper.pop()
        upper.append(q)
    return lower[:-1] + upper[:-1]


def die(msg: str, code: int = 1):
    print(f"\nERROR: {msg}\n", file=sys.stderr)
    sys.exit(code)


def warn(msg: str):
    print(f"  ! {msg}")


class Prec:
    """Coordinate rounding helper (digits chosen from the drawing size)."""
    digits = 2

    @classmethod
    def r(cls, v: float) -> float:
        v = round(float(v), cls.digits)
        return 0.0 if v == 0 else v


def digits_for(size: float) -> int:
    """Decimals needed to resolve 1/50,000 of the drawing extent."""
    if not size or size <= 0:
        return 2
    if size >= 50000:
        return 0
    return max(0, min(6, int(math.ceil(math.log10(50000.0 / size)))))


def numd(v: float) -> str:
    """Compact number for SVG path data (value already rounded)."""
    if v == int(v):
        return str(int(v))
    return repr(v)


def simplify_poly(pts, tol: float):
    """Iterative Ramer-Douglas-Peucker: drop points closer than tol to the
    line through their neighbours. This is where flattened arcs, splines and
    survey-dense polylines lose 80-95% of their vertices."""
    n = len(pts)
    if tol <= 0 or n < 3:
        return pts
    t2 = tol * tol
    keep = bytearray(n)
    keep[0] = keep[n - 1] = 1
    stack = [(0, n - 1)]
    while stack:
        i0, i1 = stack.pop()
        if i1 - i0 < 2:
            continue
        x0, y0 = pts[i0]
        x1, y1 = pts[i1]
        dx, dy = x1 - x0, y1 - y0
        seg2 = dx * dx + dy * dy
        best, bi = -1.0, -1
        for i in range(i0 + 1, i1):
            px, py = pts[i]
            if seg2 <= 1e-30:
                ex, ey = px - x0, py - y0
                d2 = ex * ex + ey * ey
            else:
                t = ((px - x0) * dx + (py - y0) * dy) / seg2
                if t < 0.0:
                    t = 0.0
                elif t > 1.0:
                    t = 1.0
                ex, ey = px - (x0 + t * dx), py - (y0 + t * dy)
                d2 = ex * ex + ey * ey
            if d2 > best:
                best, bi = d2, i
        if best > t2:
            keep[bi] = 1
            stack.append((i0, bi))
            stack.append((bi, i1))
    return [p for p, k in zip(pts, keep) if k]


# --------------------------------------------------------------------------
# affine transforms (block insert placement)
# --------------------------------------------------------------------------

def insert_matrix(entity):
    """2D affine (a, b, c, d, e, f) of an INSERT, mapping block-local (x, y)
    to WCS:  X = a*x + c*y + e ;  Y = b*x + d*y + f.  None if unavailable."""
    try:
        m = entity.matrix44()
        r0, r1, r3 = m.get_row(0), m.get_row(1), m.get_row(3)
        return (float(r0[0]), float(r0[1]), float(r1[0]),
                float(r1[1]), float(r3[0]), float(r3[1]))
    except Exception:
        return None


def mat_apply(M, x, y):
    a, b, c, d, e, f = M
    return (a * x + c * y + e, b * x + d * y + f)


def mat_scale(M):
    """Largest axis scale factor of the affine part."""
    return max(math.hypot(M[0], M[1]), math.hypot(M[2], M[3]), 1e-12)


def mat_avg_scale(M):
    return (math.hypot(M[0], M[1]) + math.hypot(M[2], M[3])) / 2.0


def mat_rotation_deg(M):
    return math.degrees(math.atan2(M[1], M[0]))


def bbox_transform(M, b):
    """WCS axis-aligned bbox of a local bbox pushed through the affine."""
    pts = [mat_apply(M, b[0], b[1]), mat_apply(M, b[2], b[1]),
           mat_apply(M, b[2], b[3]), mat_apply(M, b[0], b[3])]
    xs = [p[0] for p in pts]
    ys = [p[1] for p in pts]
    return (min(xs), min(ys), max(xs), max(ys))


# --------------------------------------------------------------------------
# DXF geometry extraction
# --------------------------------------------------------------------------

class Geom:
    """Collected geometry: polylines (lists of (x, y)) and text placements."""

    __slots__ = ("polys", "texts")

    def __init__(self):
        self.polys = []          # list[list[tuple[float, float]]]
        self.texts = []          # list[(x, y, height, rotation, string)]

    def extend(self, other: "Geom"):
        self.polys.extend(other.polys)
        self.texts.extend(other.texts)

    def bbox(self):
        xs0 = ys0 = math.inf
        xs1 = ys1 = -math.inf
        for poly in self.polys:
            for x, y in poly:
                xs0 = min(xs0, x); ys0 = min(ys0, y)
                xs1 = max(xs1, x); ys1 = max(ys1, y)
        if xs0 is math.inf:
            return None
        return (xs0, ys0, xs1, ys1)


def _plain_text(entity) -> str:
    for meth in ("plain_text", "text"):
        try:
            val = getattr(entity, meth)
            val = val() if callable(val) else val
            if val:
                return re.sub(r"\s+", " ", str(val)).strip()
        except Exception:
            continue
    try:
        return re.sub(r"\s+", " ", str(entity.dxf.text)).strip()
    except Exception:
        return ""


def collect_entity(entity, geom: Geom, flat: float, tol: float, skipped: Counter, depth: int = 0):
    """Turn one DXF entity into polylines / texts, recursing into blocks."""
    from ezdxf import path as ezpath

    etype = entity.dxftype()

    if etype == "INSERT":
        if depth > 8:
            skipped["INSERT (too deeply nested)"] += 1
            return
        try:
            for sub in entity.virtual_entities():
                collect_entity(sub, geom, flat, tol, skipped, depth + 1)
        except Exception:
            skipped["INSERT (could not explode)"] += 1
        return

    if etype in ("TEXT", "MTEXT", "ATTRIB"):
        txt = _plain_text(entity)
        if not txt:
            return
        try:
            if etype == "MTEXT":
                pt = entity.dxf.insert
                height = float(getattr(entity.dxf, "char_height", 1.0) or 1.0)
            else:
                pt = entity.dxf.insert
                if getattr(entity.dxf, "halign", 0) or getattr(entity.dxf, "valign", 0):
                    ap = getattr(entity.dxf, "align_point", None)
                    if ap is not None and (ap.x or ap.y):
                        pt = ap
                height = float(getattr(entity.dxf, "height", 1.0) or 1.0)
            rot = float(getattr(entity.dxf, "rotation", 0.0) or 0.0)
            geom.texts.append((float(pt.x), float(pt.y), height, rot, txt))
        except Exception:
            skipped[f"{etype} (bad placement)"] += 1
        return

    if etype in ("SOLID", "TRACE", "3DFACE"):
        try:
            pts = [entity.dxf.vtx0, entity.dxf.vtx1, entity.dxf.vtx3, entity.dxf.vtx2]
            ring = [(float(p.x), float(p.y)) for p in pts]
            # vtx3 == vtx2 for triangles; drop the duplicate
            ded = [p for i, p in enumerate(ring) if i == 0 or p != ring[i - 1]]
            geom.polys.append(ded + [ded[0]])
        except Exception:
            skipped[etype] += 1
        return

    if etype == "HATCH":
        try:
            for p in ezpath.from_hatch(entity):
                pts = [(float(v.x), float(v.y)) for v in p.flattening(flat)]
                if len(pts) > 1:
                    geom.polys.append(simplify_poly(pts, tol))
            return
        except Exception:
            skipped["HATCH"] += 1
            return

    if etype in ("DIMENSION", "LEADER", "MULTILEADER", "MLEADER"):
        try:
            for sub in entity.virtual_entities():
                collect_entity(sub, geom, flat, tol, skipped, depth + 1)
            return
        except Exception:
            skipped[etype] += 1
            return

    # everything else ezdxf can express as a path: LINE, ARC, CIRCLE, ELLIPSE,
    # LWPOLYLINE, POLYLINE, SPLINE, HELIX, ...
    try:
        p = ezpath.make_path(entity)
        pts = [(float(v.x), float(v.y)) for v in p.flattening(flat)]
        if len(pts) > 1:
            geom.polys.append(simplify_poly(pts, tol))
    except Exception:
        skipped[etype] += 1


def path_data(polys, flip_y=True, dx=0.0, dy=0.0, digits=None) -> str:
    """SVG path 'd' string. DXF y grows up, SVG y grows down -> negate y.
    Consecutive points that round to the same coordinate are merged."""
    dg = Prec.digits if digits is None else digits
    out = []
    for poly in polys:
        if len(poly) < 2:
            continue
        seg = []
        prev = None
        for x, y in poly:
            sx = round(x - dx, dg)
            sy = round((-y if flip_y else y) - dy, dg)
            if sx == 0:
                sx = 0.0
            if sy == 0:
                sy = 0.0
            if prev == (sx, sy):
                continue
            seg.append(("M" if not seg else "L") + numd(sx) + " " + numd(sy))
            prev = (sx, sy)
        if len(seg) >= 2:
            out.append("".join(seg))
    return "".join(out)


def poly_bbox_svg(poly):
    """Axis-aligned bbox of one polyline in SVG coords (y negated)."""
    xs0 = ys0 = math.inf
    xs1 = ys1 = -math.inf
    for x, y in poly:
        if x < xs0: xs0 = x
        if x > xs1: xs1 = x
        if -y < ys0: ys0 = -y
        if -y > ys1: ys1 = -y
    return (xs0, ys0, xs1, ys1)


# --------------------------------------------------------------------------
# DXF loading
# --------------------------------------------------------------------------

class Drawing:
    def __init__(self, path: str, simplify: float | None = None):
        try:
            import ezdxf
            from ezdxf import bbox
        except ImportError:
            die("ezdxf is not installed.  Run:  pip install ezdxf openpyxl")

        if not os.path.exists(path):
            die(f"DXF file not found: {path}")

        print(f"Reading DXF  {path}")
        try:
            self.doc = ezdxf.readfile(path)
        except Exception as exc:
            hint = ""
            low = path.lower()
            if low.endswith(".dwg"):
                hint = ("\n  This is a DWG file. ezdxf reads DXF only -- in AutoCAD use "
                        "SAVEAS and pick 'AutoCAD 2013 DXF' (or any ASCII DXF).")
            die(f"could not read the DXF: {exc}{hint}")

        self.msp = self.doc.modelspace()

        # tolerances proportional to the drawing size: curve flattening
        # (max sagitta) and polyline simplification. Both are invisible at
        # fit-to-view zoom and shrink the output enormously. The size is
        # measured robustly (2nd-98th percentile of entity anchor points), so
        # one stray entity a mile away cannot coarsen the whole drawing.
        size = self._robust_size()
        if size <= 0:
            try:
                ext = bbox.extents(self.msp, fast=True)
                size = max(ext.size.x, ext.size.y)
            except Exception:
                pass
        self.size = size if size > 0 else 100.0
        self.flat = max(self.size / 6000.0, 1e-9)
        self.tol = (self.size / 6000.0) if simplify is None else max(0.0, simplify)
        self.autodigits = digits_for(self.size)

        self.skipped = Counter()
        self.inserts = []     # {name, attribs, layer, m, bbox, pt, geom(minsert only)}
        self.texts = []       # {str, x, y, h, layer, idx}
        self.layers = defaultdict(lambda: {"geom": Geom(), "count": 0})
        self._blocks = {}     # block name -> {"geom": local Geom, "bbox", "base"}
        self._lite = {}       # block name -> reduced-detail local Geom
        self._icache = {}     # (index kind, tag, loose?) -> lookup dict
        self.reduced_blocks = 0
        self._maxscale = defaultdict(lambda: 1.0)
        self._scan()

    def _robust_size(self) -> float:
        """Drawing extent from the 2%..98% spread of cheap entity anchor
        points -- immune to a handful of stray entities far off the plan."""
        xs, ys = [], []
        for e in self.msp:
            p = None
            for attr in ("insert", "start", "center"):
                try:
                    p = getattr(e.dxf, attr)
                    break
                except Exception:
                    continue
            if p is None:
                try:
                    p = e.get_points(format="xy")[0]  # LWPOLYLINE first vertex
                except Exception:
                    continue
            try:
                xs.append(float(p[0])); ys.append(float(p[1]))
            except Exception:
                continue
        if len(xs) < 10:
            return 0.0
        xs.sort(); ys.sort()
        lo, hi = int(len(xs) * 0.02), max(int(len(xs) * 0.98), 1) - 1
        return max(xs[hi] - xs[lo], ys[hi] - ys[lo])

    def _scan(self):
        n = 0
        raw_inserts = []
        for e in self.msp:
            n += 1
            etype = e.dxftype()
            layer = str(getattr(e.dxf, "layer", "0") or "0")

            if etype == "INSERT":
                attribs = {}
                try:
                    for a in e.attribs:
                        tag = norm(getattr(a.dxf, "tag", ""))
                        if tag:
                            attribs[tag] = _plain_text(a)
                except Exception:
                    pass
                rows = int(getattr(e.dxf, "row_count", 1) or 1)
                cols = int(getattr(e.dxf, "column_count", 1) or 1)
                M = insert_matrix(e) if rows <= 1 and cols <= 1 else None
                raw_inserts.append({"ent": e, "name": str(e.dxf.name),
                                    "attribs": attribs, "layer": layer, "m": M})
                continue

            g = Geom()
            collect_entity(e, g, self.flat, self.tol, self.skipped)
            if etype in ("TEXT", "MTEXT") and g.texts:
                x, y, h, rot, s = g.texts[0]
                self.texts.append({"str": s, "x": x, "y": y, "h": h,
                                   "layer": layer, "idx": len(self.texts)})
            self.layers[layer]["geom"].extend(g)
            self.layers[layer]["count"] += 1

        # blocks are exploded ONCE per definition (in local coordinates) and
        # reused by every insert -- this is the big generation + output win
        # for fab layouts full of repeated fixtures.
        for r in raw_inserts:
            if r["m"] is not None:
                self._maxscale[r["name"]] = max(self._maxscale[r["name"]], mat_scale(r["m"]))

        for r in raw_inserts:
            name, M = r["name"], r["m"]
            if M is None:                       # MINSERT array or no matrix: old path
                g = Geom()
                collect_entity(r["ent"], g, self.flat, self.tol, self.skipped)
                b = g.bbox()
                if b:
                    pt = ((b[0] + b[2]) / 2.0, (b[1] + b[3]) / 2.0)
                else:
                    try:
                        p = r["ent"].dxf.insert
                        pt = (float(p.x), float(p.y))
                    except Exception:
                        pt = (0.0, 0.0)
                self.inserts.append({"name": name, "attribs": r["attribs"],
                                     "layer": r["layer"], "m": None,
                                     "geom": g, "bbox": b, "pt": pt})
            else:
                bd = self.blockdef(name)
                b = bbox_transform(M, bd["bbox"]) if bd["bbox"] else None
                pt = mat_apply(M, bd["base"][0], bd["base"][1])
                self.inserts.append({"name": name, "attribs": r["attribs"],
                                     "layer": r["layer"], "m": M,
                                     "geom": None, "bbox": b, "pt": pt})

        print(f"  {n:,} top-level entities · {len(self.inserts):,} block inserts · "
              f"{len(self.layers)} layers · {len(self._blocks)} block definitions used")
        if self.skipped:
            detail = ", ".join(f"{k}×{v}" for k, v in self.skipped.most_common(6))
            warn(f"skipped some entity types: {detail}")

    def blockdef(self, name):
        """Local-space geometry of one block definition, exploded once."""
        bd = self._blocks.get(name)
        if bd is None:
            g = Geom()
            base = (0.0, 0.0)
            try:
                blk = self.doc.blocks[name]
                bp = blk.base_point
                base = (float(bp[0]), float(bp[1]))
                sc = max(self._maxscale.get(name, 1.0), 1e-9)
                fl, tl = self.flat / sc, (self.tol / sc if self.tol else 0.0)
                for e in blk:
                    if e.dxftype() == "ATTDEF":
                        continue
                    collect_entity(e, g, fl, tl, self.skipped)
            except Exception:
                pass
            bd = {"geom": g, "bbox": g.bbox(), "base": base}
            self._blocks[name] = bd
        return bd

    def block_lite(self, name):
        """Block geometry with the detail policy applied. Tool blocks in fab
        drawings are often complete equipment drawings; for a layout view
        only the footprint matters. 'auto' keeps pieces at least 12% of the
        block's size (outline, major internal lines), drops the rest, and
        falls back to the convex-hull footprint when what is left is still
        heavy or does not cover the block. Blocks that are large relative to
        the drawing (bound xrefs, whole bays) are always kept in full."""
        bd = self.blockdef(name)
        full = bd["geom"]
        bb = bd["bbox"]
        if BlockDetail.mode == "full" or not full.polys or not bb:
            return full
        lt = self._lite.get(name)
        if lt is not None:
            return lt
        size = max(bb[2] - bb[0], bb[3] - bb[1])
        if size >= 0.15 * self.size or size <= 0:
            self._lite[name] = full
            return full
        rect = [(bb[0], bb[1]), (bb[2], bb[1]), (bb[2], bb[3]), (bb[0], bb[3]), (bb[0], bb[1])]

        def hull_geom():
            g = Geom()
            h = convex_hull([q for poly in full.polys for q in poly])
            g.polys.append(h + [h[0]] if len(h) >= 3 else rect)
            return g

        if BlockDetail.mode == "box":
            g = Geom(); g.polys.append(rect)
        elif BlockDetail.mode == "outline":
            g = hull_geom()
        else:                                   # auto
            kept = []
            for poly in full.polys:
                xs = [q[0] for q in poly]; ys = [q[1] for q in poly]
                if max(max(xs) - min(xs), max(ys) - min(ys)) >= 0.12 * size:
                    kept.append(poly)
            npts = sum(len(poly) for poly in kept)
            cover = 0.0
            if kept:
                kx = [q[0] for poly in kept for q in poly]; ky = [q[1] for poly in kept for q in poly]
                cover = max(max(kx) - min(kx), max(ky) - min(ky))
            if not kept or cover < 0.6 * size or npts > 1500:
                g = hull_geom()
            else:
                g = Geom(); g.polys = kept
        if sum(len(poly) for poly in g.polys) < sum(len(poly) for poly in full.polys):
            self.reduced_blocks += 1
        self._lite[name] = g
        return g

    def insert_wcs_geom(self, ins):
        """(Geom in WCS, bbox) for one insert; geometry may be empty."""
        if ins["m"] is None:
            return ins["geom"], ins["bbox"]
        bd = self.blockdef(ins["name"])
        if not bd["geom"].polys:
            return Geom(), None
        M = ins["m"]
        g = Geom()
        xs0 = ys0 = math.inf
        xs1 = ys1 = -math.inf
        for poly in self.block_lite(ins["name"]).polys:
            tp = [mat_apply(M, x, y) for x, y in poly]
            for x, y in tp:
                if x < xs0: xs0 = x
                if x > xs1: xs1 = x
                if y < ys0: ys0 = y
                if y > ys1: ys1 = y
            g.polys.append(tp)
        return g, (xs0, ys0, xs1, ys1)

    # ---- indexes used for matching -------------------------------------
    def attrib_tags(self):
        tags = Counter()
        for ins in self.inserts:
            for t in ins["attribs"]:
                tags[t] += 1
        return tags

    # The lookups below are built once per run and shared by matching,
    # auto-detection, --inspect and --trace (callers must not mutate them).
    def _cached(self, key, build):
        k = (key, Match.loose)
        idx = self._icache.get(k)
        if idx is None:
            idx = self._icache[k] = build()
        return idx

    def index_by_attrib(self, tag):
        def build():
            idx = defaultdict(list)
            for ins in self.inserts:
                v = mkey(ins["attribs"].get(tag))
                if v:
                    idx[v].append(ins)
            return idx
        return self._cached(("attrib", tag), build)

    def index_by_blockname(self):
        def build():
            idx = defaultdict(list)
            for ins in self.inserts:
                idx[mkey(ins["name"])].append(ins)
            return idx
        return self._cached(("blockname", None), build)

    def index_by_text(self):
        def build():
            idx = defaultdict(list)
            for t in self.texts:
                k = mkey(t["str"])
                if k:
                    idx[k].append(t)
            return idx
        return self._cached(("text", None), build)

    def index_by_blocktext(self):
        """Text drawn INSIDE a block definition labels every insert of that
        block. The string is stored once in the file (which is why FIND and a
        text-editor search see it once), but it identifies all placements."""
        return self._cached(("blocktext", None), self._build_blocktext)

    def _build_blocktext(self):
        idx = defaultdict(list)
        per_name = {}
        for ins in self.inserts:
            if ins["m"] is None:
                geom = ins["geom"]
                keys = {mkey(s) for (_x, _y, _h, _r, s) in (geom.texts if geom else [])
                        if mkey(s)}
            else:
                name = ins["name"]
                if name not in per_name:
                    bd = self.blockdef(name)
                    per_name[name] = {mkey(s) for (_x, _y, _h, _r, s) in bd["geom"].texts
                                      if mkey(s)}
                keys = per_name[name]
            for k in keys:
                idx[k].append(ins)
        return idx


# --------------------------------------------------------------------------
# schedule loading
# --------------------------------------------------------------------------

class Schedule:
    def __init__(self, path: str, sheet: str | None = None):
        if not os.path.exists(path):
            die(f"Schedule file not found: {path}")
        print(f"Reading schedule  {path}")
        low = path.lower()
        self.sheet_names = []
        if low.endswith((".csv", ".tsv", ".txt")):
            self._read_csv(path)
            self.sheet = "(csv)"
        else:
            self._read_xlsx(path, sheet)
        print(f"  {len(self.rows):,} data rows · {len(self.headers)} columns"
              + (f" · sheet '{self.sheet}'" if self.sheet != "(csv)" else ""))

    def _read_csv(self, path):
        with open(path, "r", encoding="utf-8-sig", newline="") as fh:
            sample = fh.read(8192)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except Exception:
                dialect = csv.excel
            rows = [r for r in csv.reader(fh, dialect)]
        self._split_header(rows)

    def _read_xlsx(self, path, sheet):
        try:
            import openpyxl
        except ImportError:
            die("openpyxl is not installed.  Run:  pip install ezdxf openpyxl")
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            die(f"could not read the spreadsheet: {exc}")
        self.sheet_names = list(wb.sheetnames)
        if sheet is not None:
            if sheet not in wb.sheetnames:
                die(f"sheet '{sheet}' not found. Sheets: {', '.join(wb.sheetnames)}")
            ws = wb[sheet]
        else:
            ws = max(wb.worksheets, key=lambda w: (w.max_row or 0))
        self.sheet = ws.title
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        self._split_header(rows)

    def _split_header(self, rows):
        rows = [r for r in rows if r and any(c is not None and str(c).strip() != "" for c in r)]
        if not rows:
            die("the schedule file has no data.")
        # header = the row in the first 10 with the most filled cells
        best_i, best_score = 0, -1.0
        for i, r in enumerate(rows[:10]):
            filled = sum(1 for c in r if c is not None and str(c).strip() != "")
            alpha = any(re.search(r"[A-Za-z]", str(c or "")) for c in r)
            score = filled + (0.5 if alpha else 0)
            if filled >= 2 and score > best_score:
                best_i, best_score = i, score
        hdr = rows[best_i]
        self.headers = [str(h).strip() if h is not None and str(h).strip() else f"Column {i+1}"
                        for i, h in enumerate(hdr)]
        self.rows = [r for r in rows[best_i + 1:]]

    # ---- column resolution ---------------------------------------------
    def column(self, i):
        """Non-empty values of column i."""
        out = []
        for r in self.rows:
            if i < len(r):
                v = r[i]
                if v is not None and str(v).strip():
                    out.append(v)
        return out

    def resolve(self, spec, patterns, label="", flag=""):
        """spec: explicit column name / 1-based index / None -> match on header text.

        Returns (index, how) where how is 'given', 'header' or '' (not found).
        """
        if spec:
            for i, h in enumerate(self.headers):
                if h.strip().lower() == str(spec).strip().lower():
                    return i, "given"
            if str(spec).isdigit():
                i = int(spec) - 1
                if 0 <= i < len(self.headers):
                    return i, "given"
            die(f"column '{spec}' not found. Columns: {', '.join(self.headers)}\n"
                f"  Use {flag} with one of those names, or a 1-based column number.")
        for pat in patterns:
            for i, h in enumerate(self.headers):
                if re.search(pat, h, re.I):
                    return i, "header"
        return -1, ""

    def detect_date_column(self, style, exclude=()):
        """Pick the column whose values actually parse as dates."""
        best, best_score = -1, 0.0
        for i in range(len(self.headers)):
            if i in exclude:
                continue
            vals = self.column(i)[:40]
            if len(vals) < 2:
                continue
            ok = sum(1 for v in vals if coerce_date(v, style) is not None)
            score = ok / len(vals)
            if score > best_score and score >= 0.6:
                best, best_score = i, score
        return best

    def detect_id_column(self, drawing_keys, exclude=()):
        """Pick the column whose values are actually found in the drawing."""
        best, best_score, best_hits = -1, 0.0, 0
        for i in range(len(self.headers)):
            if i in exclude:
                continue
            vals = self.column(i)
            if not vals:
                continue
            keys = {mkey(v) for v in vals}
            hits = sum(1 for k in keys if k in drawing_keys)
            score = hits / max(1, len(keys))
            if score > best_score and hits > 0:
                best, best_score, best_hits = i, score, hits
        return best, best_hits


# --------------------------------------------------------------------------
# dates
# --------------------------------------------------------------------------

EXCEL_EPOCH = dt.date(1899, 12, 30)


def detect_date_style(values) -> str:
    d = m = y = 0
    for v in values:
        parts = re.split(r"[/\-.]", str(v).strip())
        if len(parts) != 3:
            continue
        if len(parts[0]) == 4:
            y += 1
            continue
        try:
            a, b = int(parts[0]), int(parts[1])
        except ValueError:
            continue
        if a > 12 >= b:
            d += 1
        elif b > 12 >= a:
            m += 1
    if y >= d and y >= m and y:
        return "ymd"
    if d > m:
        return "dmy"
    if m > d:
        return "mdy"
    return "mdy"


def coerce_date(value, style: str):
    """-> datetime.date or None"""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if 20000 < float(value) < 80000:            # Excel serial
            return EXCEL_EPOCH + dt.timedelta(days=int(round(float(value))))
        return None
    s = str(value).strip()
    m = re.match(r"^(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})", s)
    if m:
        return _mk(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.match(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})", s)
    if m:
        a, b, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yy < 100:
            yy += 2000
        if style == "dmy" or (a > 12 >= b):
            return _mk(yy, b, a)
        return _mk(yy, a, b)
    for f in ("%d %b %Y", "%d %B %Y", "%b %d %Y", "%B %d %Y", "%d-%b-%y", "%d-%b-%Y"):
        try:
            return dt.datetime.strptime(s.replace(",", ""), f).date()
        except ValueError:
            pass
    return None


def _mk(y, m, d):
    try:
        if not (1990 <= y <= 2100):
            return None
        return dt.date(y, m, d)
    except ValueError:
        return None


# --------------------------------------------------------------------------
# matching schedule rows to drawing blocks
# --------------------------------------------------------------------------

def choose_strategy(dwg: Drawing, keys, forced=None):
    """Return (label, lookup dict, kind) for the best way to find tools."""
    candidates = []
    for tag, _ in dwg.attrib_tags().most_common():
        idx = dwg.index_by_attrib(tag)
        candidates.append((f'block attribute "{tag}"', idx, "attrib", sum(1 for k in keys if k in idx)))
    idx = dwg.index_by_blockname()
    candidates.append(("block name", idx, "blockname", sum(1 for k in keys if k in idx)))
    idx = dwg.index_by_blocktext()
    candidates.append(("text inside the tool block", idx, "blocktext", sum(1 for k in keys if k in idx)))
    idx = dwg.index_by_text()
    candidates.append(("text label in drawing", idx, "text", sum(1 for k in keys if k in idx)))

    if forced and forced != "auto":
        want = forced.lower()
        for label, idx, kind, hits in candidates:
            if want == kind or (want.startswith("attrib:") and label == f'block attribute "{norm(want[7:])}"'):
                return label, idx, kind, hits
        opts = ", ".join(sorted({c[2] for c in candidates} | {f'attrib:{t}' for t in dwg.attrib_tags()}))
        die(f"--match '{forced}' is not available in this drawing. Options: {opts}")

    candidates.sort(key=lambda c: (-c[3], c[2] != "attrib"))
    return candidates[0]


def layer_wanted(name: str, only, exclude) -> bool:
    """Layer filter supporting simple * wildcards, case-insensitive."""
    def hit(pattern):
        rx = "^" + ".*".join(re.escape(p) for p in str(pattern).split("*")) + "$"
        return re.match(rx, name, re.I) is not None
    if only and not any(hit(p) for p in only):
        return False
    return not any(hit(p) for p in exclude)


class InsGrid:
    """Uniform grid over insert bboxes: point -> smallest enclosing insert."""

    def __init__(self, inserts):
        boxed = [(ins, ins["bbox"]) for ins in inserts if ins["bbox"]]
        dims = sorted(max(b[2] - b[0], b[3] - b[1]) for _i, b in boxed) if boxed else []
        med = dims[len(dims) // 2] if dims else 1.0
        self.cell = max(med * 2.0, 1e-6)
        self.map = defaultdict(list)
        self.big = []
        for ins, b in boxed:
            c0, r0 = int(b[0] // self.cell), int(b[1] // self.cell)
            c1, r1 = int(b[2] // self.cell), int(b[3] // self.cell)
            if (c1 - c0 + 1) * (r1 - r0 + 1) > 4096:
                self.big.append(ins)
                continue
            for cx in range(c0, c1 + 1):
                for cy in range(r0, r1 + 1):
                    self.map[(cx, cy)].append(ins)

    def enclosing(self, x, y):
        best, best_area = None, math.inf
        cands = self.map.get((int(x // self.cell), int(y // self.cell)), [])
        for ins in list(cands) + self.big:
            b = ins["bbox"]
            if b[0] <= x <= b[2] and b[1] <= y <= b[3]:
                area = (b[2] - b[0]) * (b[3] - b[1])
                if area < best_area:
                    best, best_area = ins, area
        return best


# --------------------------------------------------------------------------
# build the payload for the viewer
# --------------------------------------------------------------------------

def hit_centre(h):
    """Drawing-space centre of a placement (block insert or text label)."""
    if "pt" in h:
        b = h.get("bbox")
        if b:
            return (b[0] + b[2]) / 2.0, (b[1] + b[3]) / 2.0
        return h["pt"]
    return h["x"], h["y"]


PREFER_DIRS = {"north": (lambda x, y: (-y, x), "north first (highest Y in the drawing)"),
               "south": (lambda x, y: (y, x), "south first (lowest Y in the drawing)"),
               "west": (lambda x, y: (x, y), "west first (lowest X in the drawing)"),
               "east": (lambda x, y: (-x, y), "east first (highest X in the drawing)")}
PREFER_ALIASES = {"n": "north", "top": "north", "upper": "north", "up": "north",
                  "s": "south", "bottom": "south", "lower": "south", "down": "south",
                  "w": "west", "left": "west", "e": "east", "right": "east"}


def parse_prefer(spec):
    """--prefer: where the scheduled tools are, when a label is placed more
    than once. Returns None or {"key": sort key over a centre (x, y),
    "inside": point test or None, "desc": text} -- placements are consumed
    in that order in type mode."""
    if not spec or not str(spec).strip():
        return None
    s = PREFER_ALIASES.get(str(spec).strip().lower(), str(spec).strip().lower())
    if s in PREFER_DIRS:
        key, desc = PREFER_DIRS[s]
        return {"key": key, "inside": None, "desc": desc}
    try:
        vals = [float(v) for v in re.split(r"[,;\s]+", s) if v]
    except ValueError:
        vals = []
    if len(vals) != 4:
        die(f"--prefer '{spec}' not understood. Use north | south | east | west, or a\n"
            "  rectangle in drawing units: --prefer X0,Y0,X1,Y1  (AutoCAD ID command\n"
            "  shows coordinates; --inspect prints where the block inserts are).")
    x0, x1 = sorted((vals[0], vals[2]))
    y0, y1 = sorted((vals[1], vals[3]))
    inside = lambda x, y: x0 <= x <= x1 and y0 <= y <= y1
    return {"key": lambda x, y: (0 if inside(x, y) else 1, x, y), "inside": inside,
            "desc": f"inside ({x0:,.0f}, {y0:,.0f})-({x1:,.0f}, {y1:,.0f}) first"}


def placement_policy(args):
    """How a row picks among several placements of its label (type mode):
    --tool-layer placements first, then the --prefer order, then west to
    east / south to north. None when neither option is given."""
    prefer = parse_prefer(getattr(args, "prefer", None))
    layers = [l for l in (getattr(args, "tool_layer", None) or []) if str(l).strip()]
    if not prefer and not layers:
        return None
    on_layer = (lambda h: layer_wanted(h["layer"], layers, [])) if layers else None
    pos = prefer["key"] if prefer else (lambda x, y: (x, y))

    def key(h):
        x, y = hit_centre(h)
        return ((0 if on_layer(h) else 1) if on_layer else 0,) + tuple(pos(x, y))

    def misses(h):
        """Why this placement is not where the options say the tools are."""
        out = []
        if on_layer and not on_layer(h):
            out.append("not on --tool-layer")
        if prefer and prefer["inside"] and not prefer["inside"](*hit_centre(h)):
            out.append("outside the --prefer rectangle")
        return out

    parts = []
    if layers:
        parts.append("on layer " + " / ".join(f"'{l}'" for l in layers) + " first")
    if prefer:
        parts.append(prefer["desc"])
    return {"key": key, "misses": misses, "desc": ", then ".join(parts),
            "restricts": bool(layers or (prefer and prefer["inside"]))}


def _box_geom(b):
    g = Geom()
    g.polys.append([(b[0], b[1]), (b[2], b[1]), (b[2], b[3]), (b[0], b[3]), (b[0], b[1])])
    return g


def trace_label(dwg: Drawing, rows, strategy, label, assign_log, by_type, policy=None):
    """--trace: explain end-to-end what happened to one schedule label."""
    lab, index, kind, _ = strategy
    key = mkey(label)
    rkey = (lambda r: r["tkey"]) if by_type else (lambda r: r["key"])

    def where(h):
        x, y = hit_centre(h)
        s = f"({x:,.1f}, {y:,.1f}) on '{h.get('layer', '?')}'"
        if policy and policy["restricts"]:
            s += " OUT" if policy["misses"](h) else " in"
        return s

    print("\n" + "─" * 68)
    print(f"TRACE  {label}")
    print("─" * 68)

    srows = [r for r in rows if rkey(r) == key]
    if srows:
        print(f"schedule : {len(srows)} usable row(s): "
              + ", ".join(f"{r['id']} ({r['day'].isoformat()})" for r in srows[:8])
              + (" …" if len(srows) > 8 else ""))
    else:
        other = [r for r in rows if key in (r["key"], r["tkey"])]
        print(f"schedule : no usable row carries this label in the "
              f"{'type' if by_type else 'ID'} column.")
        if other:
            print(f"           (it does appear in the {'ID' if by_type else 'type'} column "
                  f"of {len(other)} row(s))")
        print("           Check the warnings above for an unreadable date, a blank ID or")
        print("           a duplicate row, and the exact spelling in the spreadsheet.")

    sources = []
    for tag in dwg.attrib_tags():
        hits = dwg.index_by_attrib(tag).get(key, [])
        if hits:
            sources.append((f'block attribute "{tag}"', hits))
    for nm, idx in (("block name", dwg.index_by_blockname()),
                    ("text inside the tool block", dwg.index_by_blocktext()),
                    ("text label in drawing", dwg.index_by_text())):
        hits = idx.get(key, [])
        if hits:
            sources.append((nm, hits))
    if sources:
        for nm, hits in sources:
            print(f"drawing  : {nm}: {len(hits)} placement(s) at "
                  + ", ".join(where(h) for h in hits[:6]) + (" …" if len(hits) > 6 else ""))
        if not any(nm == lab for nm, _h in sources):
            print(f"           NOTE: matching used '{lab}', which does not carry this label;")
            print("           force the option that does with --match (attrib:TAG | blockname"
                  " | blocktext | text)")
    else:
        print("drawing  : NOT found under any option (attribute, block name, text inside a")
        print("           block, loose text label) with this exact spelling.")
        rel = Counter()
        for nm, idx in (("text inside block", dwg.index_by_blocktext()),
                        ("text label", dwg.index_by_text())):
            for k2, lst in idx.items():
                if k2 != key and len(k2) >= 3 and (k2 in key or key in k2):
                    rel[(k2, nm)] += len(lst)
        ck = compact(label)
        for nm, idx in (("text inside block", dwg.index_by_blocktext()),
                        ("text label", dwg.index_by_text())):
            for k2, lst in idx.items():
                if k2 != key and compact(k2) == ck:
                    rel[(k2, nm)] += len(lst)
        if rel:
            print("           similar labels that DO exist: "
                  + ", ".join(f"'{k2}' ×{n} ({nm})" for (k2, nm), n in rel.most_common(8)))
            print("           -> a punctuation/spacing difference is fixed by --loose; a label")
            print("              drawn as two texts (e.g. 'DT_545' + 'R') needs the schedule")
            print("              value changed to the part that is actually one text.")

    print(f"matching : {lab}" + ("  (rows take free placements of their type, in date order)"
                                 if by_type else ""))
    if by_type and policy:
        print(f"order    : {policy['desc']}")
    elif by_type and len(index.get(key, [])) > 1:
        print("order    : west to east, then south to north  (change with --tool-layer NAME"
              " and/or --prefer north|south|east|west|X0,Y0,X1,Y1)")
    got = [(r, h) for r, h in assign_log if rkey(r) == key]
    for r, h in got:
        note = ""
        if "pt" in h:
            _g, b = dwg.insert_wcs_geom(h)
            note = (f", footprint {b[2] - b[0]:,.1f} × {b[3] - b[1]:,.1f}" if b
                    else ", NO drawable geometry -> drawn as a placeholder box")
        print(f"assigned : {r['id']} ({r['day'].isoformat()}) -> placement at {where(h)}{note}")
    if srows and not got:
        print("assigned : nothing -- the row(s) ended up unmatched"
              + (" (no free placement of this type was left)" if by_type and sources else ""))
    if by_type:
        used_ids = {id(h) for _r, h in got}
        free = [h for h in index.get(key, []) if id(h) not in used_ids]
        if free:
            print(f"free     : {len(free)} placement(s) of this label got no schedule row "
                  f"and stay plain background: " + ", ".join(where(h) for h in free[:6])
                  + (" …" if len(free) > 6 else ""))
    print("viewer   : type the label into the search box and press Enter to zoom to it.")


def match_rows(dwg: Drawing, rows, strategy, args):
    """Assign schedule rows to drawing placements.

    ID mode: every placement carrying a row's ID is drawn for that row.
    Type mode (--type-col): the drawing labels tool TYPES, so each row takes
    one still-free placement of its type, in date order.
    """
    label, index, kind, _ = strategy
    matched, unmatched = [], []          # matched: (row, geom WCS, bbox WCS, insert|None)
    used_inserts, used_text_keys = set(), set()
    wcs_cache, assign_log = {}, []
    grid = InsGrid(dwg.inserts) if kind == "text" else None

    def realized(ins):
        key = id(ins)
        if key not in wcs_cache:
            wcs_cache[key] = dwg.insert_wcs_geom(ins)
        return wcs_cache[key]

    def add_hit(row, hit):
        assign_log.append((row, hit))
        if kind == "text":
            used_text_keys.add((round(hit["x"], 3), round(hit["y"], 3), norm(hit["str"])))
            ins = grid.enclosing(hit["x"], hit["y"])
            if ins is not None:
                used_inserts.add(id(ins))
                g, b = realized(ins)
                matched.append((row, g, b, ins))
            else:
                h = max(hit["h"], 1e-6)
                w = max(len(hit["str"]) * h * 0.7, h * 2)
                b = (hit["x"] - h * 0.2, hit["y"] - h * 0.4, hit["x"] + w, hit["y"] + h * 1.4)
                matched.append((row, _box_geom(b), b, None))
        else:
            used_inserts.add(id(hit))
            g, b = realized(hit)
            matched.append((row, g, b, hit))

    by_type = bool(getattr(args, "type_col", None))
    policy = placement_policy(args)
    if by_type:
        # a label placed several times: rows take its placements in a fixed
        # positional order (west to east, south to north), or in the order
        # --tool-layer / --prefer ask for (e.g. the tool layer of this
        # project, north first, when the schedule covers the top half of the
        # fab and the same label also exists at the bottom)
        order = policy["key"] if policy else (lambda h2: hit_centre(h2))
        hit_order, cursors = {}, defaultdict(int)
        sched_type, short_type, none_type = Counter(), Counter(), Counter()
        outside = []
        for row in sorted(rows, key=lambda r: (r["day"], r["id"])):
            key = row["tkey"]
            if not key:
                unmatched.append(row["id"])
                continue
            sched_type[key] += 1
            if key not in hit_order:
                lst = list(index.get(key, []))
                lst.sort(key=order)
                hit_order[key] = lst
            lst = hit_order[key]
            c = cursors[key]
            if c >= len(lst):
                unmatched.append(row["id"])
                (short_type if lst else none_type)[row["type"] or key] += 1
                continue
            cursors[key] = c + 1
            add_hit(row, lst[c])
            if policy and policy["misses"](lst[c]):
                outside.append(f"{row['id']} ({'; '.join(policy['misses'](lst[c]))})")
        if outside:
            warn(f"{len(outside)} row(s) got a placement that is NOT where --tool-layer/--prefer "
                 f"say the tools are, because no free placement of their label was left there: "
                 f"{', '.join(outside[:6])}{' …' if len(outside) > 6 else ''}")
        for t2, n2 in none_type.most_common(6):
            warn(f'type "{t2}": {n2} scheduled tool(s), but no placement with that label in the drawing')
        for t2, _n2 in short_type.most_common(6):
            k2 = norm(t2)
            warn(f'type "{t2}": {sched_type[k2]} scheduled tool(s), only '
                 f'{len(hit_order.get(k2, []))} placement(s) in the drawing')
    else:
        for row in rows:
            hits = index.get(row["key"])
            if not hits:
                unmatched.append(row["id"])
                continue
            for hit in hits:
                add_hit(row, hit)

    for lab in (getattr(args, "trace", None) or []):
        trace_label(dwg, rows, strategy, lab, assign_log, by_type, policy)

    # per layer: placements that became scheduled tools / free slots of a
    # scheduled label / unrelated blocks -- for the --layers report
    match_keys = {(r["tkey"] if by_type else r["key"]) for r in rows}
    sched_ins = set()
    if kind != "text":
        for k in match_keys:
            for h in index.get(k, []):
                sched_ins.add(id(h))
    matched_ins = {id(i) for _r, _g, _b, i in matched if i is not None}
    layer_tools = {}
    for ins in dwg.inserts:
        d = layer_tools.setdefault(ins["layer"], [0, 0, 0])
        if id(ins) in matched_ins:
            d[0] += 1
        elif id(ins) in sched_ins:
            d[1] += 1
        else:
            d[2] += 1

    if not matched:
        die(f"no schedule IDs matched the drawing (tried: {label}).\n"
            f"  Run with --inspect to see how the drawing identifies its tools,\n"
            f"  then pass e.g. --match attrib:TOOL_ID  or  --match blockname.")
    return {"matched": matched, "unmatched": unmatched, "used_inserts": used_inserts,
            "used_text_keys": used_text_keys, "layer_tools": layer_tools,
            "by_type": by_type, "kind": kind, "label": label}


def build_tools(matched, by_type):
    """Tool records for the viewer plus de-duplicated footprint symbols.
    Returns (tools, symbols, typical footprint size)."""
    sizes = sorted(max(b[2] - b[0], b[3] - b[1]) for _r, _g, b, _i in matched
                   if b and (b[2] - b[0] > 0 or b[3] - b[1] > 0))
    typical = sizes[len(sizes) // 2] if sizes else 5.0

    symbols, sym_index, tools, fallback_n = [], {}, [], 0
    for row, geom, b, ins in matched:
        if not b or ((b[2] - b[0]) <= 0 and (b[3] - b[1]) <= 0):
            # no drawable geometry (proxy object / unresolved xref): place a
            # typical-size box at the block's insertion point -- NOT at (0,0).
            cx, cy = (ins["pt"] if ins is not None
                      else (((b[0] + b[2]) / 2.0, (b[1] + b[3]) / 2.0) if b else (0.0, 0.0)))
            half = typical / 2.0
            b = (cx - half, cy - half, cx + half, cy + half)
            geom = _box_geom(b)
            if ins is not None:
                fallback_n += 1
        x0, y0, x1, y1 = b
        # local path: translate so the footprint starts at (0,0) in SVG space
        d = path_data(geom.polys, flip_y=True, dx=x0, dy=-y1)
        sig = hashlib.blake2b(d.encode(), digest_size=12).hexdigest()
        si = sym_index.get(sig)
        if si is None:
            si = len(symbols)
            sym_index[sig] = si
            symbols.append(d)
        t = {"id": row["id"], "day": row["day"].isoformat(), "s": si,
             "x": Prec.r(x0), "y": Prec.r(-y1),
             "w": Prec.r(x1 - x0), "h": Prec.r(y1 - y0),
             # private (stripped before output): conflict detection + storage
             "_hull": hull_svg(geom, (x0, -y1, x1, -y0)), "_d": row["day"],
             "_ins": id(ins) if ins is not None else None,
             "_arrive": row.get("arrive"), "_store": row.get("store", "")}
        if by_type and row.get("type"):
            t["ty"] = row["type"]
        if row["name"]:
            t["nm"] = row["name"]
        if row["group"]:
            t["g"] = row["group"]
        if row["area"]:
            t["a"] = row["area"]
        for j, v in enumerate(row.get("facets", ())):
            if v:
                t[f"f{j}"] = v
        tools.append(t)
    tools.sort(key=lambda t: (t["day"], t["id"]))

    if fallback_n:
        warn(f"{fallback_n} matched tool(s) had no drawable geometry in the DXF -- "
             f"drawn as placeholder boxes at their block insertion points.")
        warn("  (typical cause: proxy objects. Re-export the DXF with PROXYGRAPHICS=1,")
        warn("   or explode the tool blocks before SAVEAS, to get real footprints.)")
    return tools, symbols, typical


def collect_background(dwg: Drawing, used_inserts, used_text_keys, kind, args):
    """Everything that is not a scheduled tool, per drawing layer.

    Blocks placed <= 2 times (bound xrefs, one-off details) are inlined as
    ordinary layer geometry; blocks placed more often become one shared
    symbol each (deduplicated by geometry) referenced by 7-number placements.
    Returns (layers, pieces, use_items, text_refs, bsyms, inlined_blocks).
    """
    bg_uses = defaultdict(list)      # layer -> [(block name, matrix)]
    bg_inline = defaultdict(Geom)    # layer -> geometry inlined into the layer
    bg_texts = defaultdict(list)     # layer -> [(x, y, h, rot, s)] in WCS
    use_count = Counter(ins["name"] for ins in dwg.inserts
                        if id(ins) not in used_inserts and ins["m"] is not None)
    inlined_blocks = 0
    for ins in dwg.inserts:
        if id(ins) in used_inserts:
            continue
        if ins["m"] is None:                     # MINSERT array: already exploded
            if ins["geom"] is not None:
                bg_inline[ins["layer"]].extend(ins["geom"])
            continue
        bd = dwg.blockdef(ins["name"])
        if bd["geom"].polys:
            if use_count[ins["name"]] <= 2:
                g_w, _b = dwg.insert_wcs_geom(ins)
                bg_inline[ins["layer"]].extend(g_w)
                inlined_blocks += 1
            else:
                bg_uses[ins["layer"]].append((ins["name"], ins["m"]))
        if bd["geom"].texts:
            M = ins["m"]
            s_avg, rot0 = mat_avg_scale(M), mat_rotation_deg(M)
            for (tx, ty, th, trot, ts) in bd["geom"].texts:
                px, py = mat_apply(M, tx, ty)
                bg_texts[ins["layer"]].append((px, py, th * s_avg, trot + rot0, ts))

    bsyms, bsym_index, bsym_sig = [], {}, {}

    def bsym_for(name):
        si = bsym_index.get(name)
        if si is None:
            bb = dwg.blockdef(name)["bbox"]
            dgs = digits_for(max(bb[2] - bb[0], bb[3] - bb[1]) if bb else 1.0)
            d = path_data(dwg.block_lite(name).polys, flip_y=True, digits=min(6, dgs))
            sig = hashlib.blake2b(d.encode(), digest_size=12).hexdigest()
            si = bsym_sig.get(sig)            # identical geometry under another name
            if si is None:
                si = len(bsyms)
                bsym_sig[sig] = si
                bsyms.append(d)
            bsym_index[name] = si
        return si

    strong = re.compile(r"wall|outline|border|shell|boundary|perimeter", re.I)
    names = sorted(set(dwg.layers) | set(bg_uses) | set(bg_inline) | set(bg_texts))
    names = [n for n in names if layer_wanted(n, args.only_layer, args.exclude_layer)]

    layers, pieces, use_items, text_refs = [], [], [], []
    for name in names:
        g = Geom()
        if name in dwg.layers:
            g.extend(dwg.layers[name]["geom"])
        if name in bg_inline:
            g.extend(bg_inline[name])

        li = len(layers)
        seen_d = set()
        for poly in g.polys:
            d1 = path_data([poly])
            if not d1:
                continue
            if len(d1) <= 20000:              # dedupe identical small pieces (xref overlays)
                if d1 in seen_d:
                    continue
                seen_d.add(d1)
            pieces.append((li, d1, poly_bbox_svg(poly)))

        n_uses = 0
        for bname, M in bg_uses.get(name, ()):  # SVG matrix = flip . M . flip
            bb = dwg.blockdef(bname)["bbox"]
            if not bb:
                continue
            si = bsym_for(bname)
            a, b_, c, dd, e, f = M
            entry = [si, round(a, 6), round(-b_, 6), round(-c, 6), round(dd, 6),
                     Prec.r(e), Prec.r(-f)]
            w = bbox_transform(M, bb)
            use_items.append((li, entry, (w[0], -w[3], w[2], -w[1])))
            n_uses += 1

        for (x, y, h, rot, s) in g.texts:
            if kind == "text" and (round(x, 3), round(y, 3), norm(s)) in used_text_keys:
                continue
            text_refs.append((li, (x, y, h, rot, s)))
        for t in bg_texts.get(name, ()):
            text_refs.append((li, t))

        layers.append({"n": name, "tiles": [], "t": [],
                       "c": dwg.layers.get(name, {}).get("count", 0) + n_uses,
                       "strong": bool(strong.search(name))})
    return layers, pieces, use_items, text_refs, bsyms, inlined_blocks


def cap_texts(layers, text_refs, max_texts):
    """Keep the largest labels (deduplicated) and attach them to their layers."""
    seen_t, deduped = set(), []
    for li, t in text_refs:
        key = (round(t[0], 2), round(t[1], 2), t[4])
        if key in seen_t:
            continue
        seen_t.add(key)
        deduped.append((li, t))
    deduped.sort(key=lambda r: -r[1][2])
    kept = deduped[:max(0, max_texts)]
    for li, (x, y, h, rot, s) in kept:
        layers[li]["t"].append([Prec.r(x), Prec.r(-y), Prec.r(h), Prec.r(rot), s])
    if len(text_refs) > len(kept):
        warn(f"kept the {len(kept):,} largest of {len(text_refs):,} text labels "
             f"(--max-texts {max_texts} — raise it if labels are missing)")


def compute_bounds(tools, typical, pieces, use_items):
    """View bounds around the tools; far-away stray geometry is ignored."""
    txs0 = min(t["x"] for t in tools);            tys0 = min(t["y"] for t in tools)
    txs1 = max(t["x"] + t["w"] for t in tools);   tys1 = max(t["y"] + t["h"] for t in tools)
    diag = math.hypot(txs1 - txs0, tys1 - tys0)
    infl = max(3.0 * diag, 40.0 * typical, 10.0)
    rg = (txs0 - infl, tys0 - infl, txs1 + infl, tys1 + infl)
    bounds = [txs0, tys0, txs1, tys1]
    clipped = 0
    for b in [p[2] for p in pieces] + [u[2] for u in use_items]:
        if b[2] < rg[0] or b[0] > rg[2] or b[3] < rg[1] or b[1] > rg[3]:
            clipped += 1
            continue
        bounds[0] = min(bounds[0], b[0]); bounds[1] = min(bounds[1], b[1])
        bounds[2] = max(bounds[2], b[2]); bounds[3] = max(bounds[3], b[3])
    if clipped:
        print(f"  initial view fitted to the tool area "
              f"({clipped:,} far-away drawing pieces left out of the auto-fit)")
    return bounds


def tile_layers(layers, pieces, use_items, bounds, bsyms, fine_unit, n=8):
    """Spatial tiles + level of detail.

    Each piece lands in one of n×n tiles over the fitted bounds (by centre),
    or in the 'global' tile (-1) when larger than a cell or outside. Pieces
    smaller than fine_unit are 'fine detail' the viewer hides until zoomed in.
    Each layer also gets 'sw' = its share of the shared-symbol bytes it uses.
    Returns the non-empty layers.
    """
    bx0, by0, bx1, by1 = bounds
    cw, ch = max((bx1 - bx0) / n, 1e-9), max((by1 - by0) / n, 1e-9)

    def tile_of(b):
        if (b[2] - b[0] > cw or b[3] - b[1] > ch or b[0] < bx0 or b[1] < by0
                or b[2] > bx1 or b[3] > by1):
            return -1
        i = min(n - 1, int((0.5 * (b[0] + b[2]) - bx0) / cw))
        j = min(n - 1, int((0.5 * (b[1] + b[3]) - by0) / ch))
        return j * n + i

    def is_fine(b):
        return max(b[2] - b[0], b[3] - b[1]) < fine_unit

    store = [dict() for _ in layers]

    def bucket(li, b):
        ti = tile_of(b)
        T = store[li].get(ti)
        if T is None:
            T = store[li][ti] = {"bb": list(b), "c": [], "f": [], "u": [], "uf": []}
        else:
            bb = T["bb"]
            bb[0] = min(bb[0], b[0]); bb[1] = min(bb[1], b[1])
            bb[2] = max(bb[2], b[2]); bb[3] = max(bb[3], b[3])
        return T

    for li, d1, b in pieces:
        bucket(li, b)["f" if is_fine(b) else "c"].append(d1)
    for li, entry, b in use_items:
        bucket(li, b)["uf" if is_fine(b) else "u"].append(entry)
    for li, L in enumerate(layers):
        for ti in sorted(store[li]):
            T = store[li][ti]
            L["tiles"].append([ti, [Prec.r(v) for v in T["bb"]],
                               "".join(T["c"]), "".join(T["f"]), T["u"], T["uf"]])

    sym_uses = Counter(e[0] for _li, e, _b in use_items)
    for li, entry, _b in use_items:
        layers[li]["sw"] = layers[li].get("sw", 0.0) + len(bsyms[entry[0]]) / sym_uses[entry[0]]
    for L in layers:
        L["sw"] = int(L.get("sw", 0.0))
    return [L for L in layers if L["tiles"] or L["t"]]


def field_list(rows, by_type, args):
    """Ordered schedule fields the viewer can filter and colour by, and the
    key it colours by at start: --color-by, else a Module column, else the
    group column, else the area column."""
    r0 = rows[0] if rows else {}
    fields = []
    if by_type and r0.get("type_label"):
        fields.append({"k": "ty", "l": r0["type_label"]})
    if r0.get("group_label"):
        fields.append({"k": "g", "l": r0["group_label"]})
    if r0.get("area_label"):
        fields.append({"k": "a", "l": r0["area_label"]})
    for j, lab in enumerate(r0.get("facet_labels", ())):
        fields.append({"k": f"f{j}", "l": lab})
    color_by = ""
    if getattr(args, "color_by", None):
        want = str(args.color_by).strip().lower()
        hit = [f for f in fields if f["l"].strip().lower() == want]
        if not hit:
            die(f"--color-by '{args.color_by}' is not one of the viewer fields: "
                + ", ".join(f["l"] for f in fields) + "\n  (add it with --facet-col NAME)")
        color_by = hit[0]["k"]
    else:
        for pick in (lambda f: re.search(r"module", f["l"], re.I), lambda f: f["k"] == "g",
                     lambda f: f["k"] == "a"):
            hit = [f for f in fields if pick(f)]
            if hit:
                color_by = hit[0]["k"]
                break
    return fields, color_by


def day_axis(tools, args):
    days_set = sorted({t["day"] for t in tools} | {t["ad"] for t in tools if t.get("ad")})
    first = dt.date.fromisoformat(days_set[0]) - dt.timedelta(days=1)
    last = dt.date.fromisoformat(days_set[-1])
    if args.start_date:
        first = dt.date.fromisoformat(args.start_date)
    if args.end_date:
        last = dt.date.fromisoformat(args.end_date)
    span = (last - first).days + 1
    if span > 4000:
        die(f"the schedule spans {span:,} days -- that looks like a date-parsing problem.\n"
            f"  First {days_set[0]}, last {days_set[-1]}. Try --date-format dmy (or mdy/ymd).")
    return [(first + dt.timedelta(days=i)).isoformat() for i in range(span)]


# --------------------------------------------------------------------------
# move-in conflicts
# --------------------------------------------------------------------------
# Everything here works in viewer (SVG) coordinates: x as in the drawing,
# y negated. Tool records carry x, y, w, h (footprint box), _hull (convex
# outline), _d (move-in date), _ins (drawing placement identity).

CONFLICT_SEV = {"overlap": "high", "storage": "medium", "access": "medium",
                "capacity": "medium", "crowding": "low"}
CONFLICT_TITLE = {"overlap": "Footprints overlap", "crowding": "Crowded move-ins",
                  "access": "Boxed in", "capacity": "Too many move-ins in one day",
                  "storage": "No storage space"}


def hull_svg(geom, box):
    """Convex outline of a matched tool in SVG coordinates (y down)."""
    pts = [(x, -y) for poly in (geom.polys if geom else ()) for x, y in poly]
    h = convex_hull(pts) if len(pts) >= 3 else []
    if len(h) < 3:
        x0, y0, x1, y1 = box
        return [(x0, y0), (x1, y0), (x1, y1), (x0, y1)]
    return h


def poly_area(poly):
    a = 0.0
    for i, (x0, y0) in enumerate(poly):
        x1, y1 = poly[(i + 1) % len(poly)]
        a += x0 * y1 - x1 * y0
    return a / 2.0


def convex_overlap_area(subject, clip):
    """Area shared by two convex polygons (Sutherland-Hodgman clipping)."""
    if poly_area(clip) < 0:
        clip = clip[::-1]
    out = list(subject)
    for i in range(len(clip)):
        ax, ay = clip[i]
        bx, by = clip[(i + 1) % len(clip)]
        inp, out = out, []
        if not inp:
            break
        side = lambda p: (bx - ax) * (p[1] - ay) - (by - ay) * (p[0] - ax)
        s = inp[-1]
        for e in inp:
            se, ss = side(e), side(s)
            if se >= 0:
                if ss < 0:
                    t = ss / (ss - se)
                    out.append((s[0] + (e[0] - s[0]) * t, s[1] + (e[1] - s[1]) * t))
                out.append(e)
            elif ss >= 0:
                t = ss / (ss - se)
                out.append((s[0] + (e[0] - s[0]) * t, s[1] + (e[1] - s[1]) * t))
            s = e
    return abs(poly_area(out)) if len(out) >= 3 else 0.0


def box_gap(a, b):
    """Edge-to-edge distance between two boxes (0 when they touch/overlap)."""
    dx = max(0.0, max(a[0] - b[2], b[0] - a[2]))
    dy = max(0.0, max(a[1] - b[3], b[1] - a[3]))
    return math.hypot(dx, dy)


class ToolGrid:
    """Uniform grid over tool boxes for neighbour queries."""

    def __init__(self, tools, cell):
        self.cell = max(cell, 1e-6)
        self.map = defaultdict(list)
        for i, t in enumerate(tools):
            for key in self.cells(t["_box"]):
                self.map[key].append(i)

    def cells(self, b):
        c = self.cell
        for cx in range(int(b[0] // c), int(b[2] // c) + 1):
            for cy in range(int(b[1] // c), int(b[3] // c) + 1):
                yield (cx, cy)

    def near(self, b, margin):
        seen = set()
        bb = (b[0] - margin, b[1] - margin, b[2] + margin, b[3] + margin)
        for key in self.cells(bb):
            for i in self.map.get(key, ()):
                if i not in seen:
                    seen.add(i)
                    yield i


def _centre(t):
    return t["x"] + t["w"] / 2.0, t["y"] + t["h"] / 2.0


def _conflict(kind, tools, day, x, y, detail):
    return {"k": kind, "s": CONFLICT_SEV[kind], "ids": [t["id"] for t in tools],
            "day": day.isoformat() if hasattr(day, "isoformat") else day,
            "x": None if x is None else Prec.r(x), "y": None if y is None else Prec.r(y),
            "d": detail}


def find_conflicts(tools, typical, args):
    """Potential move-in conflicts between the placed tools.

    overlap   two tools drawn on the same floor space (or the same block)
    crowding  tools moving in within --conflict-window days of each other
              and closer than --conflict-gap (rigging / hook-up crowding)
    access    a tool whose neighbours on all four sides are in place before
              it arrives, with no gap wide enough to bring it in
    capacity  more than --max-per-day tools on one day
    """
    if getattr(args, "no_conflicts", False) or not tools:
        return []
    window = max(0, int(getattr(args, "conflict_window", 3) or 0))
    gap = getattr(args, "conflict_gap", None)
    gap = typical if gap is None else max(0.0, float(gap))
    for t in tools:
        t["_box"] = (t["x"], t["y"], t["x"] + t["w"], t["y"] + t["h"])
    grid = ToolGrid(tools, 2.0 * typical + gap)
    out, overlapped = [], set()

    # --- same drawing placement taken by two different schedule IDs
    by_ins = defaultdict(list)
    for t in tools:
        if t.get("_ins") is not None:
            by_ins[t["_ins"]].append(t)
    for lst in by_ins.values():
        ids = sorted({t["id"] for t in lst})
        if len(ids) > 1:
            first = [next(t for t in lst if t["id"] == i) for i in ids]
            cx, cy = _centre(first[0])
            out.append(_conflict("overlap", first, max(t["_d"] for t in first), cx, cy,
                                 "both rows point at the same block in the drawing"))
            for a in first:
                for b in first:
                    overlapped.add((min(a["id"], b["id"]), max(a["id"], b["id"])))

    # --- footprint overlap / crowding candidates (pairs of nearby tools)
    parent = {}

    def find(i):
        while parent.get(i, i) != i:
            parent[i] = parent.get(parent[i], parent[i])
            i = parent[i]
        return i

    def union(i, j):
        parent.setdefault(i, i); parent.setdefault(j, j)
        ri, rj = find(i), find(j)
        if ri != rj:
            parent[ri] = rj

    crowd_gap = gap
    for i, a in enumerate(tools):
        for j in grid.near(a["_box"], crowd_gap):
            if j <= i:
                continue
            b = tools[j]
            if a["id"] == b["id"]:
                continue                       # one ID drawn at several placements
            pair = (min(a["id"], b["id"]), max(a["id"], b["id"]))
            if pair in overlapped:
                continue
            d = box_gap(a["_box"], b["_box"])
            if d > crowd_gap:
                continue
            if d == 0.0:
                inter = convex_overlap_area(a["_hull"], b["_hull"])
                small = min(abs(poly_area(a["_hull"])), abs(poly_area(b["_hull"])), )
                if small > 0 and inter / small > 0.02:
                    overlapped.add(pair)
                    ax, ay = _centre(a); bx, by = _centre(b)
                    later, earlier = (a, b) if a["_d"] >= b["_d"] else (b, a)
                    out.append(_conflict("overlap", [earlier, later], later["_d"],
                                         (ax + bx) / 2, (ay + by) / 2,
                                         f"footprints overlap by {inter / small:.0%} of the smaller tool"))
                    continue
            if crowd_gap > 0 and abs((a["_d"] - b["_d"]).days) <= window:
                union(i, j)

    # --- crowding clusters
    groups = defaultdict(list)
    for i in range(len(tools)):
        if i in parent:
            groups[find(i)].append(tools[i])
    for lst in groups.values():
        if len(lst) < 2:
            continue
        lst.sort(key=lambda t: (t["_d"], t["id"]))
        d0, d1 = lst[0]["_d"], lst[-1]["_d"]
        cx = sum(_centre(t)[0] for t in lst) / len(lst)
        cy = sum(_centre(t)[1] for t in lst) / len(lst)
        span = (d1 - d0).days
        when = "on the same day" if span == 0 else f"within {span} day(s)"
        out.append(_conflict("crowding", lst, d0, cx, cy,
                             f"{len(lst)} tools move in {when} next to each other"))

    # --- boxed in: neighbours on every side are in place before this tool
    for i, b in enumerate(tools):
        bw, bh = b["w"], b["h"]
        if bw <= 0 or bh <= 0:
            continue
        agap = 0.6 * min(bw, bh)
        eps = 0.05 * min(bw, bh)
        bx0, by0, bx1, by1 = b["_box"]
        sides = {"E": None, "W": None, "N": None, "S": None}
        for j in grid.near(b["_box"], agap):
            a = tools[j]
            if a is b or a["id"] == b["id"] or a["_d"] >= b["_d"]:
                continue
            ax0, ay0, ax1, ay1 = a["_box"]
            yov = min(ay1, by1) - max(ay0, by0)
            xov = min(ax1, bx1) - max(ax0, bx0)
            if yov >= 0.3 * bh:
                if -eps <= ax0 - bx1 <= agap and sides["E"] is None:
                    sides["E"] = a
                if -eps <= bx0 - ax1 <= agap and sides["W"] is None:
                    sides["W"] = a
            if xov >= 0.3 * bw:
                if -eps <= ay0 - by1 <= agap and sides["S"] is None:   # SVG y down: larger y = south
                    sides["S"] = a
                if -eps <= by0 - ay1 <= agap and sides["N"] is None:
                    sides["N"] = a
        if all(sides.values()):
            cx, cy = _centre(b)
            around = ", ".join(f"{k}: {sides[k]['id']} ({sides[k]['day']})" for k in ("N", "E", "S", "W"))
            out.append(_conflict("access", [b] + [sides[k] for k in ("N", "E", "S", "W")], b["_d"],
                                 cx, cy, f"{b['id']} arrives after the tools on all four sides -- {around}"))

    # --- daily capacity
    cap = getattr(args, "max_per_day", None)
    if cap:
        per_day = defaultdict(dict)
        for t in tools:
            per_day[t["_d"]].setdefault(t["id"], t)
        for d, lst in sorted(per_day.items()):
            if len(lst) > cap:
                ts = sorted(lst.values(), key=lambda t: t["id"])
                out.append(_conflict("capacity", ts, d, None, None,
                                     f"{len(ts)} move-ins on {d.isoformat()} (limit {cap})"))

    for t in tools:
        t.pop("_box", None)
    order = {"high": 0, "medium": 1, "low": 2}
    out.sort(key=lambda c: (order[c["s"]], c["day"], c["ids"][0]))
    return out


def report_conflicts(conflicts, tools):
    if not conflicts:
        print("  Conflicts   : none found")
        return
    kinds = Counter(c["k"] for c in conflicts)
    parts = ", ".join(f"{n} {k}" for k, n in kinds.most_common())
    warn(f"{len(conflicts)} potential move-in conflict(s): {parts}")
    for c in conflicts[:8]:
        who = ", ".join(c["ids"][:5]) + (" …" if len(c["ids"]) > 5 else "")
        warn(f"    [{c['s']:<6}] {CONFLICT_TITLE[c['k']]:<28} {c['day']}  {who}")
        warn(f"             {c['d']}")
    if len(conflicts) > 8:
        warn(f"    … and {len(conflicts) - 8} more (all listed in the viewer's Conflicts tab;"
             f" --conflicts-csv FILE exports them)")
    n_tools = len({i for c in conflicts for i in c["ids"]})
    warn(f"  {n_tools} tool(s) involved; tune with --conflict-window / --conflict-gap / --max-per-day,"
         f" or --no-conflicts")


def write_conflicts_csv(path, conflicts):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["kind", "severity", "day", "tools", "detail"])
        for c in conflicts:
            w.writerow([c["k"], c["s"], c["day"], "; ".join(c["ids"]), c["d"]])
    print(f"  Conflicts written to {path}")


# --------------------------------------------------------------------------
# storage (laydown) space assignment
# --------------------------------------------------------------------------

STORAGE_LAYER_HINT = re.compile(r"stor|lay ?down|stag|ware|buffer|hold", re.I)


def parse_storage_option(spec):
    """--storage NAME=X0,Y0,X1,Y1[:CAP]  ->  zone dict (drawing coordinates)."""
    s = str(spec).strip()
    name, _, rest = s.partition("=") if "=" in s else ("", "", s)
    rest, _, cap = rest.partition(":")
    try:
        vals = [float(v) for v in re.split(r"[,;\s]+", rest.strip()) if v]
        capn = int(cap) if cap.strip() else None
    except ValueError:
        vals, capn = [], None
    if len(vals) != 4 or (capn is not None and capn <= 0):
        die(f"--storage '{spec}' not understood. Use  --storage NAME=X0,Y0,X1,Y1  (drawing\n"
            "  units, AutoCAD ID shows coordinates), optionally :CAP for a maximum tool count,\n"
            "  e.g. --storage \"Laydown B=1200,300,1800,700:12\"")
    x0, x1 = sorted((vals[0], vals[2]))
    y0, y1 = sorted((vals[1], vals[3]))
    return {"n": name.strip(), "b": (x0, y0, x1, y1), "cap": capn, "src": "option"}


def zones_from_layers(dwg: Drawing, patterns):
    """Closed polylines and block inserts on --storage-layer layers become
    storage zones, named by a text inside them (same layer first)."""
    zones = []
    for layer in sorted(dwg.layers):
        if not layer_wanted(layer, patterns, []):
            continue
        boxes = []
        for poly in dwg.layers[layer]["geom"].polys:
            if len(poly) < 4:
                continue
            (fx, fy), (lx, ly) = poly[0], poly[-1]
            if math.hypot(fx - lx, fy - ly) > dwg.size * 1e-4:
                continue                      # open polyline: not an enclosure
            xs = [p[0] for p in poly]; ys = [p[1] for p in poly]
            boxes.append((min(xs), min(ys), max(xs), max(ys)))
        for ins in dwg.inserts:
            if ins["layer"] == layer and ins["bbox"]:
                boxes.append(tuple(ins["bbox"]))
        boxes.sort(key=lambda b: (-b[3], b[0]))           # north to south, west to east
        same = [t for t in dwg.texts if t["layer"] == layer]
        for b in boxes:
            if (b[2] - b[0]) <= 0 or (b[3] - b[1]) <= 0:
                continue
            inside = lambda t: b[0] <= t["x"] <= b[2] and b[1] <= t["y"] <= b[3]
            name = ""
            cands = [t for t in same if inside(t)] or \
                    [t for t in dwg.texts if inside(t) and STORAGE_LAYER_HINT.search(t["str"])]
            if cands:
                name = max(cands, key=lambda t: t["h"])["str"].strip()
            zones.append({"n": name, "b": b, "cap": None, "src": layer})
        if not boxes:
            warn(f"--storage-layer '{layer}': no closed polyline or block insert on it "
                 f"(a storage zone must be a closed outline or a block)")
    for pat in patterns:
        if not any(layer_wanted(l, [pat], []) for l in dwg.layers):
            warn(f"--storage-layer '{pat}' matches no layer in the drawing (see --inspect)")
    return zones


def parse_pair(spec, flag):
    try:
        vals = [float(v) for v in re.split(r"[,;x\s]+", str(spec).strip()) if v]
    except ValueError:
        vals = []
    if len(vals) != 2 or min(vals) <= 0:
        die(f"{flag} '{spec}' not understood: give width,height in drawing units, e.g. {flag} 40,25")
    return vals[0], vals[1]


def assign_storage(tools, zones, args, typical):
    """Give every tool that arrives before its move-in day a storage slot.

    Each zone is cut into a grid of slots (--storage-cell, default from the
    stored tools' footprints). A tool occupies ceil(w/cw) x ceil(h/ch) slots
    from its arrival day until its move-in day; slots are reused afterwards.
    Tools are placed in arrival order, longest stays first, into the free
    zone nearest their final position (a --storage-col zone first).
    Returns (storage payload, overflow conflicts) and decorates the tools.
    """
    need = [t for t in tools if t.get("_arrive") and t["_arrive"] < t["_d"]]
    seen, uniq = set(), []
    for t in sorted(need, key=lambda t: (t["_arrive"], -(t["_d"] - t["_arrive"]).days, t["id"])):
        if t["id"] in seen:
            continue                          # one ID drawn several times: one crate
        seen.add(t["id"]); uniq.append(t)

    if getattr(args, "storage_cell", None):
        cw, ch = parse_pair(args.storage_cell, "--storage-cell")
    else:
        pool = uniq or tools
        ws = sorted(t["w"] for t in pool); hs = sorted(t["h"] for t in pool)
        q = lambda a: a[min(len(a) - 1, int(len(a) * 0.9))]
        cw, ch = max(q(ws), 1e-6) * 1.15, max(q(hs), 1e-6) * 1.15

    Z = []
    for z in zones:
        x0, y0, x1, y1 = z["b"]
        cols, rows = int((x1 - x0) // cw), int((y1 - y0) // ch)
        cap = cols * rows if z["cap"] is None else min(z["cap"], cols * rows)
        Z.append({"n": z["n"], "x": x0, "y": -y1, "w": x1 - x0, "h": y1 - y0,
                  "cols": cols, "rows": rows, "cap": cap, "src": z["src"],
                  "_cells": defaultdict(list), "_iv": [], "_n": 0, "_peak": (0, None)})
        if cols * rows == 0:
            warn(f"storage zone '{z['n']}' ({x1 - x0:,.0f} x {y1 - y0:,.0f}) is smaller than one "
                 f"slot ({cw:,.0f} x {ch:,.0f}); --storage-cell W,H sets the slot size")

    def occupancy(z, a, m):
        """Peak number of tools in zone z during [a, m)."""
        ev = []
        for i0, i1 in z["_iv"]:
            if i0 < m and a < i1:
                ev.append((max(i0, a), 1)); ev.append((min(i1, m), -1))
        ev.sort()
        cur = peak = 0
        for _d, s in ev:
            cur += s; peak = max(peak, cur)
        return peak

    def place(z, t, a, m):
        cs, rs = max(1, math.ceil(t["w"] / cw - 1e-9)), max(1, math.ceil(t["h"] / ch - 1e-9))
        if cs > z["cols"] or rs > z["rows"] or occupancy(z, a, m) >= z["cap"]:
            return False
        for r in range(z["rows"] - rs + 1):
            for c in range(z["cols"] - cs + 1):
                ok = True
                for rr in range(r, r + rs):
                    for cc in range(c, c + cs):
                        if any(i0 < m and a < i1 for i0, i1 in z["_cells"][(cc, rr)]):
                            ok = False; break
                    if not ok:
                        break
                if not ok:
                    continue
                for rr in range(r, r + rs):
                    for cc in range(c, c + cs):
                        z["_cells"][(cc, rr)].append((a, m))
                z["_iv"].append((a, m)); z["_n"] += 1
                t["sz"] = Z.index(z); t["sl"] = r * z["cols"] + c + 1
                t["sx"] = Prec.r(z["x"] + c * cw + (cs * cw - t["w"]) / 2)
                t["sy"] = Prec.r(z["y"] + r * ch + (rs * ch - t["h"]) / 2)
                return True
        return False

    zname = {norm(z["n"]): i for i, z in enumerate(Z) if z["n"]}
    overflow, wrong_zone = [], []
    for t in uniq:
        a, m = t["_arrive"], t["_d"]
        cx, cy = _centre(t)
        order = sorted(range(len(Z)), key=lambda i: math.hypot(Z[i]["x"] + Z[i]["w"] / 2 - cx,
                                                               Z[i]["y"] + Z[i]["h"] / 2 - cy))
        want = norm(t.get("_store", ""))
        if want and want in zname:
            order.remove(zname[want]); order.insert(0, zname[want])
        elif want:
            wrong_zone.append(f"{t['id']} -> '{t['_store']}'")
        placed = None
        for i in order:
            if place(Z[i], t, a, m):
                placed = i
                break
        if placed is None:
            overflow.append(_conflict("storage", [t], a, cx, cy,
                                      f"no free storage slot from {a.isoformat()} to {m.isoformat()}"
                                      f" ({(m - a).days} days) in any zone"))
        elif want and want in zname and placed != zname[want]:
            wrong_zone.append(f"{t['id']} wanted '{t['_store']}' (full), got '{Z[placed]['n']}'")
        t["ad"] = a.isoformat()
        del t["_arrive"]
    for t in tools:
        if "ad" not in t and t.get("_arrive") and t["id"] in seen:
            src = next(u for u in uniq if u["id"] == t["id"])   # same ID, other placement
            for k in ("ad", "sz", "sl", "sx", "sy"):
                if k in src:
                    t[k] = src[k]
        t.pop("_arrive", None); t.pop("_store", None)

    # peak occupancy per zone
    for z in Z:
        ev = sorted([(a, 1) for a, _m in z["_iv"]] + [(m, -1) for _a, m in z["_iv"]])
        cur = 0
        for d, s in ev:
            cur += s
            if cur > z["_peak"][0]:
                z["_peak"] = (cur, d)
    print(f"  Storage     : {len(Z)} zone(s), slot {cw:,.1f} x {ch:,.1f} drawing units; "
          f"{len(uniq)} tool(s) need storage, {len(uniq) - len(overflow)} placed"
          + (f", {len(overflow)} WITHOUT a slot" if overflow else ""))
    for z in Z:
        pk, pd = z["_peak"]
        capnote = "" if z["cap"] == z["cols"] * z["rows"] else f" (max {z['cap']} tools)"
        print(f"    {z['n'] or '(unnamed)':<24} {z['cols']}x{z['rows']} slots{capnote}"
              f" · {z['_n']} tool(s) · peak {pk}" + (f" on {pd.isoformat()}" if pd else ""))
    if wrong_zone:
        warn(f"{len(wrong_zone)} storage wish(es) from the schedule could not be honoured: "
             f"{'; '.join(wrong_zone[:5])}{' …' if len(wrong_zone) > 5 else ''}")
    payload = {"zones": [{k: (Prec.r(v) if isinstance(v, float) else v) for k, v in z.items()
                          if not k.startswith("_")} for z in Z],
               "cw": Prec.r(cw), "ch": Prec.r(ch)}
    return payload, overflow


def write_storage_csv(path, tools, storage):
    zones = storage["zones"]
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["tool", "arrives", "moves_in", "days_in_storage", "zone", "slot"])
        seen = set()
        for t in sorted(tools, key=lambda t: (t.get("ad", ""), t["id"])):
            if "ad" not in t or t["id"] in seen:
                continue
            seen.add(t["id"])
            days = (dt.date.fromisoformat(t["day"]) - dt.date.fromisoformat(t["ad"])).days
            z = zones[t["sz"]]["n"] if "sz" in t else "NO SPACE"
            w.writerow([t["id"], t["ad"], t["day"], days, z, t.get("sl", "")])
    print(f"  Storage plan written to {path}")


def build_payload(dwg: Drawing, rows, strategy, args):
    """Everything the viewer needs, as one JSON-able dict."""
    m = match_rows(dwg, rows, strategy, args)
    tools, symbols, typical = build_tools(m["matched"], m["by_type"])
    storage, conflicts = None, []
    if args.storage_zones:
        storage, conflicts = assign_storage(tools, args.storage_zones, args, typical)
    conflicts += find_conflicts(tools, typical, args)
    sev = {"high": 0, "medium": 1, "low": 2}
    conflicts.sort(key=lambda c: (sev[c["s"]], c["day"], c["ids"][0]))
    for i, c in enumerate(conflicts):
        for t in tools:
            if t["id"] in c["ids"]:
                t.setdefault("cf", []).append(i)
    for t in tools:
        for k in [k for k in t if k.startswith("_")]:
            del t[k]
    layers, pieces, use_items, text_refs, bsyms, inlined = collect_background(
        dwg, m["used_inserts"], m["used_text_keys"], m["kind"], args)
    cap_texts(layers, text_refs, args.max_texts)
    bounds = compute_bounds(tools, typical, pieces, use_items)
    fine_unit = max(dwg.size / 500.0, 1e-9)
    layers = tile_layers(layers, pieces, use_items, bounds, bsyms, fine_unit)
    if inlined:
        print(f"  {inlined:,} rarely-placed block(s) inlined into their layers; "
              f"{len(bsyms):,} shared symbol(s) for the rest")
    if dwg.reduced_blocks:
        print(f"  block detail '{BlockDetail.mode}': {dwg.reduced_blocks:,} block definition(s) "
              f"reduced to their footprint outline (keep everything with --block-detail full)")

    fields, color_by = field_list(rows, m["by_type"], args)
    payload = {
        "title": args.title or "Fab Move-In Simulator",
        "subtitle": f"{os.path.basename(args.dxf)} + {os.path.basename(args.schedule)}"
                    f"  ·  matched on {m['label']}"
                    + (f'  ·  slots assigned by "{rows[0]["type_label"]}"'
                       if m["by_type"] and rows else ""),
        "byType": m["by_type"],
        "fields": fields,
        "colorBy": color_by,
        "layers": layers,
        "layerTools": m["layer_tools"],
        "fineUnit": round(fine_unit, 6),
        "syms": symbols,
        "bsyms": bsyms,
        "tools": tools,
        "days": day_axis(tools, args),
        "bounds": [Prec.r(v) for v in bounds],
        "unmatched": m["unmatched"][:1000],
        "unmatchedTotal": len(m["unmatched"]),
        "conflicts": conflicts,
        "storage": storage,
        "arrivalLabel": rows[0].get("arrive_label", "") if rows else "",
        "generated": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    if args.max_mb:
        shrink_to_budget(payload, args.max_mb)
    n_ids = len({t["id"] for t in tools})
    return payload, n_ids, m["unmatched"], m["label"]


def _layer_stats(L):
    """(geometry bytes, block placements) of one layer across its tiles."""
    geo = sum(len(T[2]) + len(T[3]) for T in L.get("tiles", ()))
    uses = sum(len(T[4]) + len(T[5]) for T in L.get("tiles", ()))
    return geo, uses


def _layer_weight(L):
    """Approximate JSON bytes of one layer's background content, including
    its share of the shared block symbols it references."""
    geo, uses = _layer_stats(L)
    return geo + 45 * uses + 40 * len(L.get("t", ())) + int(L.get("sw", 0))


# layer names that usually carry building services / annotation, not layout
_IRRELEVANT = re.compile(
    r"pip|duct|hvac|mech|elec|cabl|tray|light|lamp|power|sprink|fire|plumb|drain|"
    r"exhaust|gas|chem|vacu|hatch|dimen|(^|[^a-z])dim($|[^a-z])|annot|note|leader|"
    r"ceil|roof|defpoint|furn|pcw|upw|cda|scrub|conduit|bus|util", re.I)


def report_layers(payload):
    """--layers: what each background layer costs, and what to drop."""
    Ls = sorted(payload["layers"], key=lambda L: -_layer_weight(L))
    total = sum(_layer_weight(L) for L in Ls) or 1
    print("\nBackground layers by size in the output (heaviest first):")
    print(f"  {'layer':<32} {'~KB':>8} {'share':>6} {'cum.':>5}   contents")
    cum, suggest = 0, []
    for L in Ls:
        w = _layer_weight(L)
        cum += w
        geo, uses = _layer_stats(L)
        if L["strong"]:
            hint = "walls/outline -- keep"
        elif _IRRELEVANT.search(L["n"]):
            hint = "looks like services/annotation -- probably not needed"
            suggest.append(L["n"])
        else:
            hint = ""
        sw = int(L.get("sw", 0))
        print(f"  {L['n'][:32]:<32} {w / 1024:>8,.0f} {100 * w / total:>5.0f}% {100 * cum / total:>4.0f}%   "
              f"{geo // 1024:,} KB lines · {uses:,} placements"
              + (f" (+{sw // 1024:,} KB of shared symbols)" if sw >= 1024 else "")
              + f" · {len(L['t']):,} texts  {hint}")
    print(f"\n  total background ~{total / 1024:,.0f} KB in {len(Ls)} layers, of which "
          f"{sum(len(b) for b in payload.get('bsyms', ())) / 1024:,.0f} KB are shared block "
          f"symbols (tools are never affected by layer filters)")
    print(f"  tool footprints: {len(payload['tools']):,} tools · {len(payload['syms']):,} distinct "
          f"symbols · {sum(len(d) for d in payload['syms']) / 1024:,.0f} KB"
          + ("   (reduce with --block-detail outline)" if BlockDetail.mode == "full" else ""))
    lt = payload.get("layerTools") or {}
    rows_lt = sorted(((n, v) for n, v in lt.items() if sum(v)), key=lambda kv: (-kv[1][0], -kv[1][1], -kv[1][2]))
    if rows_lt:
        print("\nBlock placements per layer (matching option used for this run):")
        print(f"  {'layer':<32} {'scheduled tools':>15} {'free slots':>11} {'other blocks':>13}")
        print(f"  {'':<32} {'(drawn coloured)':>15} {'(label in':>11} {'(not in':>13}")
        print(f"  {'':<32} {'':>15} {'schedule)':>11} {'schedule)':>13}")
        for n, (m, f, o) in rows_lt[:40]:
            print(f"  {n[:32]:<32} {m:>15,} {f:>11,} {o:>13,}")
        if len(rows_lt) > 40:
            print(f"  … and {len(rows_lt) - 40} more layers")
        print("  Excluding a layer never removes its scheduled tools -- only its free slots")
        print("  and other blocks leave the backdrop. Free slots are drawing placements of")
        print("  a scheduled label that no schedule row consumed (more slots than rows,")
        print("  or tools of another floor / phase).")
    if suggest:
        print("\nLayers whose names suggest they are not needed for a move-in view.")
        print("Paste this into your command to drop them (edit the list as you see fit):")
        print("  " + " ".join(f'--exclude-layer "{n}"' for n in suggest))
    print("\nOr keep only what you name:  --only-layer WALLS --only-layer \"A-GRID*\"  (wildcards ok)")
    print("Preview a layer's contribution first: untick it under View ▾ in the HTML;")
    print("the menu shows each layer's size.\n")


def shrink_to_budget(payload, max_mb):
    """Drop the heaviest background layers (never the tools) until the
    payload fits under max_mb. 'strong' layers (walls/outline) go last."""
    budget = int(max_mb * 1024 * 1024) - 90_000       # headroom for the viewer itself
    cur = len(json.dumps(payload, separators=(",", ":")))
    if cur <= budget:
        return
    warn(f"payload is {cur / 1e6:,.0f} MB, over --max-mb {max_mb:g} -- "
         f"dropping heaviest background layers (tools are always kept):")
    layers = payload["layers"]
    order = sorted(range(len(layers)), key=lambda i: (layers[i]["strong"], -_layer_weight(layers[i])))
    for i in order:
        if cur <= budget:
            break
        L = layers[i]
        w = _layer_weight(L)
        if w < 10_000:
            continue
        warn(f"    dropped layer {L['n']}  (~{w / 1024:,.0f} KB)")
        L["tiles"], L["t"] = [], []
        cur -= w
    payload["layers"] = [L for L in layers if L["tiles"] or L["t"]]
    # prune background symbols nothing references any more, remapping indexes
    used = sorted({u[0] for L in payload["layers"] for T in L["tiles"] for u in T[4] + T[5]})
    remap = {old: new for new, old in enumerate(used)}
    for L in payload["layers"]:
        for T in L["tiles"]:
            for u in T[4] + T[5]:
                u[0] = remap[u[0]]
    payload["bsyms"] = [payload["bsyms"][old] for old in used]
    cur = len(json.dumps(payload, separators=(",", ":")))
    warn(f"  background trimmed; payload is now ~{cur / 1e6:,.1f} MB. "
         f"Dropped layers stay off the map entirely.")


# --------------------------------------------------------------------------
# inspect mode
# --------------------------------------------------------------------------

def insert_extents(dwg: Drawing):
    """2%..98% spread of block insert centres, or None: where the fab is."""
    xs, ys = [], []
    for ins in dwg.inserts:
        x, y = hit_centre(ins)
        xs.append(x); ys.append(y)
    if len(xs) < 2:
        return None
    xs.sort(); ys.sort()
    lo, hi = int(len(xs) * 0.02), max(int(len(xs) * 0.98) - 1, 0)
    return xs[lo], ys[lo], xs[hi], ys[hi]


def do_inspect(dwg: Drawing, sch: Schedule | None, full: bool = False):
    lim = None if full else 40      # None = no cap (--inspect-full)
    print("\n" + "=" * 68)
    print("DXF CONTENTS" + ("  (full listing)" if full else ""))
    print("=" * 68)

    print(f"\nLayers ({len(dwg.layers)}):")
    for name, d in sorted(dwg.layers.items(), key=lambda kv: -kv[1]["count"])[:lim]:
        print(f"  {name:<28} {d['count']:>7,} entities")
    stor = [n for n in dwg.layers if STORAGE_LAYER_HINT.search(n)]
    if stor:
        print("  -> layers that look like storage / laydown areas: " + ", ".join(f"'{n}'" for n in stor[:6]))
        print("     use  --storage-layer NAME  to assign waiting tools a slot inside their outlines")

    blocks = Counter(i["name"] for i in dwg.inserts)
    print(f"\nBlock names ({len(blocks)} distinct, {len(dwg.inserts):,} inserts):")
    ext = insert_extents(dwg)
    if ext:
        x0, y0, x1, y1 = ext
        print(f"  most inserts lie in X {x0:,.0f} .. {x1:,.0f}, Y {y0:,.0f} .. {y1:,.0f}"
              f"  (mid X {(x0 + x1) / 2:,.0f}, mid Y {(y0 + y1) / 2:,.0f})")
        print("  -> with --type-col, choose which copy of a repeated label a row takes:")
        print("     --tool-layer NAME   --prefer north|south|east|west   --prefer X0,Y0,X1,Y1")
    for name, n in blocks.most_common(None if full else 30):
        print(f"  {name:<28} {n:>7,} inserts")

    tags = dwg.attrib_tags()
    if tags:
        print(f"\nBlock attribute tags ({len(tags)}):")
        for tag, n in tags.most_common(None if full else 20):
            samples = []
            for ins in dwg.inserts:
                v = ins["attribs"].get(tag)
                if v and v not in samples:
                    samples.append(v)
                if len(samples) >= 5:
                    break
            print(f"  {tag:<28} {n:>7,} blocks   e.g. {', '.join(samples)}")
        print("\n  -> to match on one of these:  --match attrib:<TAG>")
    else:
        print("\nBlock attribute tags: none found")
        print("  (tools are probably identified by block name or by a text label)")

    if dwg.texts:
        show_n = len(dwg.texts) if full else 15
        print(f"\nText labels ({len(dwg.texts):,})" + ("" if full else ", first 15") + ":")
        for t in dwg.texts[:show_n]:
            print(f"  {t['str'][:60]}")

    bt = dwg.index_by_blocktext()
    if bt:
        print(f"\nText inside block definitions ({len(bt)} distinct strings):")
        print("  (stored once per block, shown on every insert -- FIND/text search")
        print("   sees these only once, but they identify all placements)")
        ranked = sorted(bt.items(), key=lambda kv: -len(kv[1]))
        cut = len(ranked) if full else 30
        for k, lst in ranked[:cut]:
            print(f"  {k[:40]:<40} on {len(lst):>6,} block insert(s)")
        if len(ranked) > cut:
            print(f"  … and {len(ranked) - cut} more distinct strings"
                  f"  (use --inspect-full to list all)")

    if sch is None:
        print()
        return

    print("\n" + "=" * 68)
    print("SCHEDULE CONTENTS")
    print("=" * 68)
    if sch.sheet_names:
        print(f"\nSheets: {', '.join(sch.sheet_names)}   (using '{sch.sheet}')")
    print(f"\nColumns ({len(sch.headers)}), with the first few values:")
    for i, h in enumerate(sch.headers):
        vals = []
        for r in sch.rows[:6]:
            v = r[i] if i < len(r) else None
            if v is not None and str(v).strip():
                vals.append(str(v)[:22])
        print(f"  [{i+1}] {h:<26} {', '.join(vals)}")

    print("\nHow well each match option would do:")
    ordered, seen = [], set()
    for r in sch.rows:
        if r and r[0] is not None:
            k = mkey(r[0])
            if k and k not in seen:
                seen.add(k)
                ordered.append(k)

    def match_line(name, idx):
        found = [k for k in ordered if k in idx]
        ex = f"   e.g. {', '.join(found[:4])}" if found else ""
        print(f"  {name:<26} {len(found):>6,} of {len(ordered):,}{ex}")

    for tag in dwg.attrib_tags():
        match_line(f"attrib:{tag}", dwg.index_by_attrib(tag))
    match_line("blockname", dwg.index_by_blockname())
    match_line("blocktext", dwg.index_by_blocktext())
    match_line("text", dwg.index_by_text())
    print("\n(the counts above test column 1 of the schedule; the e.g. values are")
    print(" schedule values that were actually found in the drawing)\n")

    # near misses: schedule values that differ from a drawing label only by
    # punctuation or spacing (DT-112 vs DT_112) -- solvable with --loose
    if not Match.loose:
        allkeys = set(bt) | set(dwg.index_by_blockname()) | set(dwg.index_by_text())
        for tag in dwg.attrib_tags():
            allkeys |= set(dwg.index_by_attrib(tag))
        cmap = {}
        for k in allkeys:
            cmap.setdefault(compact(k), k)
        pairs = []
        for v in ordered:
            if v in allkeys:
                continue
            c = compact(v)
            if c and c in cmap:
                pairs.append((v, cmap[c]))
        if pairs:
            print(f"Near misses: {len(pairs)} first-column value(s) differ from a drawing")
            print("label only by punctuation or spacing:")
            for v, k in pairs[:10]:
                print(f"    schedule '{v}'   vs   drawing '{k}'")
            if len(pairs) > 10:
                print(f"    … and {len(pairs) - 10} more")
            print("  -> rerun with --loose to ignore those differences when matching\n")


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------

ID_PATTERNS = [r"^(tool|block|eq(uip(ment)?)?|asset)[ _-]?(id|tag|no|number)$",
               r"tool.*id|block.*id|id.*tool|asset.*tag", r"^id$", r"tool|equip|asset"]
DATE_PATTERNS = [r"move.?in.*date|date.*move.?in", r"move.?in", r"install", r"hook.?up",
                 r"deliver|arriv|ship", r"date", r"schedul|plan"]
NAME_PATTERNS = [r"^(tool[ _-]?)?name$", r"descript"]
# extra filter / colour facets picked up automatically (besides group / area / type)
FACET_PATTERNS = [r"^model|tool.?model|model.?(no|number|name)", r"vendor|supplier|manufacturer|maker|^oem",
                  r"^module|module", r"^owner|^status|^priority|^process|^category"]
GROUP_PATTERNS = [r"phase|stage|wave|batch", r"group|category|type|process"]
AREA_PATTERNS = [r"bay|area|zone|module|room|cleanroom|sector"]
ARRIVAL_PATTERNS = [r"arriv|deliver|ship|eta|receiv|on.?site|dock"]
STORE_PATTERNS = [r"stor|lay.?down|stag|warehouse"]


def build_parser():
    ap = argparse.ArgumentParser(
        prog="fab_movein.py",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description="Build a day-by-day tool move-in animation from a DXF layout "
                    "and a move-in schedule. Everything runs locally.",
        epilog="""examples:
  python fab_movein.py --dxf fab.dxf --schedule movein.xlsx
  python fab_movein.py --dxf fab.dxf --schedule movein.xlsx --inspect
  python fab_movein.py --dxf fab.dxf --schedule movein.xlsx --type-col Block_ID --layers
  python fab_movein.py --dxf fab.dxf --schedule movein.xlsx \\
      --match attrib:TOOL_ID --id-col "Tool ID" --date-col "Move In Date" \\
      --group-col Phase --area-col Bay -o q4-movein.html --open
""")
    g = ap.add_argument_group("input")
    g.add_argument("--dxf", required=True, help="fab layout DXF file")
    g.add_argument("--schedule", required=True, help="move-in schedule (.xlsx/.xlsm/.csv)")
    g.add_argument("--sheet", help="worksheet name (default: the biggest sheet)")

    g = ap.add_argument_group("schedule columns (all auto-detected when omitted)")
    g.add_argument("--id-col", help="column holding the tool / block ID")
    g.add_argument("--date-col", help="column holding the move-in date")
    g.add_argument("--name-col", help="column holding a descriptive tool name")
    g.add_argument("--group-col", help="column to colour by (e.g. Phase)")
    g.add_argument("--area-col", help="column holding the bay / area")
    g.add_argument("--facet-col", action="append", default=[], metavar="NAME",
                   help="extra column to filter and colour by in the viewer (repeatable). Model, "
                        "vendor / supplier and module columns are picked up automatically")
    g.add_argument("--color-by", metavar="NAME",
                   help="column the viewer colours tools by at start (default: a Module column "
                        "if there is one, else the group column, else the area column)")
    g.add_argument("--type-col", metavar="NAME",
                   help="column holding the tool TYPE, for drawings that label types rather "
                        "than individual tools (several tools all marked e.g. DT_112). Each "
                        "schedule row is assigned one free placement of its type, in date order")
    g.add_argument("--date-format", default="auto", choices=["auto", "dmy", "mdy", "ymd"],
                   help="how to read ambiguous numeric dates")

    g = ap.add_argument_group("matching")
    g.add_argument("--match", default="auto",
                   help="how to find tools in the DXF: auto | attrib:TAG | blockname | "
                        "blocktext | text")
    g.add_argument("--loose", action="store_true",
                   help="ignore punctuation and spacing when matching IDs "
                        "(DT-112 matches DT_112 and DT 112)")
    g.add_argument("--tool-layer", action="append", default=[], metavar="NAME",
                   help="type mode: drawing layer(s) holding the scheduled tools (repeatable; "
                        "* wildcards allowed). When a label is placed several times, copies "
                        "on these layers are taken first, e.g. --tool-layer \"0_ENG 3K tool\"")
    g.add_argument("--prefer", metavar="WHERE",
                   help="type mode: which placement a row takes when its label appears "
                        "several times in the drawing -- north | south | east | west "
                        "(that side first; north = higher Y), or a rectangle in drawing "
                        "units X0,Y0,X1,Y1 (placements inside it first). Use it when the "
                        "schedule covers one part of the fab and the same labels also "
                        "exist elsewhere")

    g = ap.add_argument_group("move-in conflicts (detected and marked by default)")
    g.add_argument("--conflict-window", type=int, default=3, metavar="DAYS",
                   help="tools moving in within this many days of each other AND closer than "
                        "--conflict-gap are flagged as crowded (default 3; 0 = same day only)")
    g.add_argument("--conflict-gap", type=float, default=None, metavar="UNITS",
                   help="edge-to-edge distance (drawing units) below which two tools count as "
                        "neighbours for crowding (default: one typical tool size; 0 disables)")
    g.add_argument("--max-per-day", type=int, default=None, metavar="N",
                   help="flag days with more than N move-ins (rigging / crew capacity)")
    g.add_argument("--no-conflicts", action="store_true",
                   help="skip conflict detection (overlaps, crowding, boxed-in tools, capacity)")
    g.add_argument("--conflicts-csv", metavar="FILE", help="also write the conflict list as CSV")

    g = ap.add_argument_group("storage (laydown) space")
    g.add_argument("--storage", action="append", default=[], metavar="NAME=X0,Y0,X1,Y1[:CAP]",
                   help="a storage zone as a rectangle in drawing units, optionally with a "
                        "maximum tool count (repeatable), e.g. --storage \"Laydown B=1200,300,1800,700:12\"")
    g.add_argument("--storage-layer", action="append", default=[], metavar="NAME",
                   help="drawing layer(s) whose closed outlines / blocks are storage zones "
                        "(repeatable; * wildcards). A text inside the outline names the zone")
    g.add_argument("--arrival-col", metavar="NAME",
                   help="column holding the delivery / arrival date (auto-detected). Tools that "
                        "arrive before their move-in day get a storage slot for the wait")
    g.add_argument("--storage-lead", type=int, default=None, metavar="DAYS",
                   help="without an arrival column: every tool arrives this many days before "
                        "its move-in (default 7 when storage zones are given)")
    g.add_argument("--storage-cell", metavar="W,H",
                   help="storage slot size in drawing units (default: from the tools' footprints)")
    g.add_argument("--storage-col", metavar="NAME",
                   help="column naming the storage zone a tool should use (honoured when free)")
    g.add_argument("--storage-csv", metavar="FILE", help="write the storage plan (tool, zone, slot, dates) as CSV")

    g = ap.add_argument_group("output")
    g.add_argument("-o", "--out", default="fab-move-in-simulation.html", help="output HTML file")
    g.add_argument("--title", help="title shown in the page")
    g.add_argument("--start-date", help="clip the timeline start (YYYY-MM-DD)")
    g.add_argument("--end-date", help="clip the timeline end (YYYY-MM-DD)")
    g.add_argument("--open", action="store_true", dest="open_",
                   help="open the result in your browser when done")

    g = ap.add_argument_group("size and detail")
    g.add_argument("--exclude-layer", action="append", default=[], metavar="NAME",
                   help="drop a drawing layer (repeatable; * wildcards allowed), e.g. "
                        "--exclude-layer PIPING")
    g.add_argument("--only-layer", action="append", default=[], metavar="NAME",
                   help="keep only these drawing layers (repeatable; * wildcards allowed)")
    g.add_argument("--block-detail", default="auto", choices=["auto", "full", "outline", "box"],
                   help="how much of each block's inner geometry to keep: auto = outline + "
                        "major lines (default), full = everything, outline = convex-hull "
                        "footprint, box = bounding rectangle. Whole-floor blocks are always full")
    g.add_argument("--max-mb", type=float, default=None, metavar="MB",
                   help="hard cap on output size: drop the heaviest background "
                        "layers (never the tools) until the file fits, e.g. --max-mb 40")
    g.add_argument("--simplify", type=float, default=None, metavar="UNITS",
                   help="polyline simplification tolerance in drawing units "
                        "(default: auto ≈ drawing size / 6000; 0 disables)")
    g.add_argument("--max-texts", type=int, default=4000, metavar="N",
                   help="cap on background text labels, largest kept (default 4000)")
    g.add_argument("--precision", type=int, default=None,
                   help="coordinate decimals (default: auto from drawing size)")

    g = ap.add_argument_group("diagnostics (print a report and exit)")
    g.add_argument("--inspect", action="store_true",
                   help="what's in the DXF and the spreadsheet, and what would match")
    g.add_argument("--inspect-full", action="store_true",
                   help="like --inspect, but list everything without truncation")
    g.add_argument("--layers", action="store_true",
                   help="every background layer by its size in the output, what it holds, "
                        "and a suggested --exclude-layer list")
    g.add_argument("--trace", action="append", default=[], metavar="LABEL",
                   help="explain what happened to one schedule label / tool ID: where it "
                        "is in the drawing, which placement it got, or why not (repeatable)")
    ap.add_argument("--version", action="version", version=f"%(prog)s {__version__}")
    return ap


def resolve_columns(sch: Schedule, args, drawing_keys):
    """Work out which spreadsheet columns hold what. Returns a dict with the
    column indexes (-1 = none), the date style and a flag telling whether a
    --type-col hint was already printed."""
    i_id, how_id = sch.resolve(args.id_col, ID_PATTERNS, "ID column", "--id-col")
    i_date, how_date = sch.resolve(args.date_col, DATE_PATTERNS, "Date column", "--date-col")
    i_type = -1
    if args.type_col:
        i_type, _ = sch.resolve(args.type_col, [], "Type column", "--type-col")

    # a header-matched ID column that finds nothing is worse than no guess at all
    # (unless we match by type -- then the ID column need not appear in the drawing)
    i_id_hdr = -1
    if how_id == "header" and i_type < 0:
        keys = {norm(v) for v in sch.column(i_id)}
        if not any(k in drawing_keys for k in keys):
            i_id_hdr, i_id, how_id = i_id, -1, ""
    hinted_type = False
    if i_id < 0:
        excl = {i for i in (i_date, i_type) if i >= 0}
        i_id, _hits = sch.detect_id_column(drawing_keys, exclude=excl)
        how_id = "content" if i_id >= 0 else ""
        if i_id >= 0 and i_id_hdr >= 0:
            # the column that matches the drawing repeats its values heavily:
            # that's a tool TYPE column, not per-tool IDs -- don't dedupe on it
            vals = [norm(v) for v in sch.column(i_id) if norm(v)]
            if vals and len(set(vals)) < 0.8 * len(vals):
                warn(f'column "{sch.headers[i_id]}" matches the drawing, but its values repeat --')
                warn('  it looks like a tool TYPE column (several tools per label).')
                warn(f'  Keeping "{sch.headers[i_id_hdr]}" as the ID column. To place every tool')
                warn(f'  on a free slot of its type, rerun with:  --type-col "{sch.headers[i_id]}"')
                hinted_type = True
                i_id, how_id = i_id_hdr, "header"
        if i_id < 0:
            die("could not work out which column holds the tool / block IDs.\n"
                f"  Columns found: {', '.join(sch.headers)}\n"
                "  None of them contained values that appear in the drawing.\n"
                "  Run with --inspect to compare the two, then pass --id-col \"<column name>\".")

    style = args.date_format
    if style == "auto" and i_date >= 0:
        style = detect_date_style([v for v in sch.column(i_date) if isinstance(v, str)])
    elif style == "auto":
        style = "mdy"

    if i_date >= 0 and how_date == "header":
        vals = sch.column(i_date)[:40]
        if vals and sum(1 for v in vals if coerce_date(v, style) is not None) / len(vals) < 0.5:
            i_date, how_date = -1, ""      # header said "date" but the values aren't dates
    if i_date < 0:
        i_date = sch.detect_date_column(style, exclude={i_id})
        how_date = "content" if i_date >= 0 else ""
        if i_date < 0:
            die("could not work out which column holds the move-in dates.\n"
                f"  Columns found: {', '.join(sch.headers)}\n"
                "  None of them parsed as dates. If they are text like '3.Q 2026',\n"
                "  add a real date column, or pass --date-col \"<column name>\"\n"
                "  together with --date-format dmy|mdy|ymd.")
        if args.date_format == "auto":
            style = detect_date_style([v for v in sch.column(i_date) if isinstance(v, str)])

    taken = {i_id, i_date}
    i_arr, how_arr = sch.resolve(args.arrival_col, ARRIVAL_PATTERNS, "Arrival column", "--arrival-col")
    if i_arr in taken:
        if how_arr == "given":
            die("--arrival-col cannot be the ID or move-in date column.")
        i_arr = -1
    if i_arr >= 0:
        vals = sch.column(i_arr)[:40]
        if not vals or sum(1 for v in vals if coerce_date(v, style) is not None) / len(vals) < 0.5:
            if how_arr == "given":
                die(f'--arrival-col "{sch.headers[i_arr]}" does not hold readable dates.')
            i_arr = -1
    i_store = -1
    if args.storage_zones:
        i_store, _ = sch.resolve(args.storage_col, STORE_PATTERNS, "Storage column", "--storage-col")
        if i_store in taken or i_store == i_arr:
            i_store = -1
    taken |= {i for i in (i_arr, i_store) if i >= 0}
    i_name, _ = sch.resolve(args.name_col, NAME_PATTERNS, "Name column", "--name-col")
    i_group, _ = sch.resolve(args.group_col, GROUP_PATTERNS, "Group column", "--group-col")
    i_area, _ = sch.resolve(args.area_col, AREA_PATTERNS, "Area column", "--area-col")
    if i_name in taken:
        i_name = -1
    if i_group in taken:
        i_group = -1
    if i_area in taken or (i_area >= 0 and i_area == i_group):
        i_area = -1
    # facets: explicit --facet-col first, then well-known columns (model, vendor,
    # module, ...) that repeat their values, i.e. are worth filtering by
    used = {i for i in (i_id, i_date, i_name, i_group, i_area, i_type, i_arr, i_store) if i >= 0}
    facets = []
    for spec in args.facet_col:
        j, _ = sch.resolve(spec, [], "Facet column", "--facet-col")
        if j in (i_id, i_date):
            die(f"--facet-col '{spec}' is the ID / date column; pick another column.")
        if j not in facets:
            facets.append(j)
    for pat in FACET_PATTERNS:
        for j, h in enumerate(sch.headers):
            if j in used or j in facets or not re.search(pat, h, re.I):
                continue
            vals = [norm(v) for v in sch.column(j) if norm(v)]
            if len(vals) >= 2 and 2 <= len(set(vals)) < len(vals):
                facets.append(j)
    if i_type >= 0:
        if i_type == i_date:
            die("--type-col cannot be the date column.")
        # --type-col may equal the ID column: the schedule then has no unique
        # per-tool IDs, and repeated labels are numbered DT_112 (2), (3), ...
        if i_group < 0 and i_type != i_id:
            i_group = i_type       # colour by type unless another grouping was chosen

    note = {"given": "", "header": "", "content": "  (auto-detected from the values)"}
    print(f"  ID column   : {sch.headers[i_id]}{note.get(how_id, '')}")
    print(f"  Date column : {sch.headers[i_date]}{note.get(how_date, '')}")
    for lbl, i in (("Name", i_name), ("Group", i_group), ("Area", i_area),
                   ("Arrival", i_arr), ("Storage", i_store)):
        if i >= 0:
            print(f"  {lbl:<12}: {sch.headers[i]}")
    if facets:
        print(f"  Filters     : {', '.join(sch.headers[j] for j in facets)}  (plus ID"
              f"{', type' if i_type >= 0 else ''}{', group' if i_group >= 0 else ''}"
              f"{', area' if i_area >= 0 else ''} -- filter and colour by any of them in the viewer)")
    if i_type >= 0:
        print(f"  Type        : {sch.headers[i_type]}  "
              f"(each row gets one free placement of its type, in date order)")
    return {"id": i_id, "date": i_date, "name": i_name, "group": i_group, "area": i_area,
            "type": i_type, "arrive": i_arr, "store": i_store, "facets": facets,
            "style": style, "hinted_type": hinted_type}


def build_rows(sch: Schedule, cols):
    """Usable schedule rows: an ID, a readable date, and the optional columns."""
    cell = lambda r, i: (r[i] if 0 <= i < len(r) else None)
    i_id, i_date, i_type = cols["id"], cols["date"], cols["type"]
    rows, bad_dates, seen, dup_rows = [], [], {}, 0
    for r in sch.rows:
        raw = cell(r, i_id)
        if raw is None or not str(raw).strip():
            continue
        tid = str(raw).strip()
        key = mkey(tid)
        day = coerce_date(cell(r, i_date), cols["style"])
        if day is None:
            bad_dates.append(tid)
            continue
        n_seen = seen.get(key, 0)
        if n_seen:
            if i_type < 0:
                dup_rows += 1
                continue
            tid = f"{tid} ({n_seen + 1})"   # same label, several physical tools
        seen[key] = n_seen + 1
        txt = lambda i: ("" if i < 0 or cell(r, i) is None else str(cell(r, i)).strip())
        arrive = coerce_date(cell(r, cols["arrive"]), cols["style"]) if cols["arrive"] >= 0 else None
        if arrive is None and cols.get("lead"):
            arrive = day - dt.timedelta(days=cols["lead"])
        rows.append({"id": tid, "key": key, "day": day, "name": txt(cols["name"]),
                     "group": txt(cols["group"]), "area": txt(cols["area"]),
                     "type": txt(i_type), "tkey": mkey(txt(i_type)),
                     "arrive": arrive, "store": txt(cols["store"]),
                     "facets": [txt(j) for j in cols["facets"]],
                     "facet_labels": [sch.headers[j] for j in cols["facets"]],
                     "arrive_label": sch.headers[cols["arrive"]] if cols["arrive"] >= 0 else "",
                     "group_label": sch.headers[cols["group"]] if cols["group"] >= 0 else "",
                     "area_label": sch.headers[cols["area"]] if cols["area"] >= 0 else "",
                     "type_label": sch.headers[i_type] if i_type >= 0 else ""})

    if not rows:
        die("no usable rows: every row was missing an ID or a readable date.\n"
            "  Try --inspect to see the columns, then set --id-col / --date-col.")
    if bad_dates:
        warn(f"{len(bad_dates)} row(s) had an unreadable date and were skipped "
             f"(e.g. {', '.join(bad_dates[:5])})")
    if dup_rows:
        warn(f"{dup_rows} duplicate-ID row(s) ignored (the same ID on several rows).")
        warn("  If those are multiple units of one type, add --type-col so each row"
             " gets its own placement.")
    return rows


def suggest_type_col(sch: Schedule, rows, cols, drawing_keys):
    """Direct IDs barely present in the drawing, but another column matches?
    Then the drawing probably labels tool TYPES, not individual tools."""
    hitrate = sum(1 for r in rows if r["key"] in drawing_keys) / len(rows)
    if hitrate >= 0.2 or cols["hinted_type"]:
        return
    best_j, best_rate = -1, 0.0
    for j in range(len(sch.headers)):
        if j in (cols["id"], cols["date"]):
            continue
        vals = [mkey(v) for v in sch.column(j) if mkey(v)]
        if len(vals) < 2:
            continue
        rate = sum(1 for v in vals if v in drawing_keys) / len(vals)
        if rate > best_rate:
            best_rate, best_j = rate, j
    if best_rate >= 0.5:
        warn(f"only {hitrate:.0%} of the schedule IDs appear anywhere in the drawing,")
        warn(f'  but column "{sch.headers[best_j]}" matches {best_rate:.0%} of its values.')
        warn("  If the drawing labels tool TYPES (several tools sharing one label),")
        warn(f'  rerun with:  --type-col "{sch.headers[best_j]}"')


def main(argv=None):
    args = build_parser().parse_args(argv)

    # Windows consoles default to cp1252/cp437; layer and block names from a DXF
    # may carry characters they cannot encode. Never let a print() crash the run.
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(errors="replace")
        except Exception:
            pass

    Match.loose = bool(args.loose)
    if Match.loose:
        print("Loose matching: punctuation and spacing in IDs are ignored")
    BlockDetail.mode = args.block_detail

    dwg = Drawing(args.dxf, simplify=args.simplify)
    Prec.digits = dwg.autodigits if args.precision is None else max(0, min(6, args.precision))
    sch = Schedule(args.schedule, args.sheet)

    if args.inspect or args.inspect_full:
        do_inspect(dwg, sch, full=args.inspect_full)
        return 0

    # every identifier the drawing knows about, for content-based ID detection
    drawing_keys = (set(dwg.index_by_blockname()) | set(dwg.index_by_text())
                    | set(dwg.index_by_blocktext()))
    for tag in dwg.attrib_tags():
        drawing_keys |= set(dwg.index_by_attrib(tag))

    args.storage_zones = [parse_storage_option(sp) for sp in args.storage]
    if args.storage_layer:
        args.storage_zones += zones_from_layers(dwg, args.storage_layer)
    for i, z in enumerate(args.storage_zones):
        if not z["n"]:
            z["n"] = f"Storage {i + 1}"
    if args.storage_zones:
        for z in args.storage_zones:
            x0, y0, x1, y1 = z["b"]
            print(f"  Storage zone: {z['n']:<24} X {x0:,.0f}..{x1:,.0f}  Y {y0:,.0f}..{y1:,.0f}"
                  + ("" if z["cap"] is None else f"  max {z['cap']} tools")
                  + (f"  (layer '{z['src']}')" if z["src"] != "option" else ""))
    elif args.storage_lead or args.storage_cell or args.storage_col or args.storage_csv:
        warn("--storage-lead / --storage-cell / --storage-col / --storage-csv need storage zones: "
             "add --storage NAME=X0,Y0,X1,Y1 or --storage-layer NAME")

    cols = resolve_columns(sch, args, drawing_keys)
    if args.storage_zones and cols["arrive"] < 0:
        cols["lead"] = 7 if args.storage_lead is None else max(0, args.storage_lead)
        if cols["lead"]:
            print(f"  Arrival     : no arrival column -- every tool arrives {cols['lead']} day(s) "
                  f"before its move-in (--storage-lead DAYS changes this)")
        else:
            warn("--storage-lead 0 and no arrival column: no tool needs storage")
    elif args.storage_lead is not None and cols["arrive"] >= 0:
        cols["lead"] = max(0, args.storage_lead)
        print(f"  Arrival     : rows without a readable arrival date arrive {cols['lead']} day(s) early")
    rows = build_rows(sch, cols)

    if cols["type"] >= 0:
        match_keys = {r["tkey"] for r in rows if r["tkey"]}
        if not match_keys:
            die(f"type column '{sch.headers[cols['type']]}' has no values.")
    else:
        match_keys = {r["key"] for r in rows}
        suggest_type_col(sch, rows, cols, drawing_keys)

    strategy = choose_strategy(dwg, match_keys, args.match)
    print(f"  Matching on : {strategy[0]}")
    policy = placement_policy(args)
    if policy and cols["type"] >= 0:
        print(f"  Placements  : {policy['desc']}")
        for lay in args.tool_layer:
            if not any(layer_wanted(i["layer"], [lay], []) for i in dwg.inserts):
                warn(f"--tool-layer '{lay}' matches no block insert in the drawing "
                     f"(layer names are listed by --inspect)")
    elif policy:
        warn("--tool-layer / --prefer only matter with --type-col: without it every placement "
             "carrying a scheduled ID is drawn, so there is nothing to choose between.")

    payload, n_ids, unmatched, label = build_payload(dwg, rows, strategy, args)

    if args.layers:
        report_layers(payload)
        return 0

    print(f"\nMatched {n_ids:,} of {len(rows):,} scheduled tools to the drawing"
          f" ({len(payload['syms'])} distinct footprints)")
    if unmatched:
        warn(f"{len(unmatched)} not found in the drawing: "
             f"{', '.join(unmatched[:8])}{' …' if len(unmatched) > 8 else ''}")
        if len(unmatched) > len(rows) * 0.5:
            warn("that's more than half -- try --inspect and a different --match option")
    if not args.no_conflicts:
        report_conflicts(payload["conflicts"], payload["tools"])
    if args.conflicts_csv:
        write_conflicts_csv(args.conflicts_csv, payload["conflicts"])
    if args.storage_csv and payload["storage"]:
        write_storage_csv(args.storage_csv, payload["tools"], payload["storage"])

    html = render_html(payload)
    with open(args.out, "w", encoding="utf-8") as fh:
        fh.write(html)
    size = os.path.getsize(args.out) / 1024
    n_uses = sum(_layer_stats(L)[1] for L in payload["layers"])
    print(f"\nWrote {args.out}  ({size:,.0f} KB · {len(payload['days']):,} days · "
          f"{len(payload['layers'])} layers · {n_uses:,} reused block placements)")
    print("Open it in any browser -- it is self-contained and works offline.")
    print("(Double-click the file. Don't serve it through a local web server --")
    print(" very large files can abort there with 'WinError 10053'.)")
    if size > 25 * 1024:
        heavy = sorted(payload["layers"], key=lambda L: -_layer_weight(L))[:6]
        warn("the output is large. Heaviest layers:")
        for L in heavy:
            geo, uses = _layer_stats(L)
            warn(f"    {L['n']:<28} ~{_layer_weight(L) / 1024:>8,.0f} KB  "
                 f"({geo // 1024:,} KB geometry · {uses:,} placements"
                 f" · {int(L.get('sw', 0)) // 1024:,} KB symbols · {len(L.get('t', ())):,} texts)")
        warn("  full table + suggested excludes:  add --layers to this command")
        warn("  drop layers you don't need:  --exclude-layer NAME  (wildcards ok, repeatable)")
        warn("  or cap the size outright:    --max-mb 40   (drops heaviest background layers)")
        warn("  other knobs: --simplify <units> (coarser curves) · --max-texts <n>")

    if args.open_:
        import webbrowser
        webbrowser.open("file://" + os.path.abspath(args.out))
    return 0


def render_html(payload) -> str:
    data = json.dumps(payload, separators=(",", ":"), ensure_ascii=False)
    # '<' only ever occurs inside JSON string values, so this is a safe way to
    # make sure the payload can never terminate the <script> element early.
    data = data.replace("<", "\\u003c")
    return _VIEWER_TEMPLATE.replace("/*__DATA__*/", data)


if __name__ == "__main__":
    sys.exit(main())
